import unittest
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from automacoes.scripts.mogo_cadastro_clientes_weekly import (
    analyze_records,
    build_summary_message,
    compare_snapshots,
    sanitize_customer,
    write_excel,
)


class MogoCadastroClientesWeeklyTests(unittest.TestCase):
    def test_sanitize_customer_keeps_operational_fields_and_drops_sensitive_fields(self):
        raw = {
            "Id": 123,
            "Nome": " Cliente Teste ",
            "Email": "cliente@exemplo.com",
            "TelefoneCelular": "(21) 99999-9999",
            "CNPJ_CPF": "111.444.777-35",
            "Senha": "nao-pode-salvar",
            "TokenFirebase": "nao-pode-salvar",
            "Cartoes": [{"final": "1234"}],
            "ProviderKeyGoogle": "nao-pode-salvar",
        }

        sanitized = sanitize_customer(raw)

        self.assertEqual(sanitized["id"], "123")
        self.assertEqual(sanitized["nome"], "Cliente Teste")
        self.assertEqual(sanitized["email"], "cliente@exemplo.com")
        self.assertEqual(sanitized["telefone_celular"], "(21) 99999-9999")
        self.assertEqual(sanitized["documento"], "111.444.777-35")
        self.assertNotIn("Senha", sanitized)
        self.assertNotIn("TokenFirebase", sanitized)
        self.assertNotIn("Cartoes", sanitized)
        self.assertNotIn("ProviderKeyGoogle", sanitized)

    def test_sanitize_customer_keeps_extra_operational_fields_and_drops_sensitive_unknowns(self):
        raw = {
            "Id": 123,
            "Nome": "Cliente Teste",
            "CodClienteIFood": "ifood-123",
            "QuantidadePessoas": 4,
            "Restricao": "sem gluten",
            "Passaporte": "AB123456",
            "MalgaCustomerId": "nao-pode-salvar",
            "ProviderKeyFacebook": "nao-pode-salvar",
        }

        sanitized = sanitize_customer(raw)

        self.assertEqual(sanitized["cod_cliente_ifood"], "ifood-123")
        self.assertEqual(sanitized["quantidade_pessoas"], "4")
        self.assertEqual(sanitized["restricao"], "sem gluten")
        self.assertEqual(sanitized["passaporte"], "AB123456")
        self.assertNotIn("malga_customer_id", sanitized)
        self.assertNotIn("provider_key_facebook", sanitized)

    def test_analyze_records_flags_dirty_and_improvement_opportunities(self):
        records = [
            sanitize_customer(
                {
                    "Id": 1,
                    "Nome": "A",
                    "Email": "email-invalido",
                    "TelefoneCelular": "123",
                    "CNPJ_CPF": "11111111111",
                    "Logradouro": "",
                    "Bairro": "",
                    "Numero": "",
                    "Nascimento": "",
                }
            ),
            sanitize_customer(
                {
                    "Id": 2,
                    "Nome": "Cliente Bom",
                    "Email": "cliente@cakeco.com.br",
                    "TelefoneCelular": "(21) 99999-9999",
                    "CNPJ_CPF": "111.444.777-35",
                    "Logradouro": "Rua A",
                    "Bairro": "Leblon",
                    "Numero": "10",
                    "Nascimento": "",
                }
            ),
        ]

        analysis = analyze_records(records)

        self.assertEqual(analysis["total_clientes"], 2)
        self.assertEqual(analysis["dirty_counts"]["nome_suspeito"], 1)
        self.assertEqual(analysis["dirty_counts"]["telefone_invalido"], 1)
        self.assertEqual(analysis["dirty_counts"]["email_invalido"], 1)
        self.assertEqual(analysis["dirty_counts"]["documento_invalido"], 1)
        self.assertEqual(analysis["dirty_counts"]["endereco_incompleto"], 1)
        self.assertEqual(analysis["improvement_counts"]["sem_aniversario"], 2)
        self.assertEqual(analysis["dirty_records"][0]["id"], "1")
        self.assertEqual(analysis["dirty_records"][0]["documento"], "***")

    def test_analyze_records_detects_duplicate_phone_and_document(self):
        records = [
            sanitize_customer({"Id": 1, "Nome": "Cliente A", "TelefoneCelular": "21999999999", "CNPJ_CPF": "11144477735"}),
            sanitize_customer({"Id": 2, "Nome": "Cliente B", "TelefoneCelular": "(21) 99999-9999", "CNPJ_CPF": "111.444.777-35"}),
        ]

        analysis = analyze_records(records)

        self.assertEqual(analysis["duplicate_counts"]["telefone"], 1)
        self.assertEqual(analysis["duplicate_counts"]["documento"], 1)
        self.assertEqual(analysis["duplicates"]["telefone"][0]["ids"], ["1", "2"])
        self.assertEqual(analysis["duplicates"]["documento"][0]["ids"], ["1", "2"])

    def test_compare_snapshots_reports_new_removed_and_tracked_field_changes(self):
        previous = [
            sanitize_customer({"Id": 1, "Nome": "Cliente A", "Email": "a@cakeco.com.br", "TelefoneCelular": "21999999999"}),
            sanitize_customer({"Id": 2, "Nome": "Cliente Removido", "Email": "removido@cakeco.com.br"}),
        ]
        current = [
            sanitize_customer({"Id": 1, "Nome": "Cliente A", "Email": "novo@cakeco.com.br", "TelefoneCelular": "21999999999"}),
            sanitize_customer({"Id": 3, "Nome": "Cliente Novo", "Email": "novo-cliente@cakeco.com.br"}),
        ]

        comparison = compare_snapshots(previous, current)

        self.assertEqual(comparison["new_count"], 1)
        self.assertEqual(comparison["removed_count"], 1)
        self.assertEqual(comparison["changed_count"], 1)
        self.assertEqual(comparison["new_records"][0]["id"], "3")
        self.assertEqual(comparison["removed_records"][0]["id"], "2")
        self.assertEqual(comparison["changed_records"][0]["id"], "1")
        self.assertEqual(comparison["changed_records"][0]["changes"][0]["field"], "email")
        self.assertEqual(comparison["changed_records"][0]["changes"][0]["old"], "a@c***.br")
        self.assertEqual(comparison["changed_records"][0]["changes"][0]["new"], "n***@c***.br")

    def test_compare_snapshots_treats_missing_previous_snapshot_as_initial_baseline(self):
        current = [
            sanitize_customer({"Id": 1, "Nome": "Cliente A"}),
            sanitize_customer({"Id": 2, "Nome": "Cliente B"}),
        ]

        comparison = compare_snapshots([], current)

        self.assertTrue(comparison["initial_baseline"])
        self.assertEqual(comparison["new_count"], 0)
        self.assertEqual(comparison["removed_count"], 0)
        self.assertEqual(comparison["changed_count"], 0)
        self.assertEqual(comparison["new_records"], [])

    def test_build_summary_message_mentions_friday_snapshot_and_main_counts(self):
        message = build_summary_message(
            run_date=date(2026, 6, 12),
            analysis={"total_clientes": 10, "dirty_records_count": 2, "improvement_records_count": 3},
            comparison={"new_count": 1, "removed_count": 0, "changed_count": 4},
            xlsx_path="/tmp/analise.xlsx",
        )

        self.assertIn("Cadastro de clientes Mogo", message)
        self.assertIn("12/06/2026", message)
        self.assertIn("10 clientes", message)
        self.assertIn("2 com cadastro sujo", message)
        self.assertIn("4 alterados", message)
        self.assertIn("analise.xlsx", message)

    def test_write_excel_includes_full_customer_registry_sheet(self):
        records = [
            sanitize_customer(
                {
                    "Id": 1,
                    "Nome": "Cliente A",
                    "Email": "a@cakeco.com.br",
                    "CodClienteIFood": "ifood-1",
                    "Senha": "nao-pode-salvar",
                }
            )
        ]
        analysis = analyze_records(records)
        comparison = compare_snapshots([], records)

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "cadastro.xlsx"
            write_excel(path, records, analysis, comparison)

            workbook = openpyxl.load_workbook(path)
            self.assertIn("Cadastro Completo", workbook.sheetnames)
            ws = workbook["Cadastro Completo"]
            headers = [cell.value for cell in ws[1]]
            self.assertIn("cod_cliente_ifood", headers)
            self.assertNotIn("senha", headers)
            self.assertEqual(ws.cell(row=2, column=headers.index("email") + 1).value, "a@cakeco.com.br")

    def test_write_excel_keeps_full_email_and_phone_in_work_sheets(self):
        previous = [
            sanitize_customer(
                {
                    "Id": 1,
                    "Nome": "Cliente A",
                    "Email": "antigo@cakeco.com.br",
                    "TelefoneCelular": "21988887777",
                }
            ),
            sanitize_customer(
                {
                    "Id": 2,
                    "Nome": "Cliente Removido",
                    "Email": "removido@cakeco.com.br",
                    "TelefoneCelular": "21977776666",
                }
            ),
        ]
        records = [
            sanitize_customer(
                {
                    "Id": 1,
                    "Nome": "A",
                    "Email": "novo@cakeco.com.br",
                    "TelefoneCelular": "123",
                    "CNPJ_CPF": "11111111111",
                }
            ),
            sanitize_customer(
                {
                    "Id": 3,
                    "Nome": "Cliente Novo",
                    "Email": "novo-cliente@cakeco.com.br",
                    "TelefoneCelular": "21999998888",
                }
            ),
        ]
        analysis = analyze_records(records)
        comparison = compare_snapshots(previous, records)

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "cadastro.xlsx"
            write_excel(path, records, analysis, comparison, previous_records=previous)

            workbook = openpyxl.load_workbook(path)

            dirty = workbook["Cadastro Sujo"]
            dirty_headers = [cell.value for cell in dirty[1]]
            self.assertEqual(dirty.cell(row=2, column=dirty_headers.index("email") + 1).value, "novo@cakeco.com.br")
            self.assertEqual(dirty.cell(row=2, column=dirty_headers.index("telefone_celular") + 1).value, "123")

            new = workbook["Novos"]
            new_headers = [cell.value for cell in new[1]]
            self.assertEqual(new.cell(row=2, column=new_headers.index("email") + 1).value, "novo-cliente@cakeco.com.br")
            self.assertEqual(new.cell(row=2, column=new_headers.index("telefone_celular") + 1).value, "21999998888")

            removed = workbook["Removidos"]
            removed_headers = [cell.value for cell in removed[1]]
            self.assertEqual(removed.cell(row=2, column=removed_headers.index("email") + 1).value, "removido@cakeco.com.br")
            self.assertEqual(removed.cell(row=2, column=removed_headers.index("telefone_celular") + 1).value, "21977776666")

            changes = workbook["Alteracoes"]
            change_rows = list(changes.iter_rows(min_row=2, values_only=True))
            self.assertIn(("1", "A", "email", "antigo@cakeco.com.br", "novo@cakeco.com.br"), change_rows)
            self.assertIn(("1", "A", "telefone_celular", "21988887777", "123"), change_rows)


if __name__ == "__main__":
    unittest.main()
