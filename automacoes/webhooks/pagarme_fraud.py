#!/usr/bin/env python3
"""Risk scoring for Pagar.me charge webhooks.

V1 goal: flag paid charges that should be confirmed before delivery.
No automatic cancellation/refund happens here.
"""

from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Protocol

WINDOW_MINUTES = 48 * 60
WINDOW_LABEL = "48h"
ALERT_THRESHOLD = 50
KNOWN_FRAUD_DELIVERY_ADDRESSES = (
    ("euclides da cunha", "106", "Rua Euclides da Cunha, 106"),
    ("major rubens vaz", "122", "Rua Major Rubens Vaz, 122"),
    ("major rubens vaz", "127", "Rua Major Rubens Vaz, 127"),
    ("santos dumont", "55", "Praça Santos Dumont, 55"),
)
DEFAULT_HOTLIST_PATH = Path(__file__).resolve().parents[1] / "data" / "pagarme_fraud_hotlist.json"


def _parse_dt(value: str | None) -> datetime:
    if not value:
        return datetime.now(timezone.utc)
    value = value.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(value)
    except ValueError:
        return datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def normalize_text(value: str | None) -> str:
    value = value or ""
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def only_digits(value: str | None) -> str:
    return re.sub(r"\D+", "", value or "")


def normalized_sha256(value: str | None) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return ""
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def digits_sha256(value: str | None) -> str:
    digits = only_digits(value)
    if not digits:
        return ""
    return hashlib.sha256(digits.encode("utf-8")).hexdigest()


def card_sha256(brand: str | None, last4: str | None) -> str:
    brand_norm = normalize_text(brand)
    last4_digits = only_digits(last4)
    if not brand_norm or not last4_digits:
        return ""
    return hashlib.sha256(f"{brand_norm}|{last4_digits}".encode("utf-8")).hexdigest()


def _first_present(row: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return str(value)
    return ""


def _clean_mogo_value(value: str | None) -> str:
    value = value or ""
    value = re.sub(r"<[^>]+>", " ", value)
    value = value.replace("&nbsp;", " ")
    return re.sub(r"\s+", " ", value).strip()


def _parse_brlish_number(value: str | int | float | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"[^0-9,.-]+", "", text)
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _format_address_parts(row: dict[str, Any]) -> tuple[str, str]:
    address_parts = [
        _clean_mogo_value(_first_present(row, ("Logradouro", "logradouro", "endereco", "Endereço"))),
        _clean_mogo_value(_first_present(row, ("Numero", "Número", "numero", "Nº"))),
        _clean_mogo_value(_first_present(row, ("Complemento", "complemento"))),
    ]
    address = ", ".join(part for part in address_parts if part)
    city = _clean_mogo_value(_first_present(row, ("Cidade", "cidade")))
    state = _clean_mogo_value(_first_present(row, ("Estado", "UF", "estado", "uf")))
    neighborhood_parts = [
        _clean_mogo_value(_first_present(row, ("Bairro", "bairro"))),
        f"{city}/{state}" if city and state else city or state,
    ]
    return address, " - ".join(part for part in neighborhood_parts if part)


def _extract_phone_from_obj(value: Any) -> str:
    if isinstance(value, str):
        return only_digits(value)
    if not isinstance(value, dict):
        return ""
    country = only_digits(str(value.get("country_code") or value.get("ddi") or ""))
    area = only_digits(str(value.get("area_code") or value.get("ddd") or ""))
    number = only_digits(str(value.get("number") or value.get("phone") or value.get("telefone") or ""))
    compact = only_digits(str(value.get("full_number") or value.get("fullNumber") or ""))
    return compact or f"{country}{area}{number}".lstrip("0")


def extract_customer_phone(customer: dict[str, Any]) -> str:
    direct = _extract_phone_from_obj(customer.get("phone") or customer.get("telefone") or customer.get("whatsapp"))
    if direct:
        return direct
    phones = customer.get("phones") or {}
    if isinstance(phones, dict):
        for key in ("mobile_phone", "home_phone", "customer_phone", "phone"):
            phone = _extract_phone_from_obj(phones.get(key))
            if phone:
                return phone
    return ""


def names_compatible(customer_name: str | None, holder_name: str | None) -> bool:
    """Compare names while accepting middle-name initials.

    Example accepted: "Joao Victor Martins" vs "JOAO V MARTINS".
    """
    customer = normalize_text(customer_name).split()
    holder = normalize_text(holder_name).split()
    if not customer or not holder:
        return True
    if customer == holder:
        return True
    # Require first and last names to match exactly when both exist.
    if len(customer) >= 2 and len(holder) >= 2:
        if customer[0] != holder[0] or customer[-1] != holder[-1]:
            return False
        # Middle tokens may be full names or matching initials.
        for h_token, c_token in zip(holder[1:-1], customer[1:-1]):
            if h_token == c_token:
                continue
            if len(h_token) == 1 and c_token.startswith(h_token):
                continue
            return False
        return True
    return normalize_text(customer_name) == normalize_text(holder_name)


def _meaningful_name_tokens(customer_name: str | None) -> list[str]:
    ignored = {"da", "de", "do", "das", "dos", "e"}
    return [
        token for token in normalize_text(customer_name).split()
        if len(token) >= 3 and token not in ignored
    ]


def customer_name_part_in_email_or_holder(customer_name: str | None, email: str | None, holder_name: str | None) -> bool:
    """Accept holder mismatch when a meaningful customer-name token appears in email or holder.

    This avoids noisy alerts for cases like customer "Iasminy" with holder
    "IASMINY VERGETTI" and email "vergetti.iasminy@gmail.com".
    """
    tokens = _meaningful_name_tokens(customer_name)
    if not tokens:
        return False

    holder_tokens = set(normalize_text(holder_name).split())
    email_user = normalize_text((email or "").split("@", 1)[0])
    email_tokens = set(email_user.split())
    compact_email_user = email_user.replace(" ", "")

    for token in tokens:
        if token in holder_tokens or token in email_tokens:
            return True
        if len(token) >= 4 and token in compact_email_user:
            return True
    return False


def _order_address_key(order: "MogoOrderSummary" | None) -> str:
    if order is None:
        return ""
    return normalize_text(" ".join(part for part in (order.address, order.neighborhood) if part))


def _related_address_key(order: "MogoOrderSummary" | None) -> str:
    if not order or not order.address:
        return ""
    value = f" {normalize_text(order.address)} "
    replacements = {
        " r ": " rua ", " av ": " avenida ", " ap ": " apto ",
        " apartamento ": " apto ", " n ": " ", " numero ": " ",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    return re.sub(r"\s+", " ", value).strip()


def _known_fraud_delivery_address(order: "MogoOrderSummary" | None) -> str:
    if order is None:
        return ""
    address = normalize_text(order.address)
    tokens = set(address.split())
    return _known_fraud_address_label(address, tokens)


def _known_fraud_address_label(address: str, tokens: set[str] | None = None) -> str:
    address = normalize_text(address)
    tokens = tokens or set(address.split())
    for street, number, label in KNOWN_FRAUD_DELIVERY_ADDRESSES:
        if street in address and number in tokens:
            return label
    return ""


def _format_pagarme_card_billing_address(card: dict[str, Any]) -> str:
    address = card.get("billing_address") or {}
    if not isinstance(address, dict):
        return ""
    parts = [
        str(address.get("line_1") or ""),
        str(address.get("line_2") or ""),
        str(address.get("city") or ""),
        str(address.get("state") or ""),
        str(address.get("zip_code") or ""),
    ]
    return " ".join(part for part in parts if part)


def _extract_card_first_six_from_raw(raw: dict[str, Any]) -> str:
    data = raw.get("data") or {}
    if not isinstance(data, dict):
        return ""
    tx = data.get("last_transaction") or {}
    if not isinstance(tx, dict):
        return ""
    card = tx.get("card") or {}
    if not isinstance(card, dict):
        return ""
    return only_digits(str(card.get("first_six_digits") or ""))


def _extract_card_identity_from_raw(raw: dict[str, Any]) -> tuple[str, str, str, str]:
    data = raw.get("data") or {}
    if not isinstance(data, dict):
        return ("", "", "", "")
    tx = data.get("last_transaction") or {}
    if not isinstance(tx, dict):
        return ("", "", "", "")
    card = tx.get("card") or {}
    if not isinstance(card, dict):
        return ("", "", "", "")
    return (
        only_digits(str(card.get("first_six_digits") or "")),
        only_digits(str(card.get("last_four_digits") or "")),
        only_digits(str(card.get("exp_month") or "")),
        only_digits(str(card.get("exp_year") or "")),
    )


def _names_share_meaningful_part(left: str | None, right: str | None) -> bool:
    if names_compatible(left, right):
        return True
    left_tokens = set(_meaningful_name_tokens(left))
    right_tokens = set(_meaningful_name_tokens(right))
    if not left_tokens or not right_tokens:
        return False
    smaller = left_tokens if len(left_tokens) <= len(right_tokens) else right_tokens
    larger = right_tokens if smaller is left_tokens else left_tokens
    overlap = smaller & larger
    if len(smaller) == 1:
        return bool(overlap)
    return len(overlap) / len(smaller) >= 0.75


@dataclass(frozen=True)
class ChargeEvent:
    hook_id: str
    event_type: str
    charge_id: str
    status: str
    created_at: datetime
    amount: int
    customer_name: str
    customer_email: str
    customer_document: str
    customer_phone: str
    card_brand: str
    card_last4: str
    holder_name: str
    holder_document: str
    card_billing_address: str
    acquirer_message: str
    acquirer_return_code: str
    payment_method: str
    raw: dict[str, Any]

    @property
    def identity_key(self) -> str:
        document = normalize_text(self.customer_document)
        email = normalize_text(self.customer_email)
        name = normalize_text(self.customer_name)
        return document or email or name

    @property
    def card_key(self) -> str:
        return "|".join([
            normalize_text(self.card_brand),
            normalize_text(self.card_last4),
            normalize_text(self.holder_name),
        ])

    @property
    def is_pix(self) -> bool:
        return normalize_text(self.payment_method) == "pix"

    @property
    def is_card(self) -> bool:
        return normalize_text(self.payment_method) in {"credit card", "creditcard", "cartao", "cartao credito"}

    @property
    def is_paid(self) -> bool:
        return self.event_type == "charge.paid" or self.status == "paid"

    @property
    def is_failed(self) -> bool:
        return self.event_type == "charge.payment_failed" or self.status in {"failed", "not_authorized"}


@dataclass(frozen=True)
class MogoOrderSummary:
    order_number: str = ""
    status: str = ""
    customer_name: str = ""
    date: str = ""
    delivery_date: str = ""
    delivery_time: str = ""
    fulfillment: str = ""
    address: str = ""
    neighborhood: str = ""
    amount: str = ""
    origin: str = ""
    item: str = ""
    phone: str = ""
    document: str = ""
    email: str = ""


@dataclass(frozen=True)
class RelatedCustomerProfile:
    match_kind: str
    match_reason: str
    name: str = ""
    phone: str = ""
    email: str = ""
    document: str = ""
    address: str = ""
    last_purchase_date: str = ""
    last_purchase_amount: str = ""
    valid_purchase_count: int = 0


@dataclass(frozen=True)
class CustomerHistoryResult:
    has_prior_valid_purchase: bool
    matched_by: str | None
    status: str
    error: str | None = None
    order: MogoOrderSummary | None = None
    valid_purchase_count: int = 0
    operational_order: MogoOrderSummary | None = None
    related_profiles: tuple[RelatedCustomerProfile, ...] = ()


class CustomerHistoryChecker(Protocol):
    def lookup(self, charge: ChargeEvent) -> CustomerHistoryResult:
        ...


class NoopCustomerHistoryChecker:
    def lookup(self, charge: ChargeEvent) -> CustomerHistoryResult:
        return CustomerHistoryResult(False, None, "not_configured", None)


class CompositeCustomerHistoryChecker:
    def __init__(self, *checkers: CustomerHistoryChecker):
        self.checkers = checkers

    def lookup(self, charge: ChargeEvent) -> CustomerHistoryResult:
        merged = CustomerHistoryResult(False, None, "not_found", None)
        for checker in self.checkers:
            result = checker.lookup(charge)
            if result.has_prior_valid_purchase and not merged.has_prior_valid_purchase:
                merged = CustomerHistoryResult(
                    result.has_prior_valid_purchase,
                    result.matched_by,
                    result.status,
                    result.error,
                    result.order,
                    result.valid_purchase_count,
                    merged.operational_order or result.operational_order,
                    result.related_profiles or merged.related_profiles,
                )
            elif result.status == "error" and not merged.has_prior_valid_purchase and merged.status != "error":
                merged = CustomerHistoryResult(False, None, "error", result.error, None, 0, merged.operational_order or result.operational_order, result.related_profiles or merged.related_profiles)

            if result.operational_order and not merged.operational_order:
                merged = CustomerHistoryResult(
                    merged.has_prior_valid_purchase,
                    merged.matched_by,
                    merged.status,
                    merged.error,
                    merged.order,
                    merged.valid_purchase_count,
                    result.operational_order,
                    merged.related_profiles or result.related_profiles,
                )
            elif result.related_profiles and not merged.related_profiles:
                merged = CustomerHistoryResult(
                    merged.has_prior_valid_purchase, merged.matched_by, merged.status,
                    merged.error, merged.order, merged.valid_purchase_count,
                    merged.operational_order, result.related_profiles,
                )
        return merged


@dataclass(frozen=True)
class FraudHotlist:
    holder_name_hashes: frozenset[str]
    customer_name_hashes: frozenset[str] = frozenset()
    customer_email_hashes: frozenset[str] = frozenset()
    customer_document_hashes: frozenset[str] = frozenset()
    card_hashes: frozenset[str] = frozenset()

    @classmethod
    def empty(cls) -> "FraudHotlist":
        return cls(frozenset())

    @classmethod
    def from_holder_names(cls, names: list[str]) -> "FraudHotlist":
        return cls(frozenset(hash_ for hash_ in (normalized_sha256(name) for name in names) if hash_))

    @classmethod
    def from_customer_documents(cls, documents: list[str]) -> "FraudHotlist":
        return cls(frozenset(), customer_document_hashes=frozenset(hash_ for hash_ in (digits_sha256(document) for document in documents) if hash_))

    @classmethod
    def from_cards(cls, cards: list[tuple[str, str]]) -> "FraudHotlist":
        return cls(frozenset(), card_hashes=frozenset(hash_ for hash_ in (card_sha256(brand, last4) for brand, last4 in cards) if hash_))

    @classmethod
    def from_path(cls, path: str | Path) -> "FraudHotlist":
        path = Path(path)
        if not path.exists():
            return cls.empty()
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return cls.empty()

        def read_hashes(key: str) -> frozenset[str]:
            raw_hashes = payload.get(key, [])
            if not isinstance(raw_hashes, list):
                return frozenset()
            return frozenset({
                str(value).strip().lower()
                for value in raw_hashes
                if isinstance(value, str) and re.fullmatch(r"[0-9a-fA-F]{64}", value.strip())
            })

        return cls(
            holder_name_hashes=read_hashes("holder_name_sha256"),
            customer_name_hashes=read_hashes("customer_name_sha256"),
            customer_email_hashes=read_hashes("customer_email_sha256"),
            customer_document_hashes=read_hashes("customer_document_sha256"),
            card_hashes=read_hashes("card_sha256"),
        )

    def matches(self, charge: ChargeEvent) -> bool:
        # The operational hotlist is anchored on the identity the customer used
        # to create the Pagar.me/Mogo customer: name, CPF/document and email.
        # Card holder and card last4 are too weak to block delivery on their own.
        hashes = [
            (normalized_sha256(charge.customer_name), self.customer_name_hashes),
            (normalized_sha256(charge.customer_email), self.customer_email_hashes),
            (digits_sha256(charge.customer_document), self.customer_document_hashes),
        ]
        return any(hash_ and hash_ in hotlist_hashes for hash_, hotlist_hashes in hashes)


class LocalMogoHistoryChecker:
    """Lookup prior valid purchases in local Mogo JSON exports.

    The checker indexes only records that represent paid/concluded history.
    Matching priority is document, email, phone, then careful name fallback.
    """

    VALID_STATUS = {"pago", "paga", "entregue", "concluido", "concluida", "finalizado", "finalizada"}

    def __init__(self, reports_root: str | Path):
        self.reports_root = Path(reports_root)
        self._loaded = False
        self._documents: set[str] = set()
        self._emails: set[str] = set()
        self._phones: set[str] = set()
        self._names: set[str] = set()
        self._document_orders: dict[str, MogoOrderSummary] = {}
        self._email_orders: dict[str, MogoOrderSummary] = {}
        self._phone_orders: dict[str, MogoOrderSummary] = {}
        self._name_orders: dict[str, MogoOrderSummary] = {}
        self._document_order_ids: dict[str, set[str]] = {}
        self._email_order_ids: dict[str, set[str]] = {}
        self._phone_order_ids: dict[str, set[str]] = {}
        self._name_order_ids: dict[str, set[str]] = {}
        self._valid_orders: list[MogoOrderSummary] = []
        self._valid_order_ids: set[str] = set()
        self._operational_orders: list[MogoOrderSummary] = []

    def lookup(self, charge: ChargeEvent) -> CustomerHistoryResult:
        try:
            self._load_once()
        except Exception as exc:
            return CustomerHistoryResult(False, None, "error", exc.__class__.__name__)

        operational_order = self._find_operational_order(charge)
        related_profiles = self._find_related_profiles(charge, operational_order)
        name_address_order = self._find_valid_purchase_by_name_and_address(charge, operational_order)
        if name_address_order:
            return CustomerHistoryResult(True, "name_address", "valid_purchase", None, name_address_order, 1, operational_order, related_profiles)

        document = only_digits(charge.customer_document)
        if document and document in self._documents:
            return CustomerHistoryResult(True, "document", "valid_purchase", None, self._document_orders.get(document), self._valid_count(self._document_order_ids, document), operational_order, related_profiles)

        email = normalize_text(charge.customer_email)
        if email and email in self._emails:
            return CustomerHistoryResult(True, "email", "valid_purchase", None, self._email_orders.get(email), self._valid_count(self._email_order_ids, email), operational_order, related_profiles)

        phone = only_digits(charge.customer_phone)
        if phone:
            for indexed in self._phones:
                if phone == indexed or phone.endswith(indexed) or indexed.endswith(phone):
                    return CustomerHistoryResult(True, "phone", "valid_purchase", None, self._phone_orders.get(indexed), self._valid_count(self._phone_order_ids, indexed), operational_order, related_profiles)

        name = normalize_text(charge.customer_name)
        if name:
            for indexed_name in self._names:
                if names_compatible(name, indexed_name):
                    return CustomerHistoryResult(True, "name", "valid_purchase", None, self._name_orders.get(indexed_name), self._valid_count(self._name_order_ids, indexed_name), operational_order, related_profiles)

        return CustomerHistoryResult(False, None, "not_found", None, None, 0, operational_order, related_profiles)

    def _find_related_profiles(
        self,
        charge: ChargeEvent,
        operational_order: MogoOrderSummary | None,
    ) -> tuple[RelatedCustomerProfile, ...]:
        current_address = _related_address_key(operational_order)
        holder_mismatch = bool(charge.holder_name and not names_compatible(charge.customer_name, charge.holder_name))
        profiles: list[tuple[int, RelatedCustomerProfile]] = []
        seen: set[str] = set()
        for order in self._valid_orders:
            identity = only_digits(order.document) or normalize_text(order.email) or only_digits(order.phone) or normalize_text(order.customer_name)
            if not identity or identity in seen:
                continue
            if (
                (only_digits(charge.customer_document) and only_digits(charge.customer_document) == only_digits(order.document))
                or (normalize_text(charge.customer_email) and normalize_text(charge.customer_email) == normalize_text(order.email))
                or names_compatible(charge.customer_name, order.customer_name)
            ):
                continue
            address_match = bool(current_address and current_address == _related_address_key(order))
            strong_holder = bool(holder_mismatch and names_compatible(charge.holder_name, order.customer_name))
            holder_tokens = set(_meaningful_name_tokens(charge.holder_name))
            order_tokens = set(_meaningful_name_tokens(order.customer_name))
            phone_match = bool(
                only_digits(charge.customer_phone)
                and only_digits(order.phone)
                and (
                    only_digits(charge.customer_phone).endswith(only_digits(order.phone))
                    or only_digits(order.phone).endswith(only_digits(charge.customer_phone))
                )
            )
            email_match = bool(
                normalize_text(charge.customer_email)
                and normalize_text(charge.customer_email) == normalize_text(order.email)
            )
            partial_holder = bool(
                holder_mismatch and holder_tokens & order_tokens
                and (address_match or phone_match or email_match)
            )
            if strong_holder:
                kind, reason, priority = "strong_holder_name", "correspondência forte pelo titular", 3
            elif partial_holder:
                kind, reason, priority = "partial_holder_confirmed", "possível relação pelo titular com confirmação adicional", 2
            elif address_match:
                kind, reason, priority = "exact_address", "endereço completo igual", 2
            else:
                continue
            count_key = normalize_text(order.customer_name)
            count = self._valid_count(self._name_order_ids, count_key) or 1
            profiles.append((priority, RelatedCustomerProfile(
                match_kind=kind,
                match_reason=reason,
                name=order.customer_name,
                phone=order.phone,
                email=order.email,
                document=order.document,
                address=" - ".join(part for part in (order.address, order.neighborhood) if part),
                last_purchase_date=order.date,
                last_purchase_amount=order.amount,
                valid_purchase_count=count,
            )))
            seen.add(identity)
        profiles.sort(key=lambda item: item[0], reverse=True)
        return tuple(profile for _, profile in profiles[:3])

    def _find_valid_purchase_by_name_and_address(self, charge: ChargeEvent, operational_order: MogoOrderSummary | None) -> MogoOrderSummary | None:
        current_address = _order_address_key(operational_order)
        if not current_address:
            return None
        current_name = operational_order.customer_name if operational_order and operational_order.customer_name else charge.customer_name
        for order in self._valid_orders:
            if not _order_address_key(order) or _order_address_key(order) != current_address:
                continue
            if _names_share_meaningful_part(current_name, order.customer_name):
                return order
        return None

    def _load_once(self) -> None:
        if self._loaded:
            return
        if not self.reports_root.exists():
            self._loaded = True
            return

        for path in self.reports_root.rglob("*.json"):
            try:
                payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
            except Exception:
                continue
            for row in self._iter_rows(payload):
                if self._is_operational_order_row(path, row):
                    self._index_operational_row(row)
                if self._is_valid_purchase_row(path, row):
                    self._index_row(row)
        for path in self._operational_xlsx_paths():
            for row in self._iter_xlsx_rows(path):
                self._index_operational_row(row)
        self._loaded = True

    def _operational_xlsx_paths(self) -> list[Path]:
        folders = ("Pendentes", "Na Entrega", "Pedidos Entregues")
        paths: list[Path] = []
        for folder in folders:
            candidate = self.reports_root / folder
            if candidate.exists():
                paths.extend(sorted(candidate.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)[:10])
        return paths

    def _iter_xlsx_rows(self, path: Path) -> list[dict[str, Any]]:
        try:
            import openpyxl
        except Exception:
            return []
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows)]
            return [
                {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
                for row in rows
            ]
        except Exception:
            return []

    def _iter_rows(self, payload: Any) -> list[dict[str, Any]]:
        if isinstance(payload, list):
            return [row for row in payload if isinstance(row, dict)]
        if not isinstance(payload, dict):
            return []
        for key in ("registros", "rows", "data", "items", "records"):
            value = payload.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]
        return []

    def _is_valid_purchase_row(self, path: Path, row: dict[str, Any]) -> bool:
        folder = normalize_text(path.parent.name)
        status = normalize_text(_first_present(row, ("A10", "StatusEntrega", "StatusPedido", "StatusPago", "status", "situacao", "posicao")))
        if status:
            return status in self.VALID_STATUS
        if "historico pagamento" in folder:
            return bool(_first_present(row, ("dataPag", "idPag", "numPed", "chave")))
        if "analise cadastro clientes" in folder:
            first_order = _first_present(row, ("primeiro_pedido", "Primeiro Pedido"))
            last_order = _first_present(row, ("ultimo_pedido", "Último Pedido", "Ultimo Pedido"))
            total_orders = _parse_brlish_number(_first_present(row, ("total_pedidos", "Total Pedidos")))
            total_delivery = _parse_brlish_number(_first_present(row, ("total_delivery", "Total Delivery")))
            has_order_window = bool(first_order or last_order)
            has_paid_total = any(value is not None and value > 0 for value in (total_orders, total_delivery))
            return has_order_window and has_paid_total
        return False

    def _is_operational_order_row(self, path: Path, row: dict[str, Any]) -> bool:
        folder = normalize_text(path.parent.name)
        if any(name in folder for name in ("pendentes", "na entrega", "pedidos entregues")):
            return True
        return bool(_first_present(row, ("NumeroPedido", "Nº Pedido"))) and bool(_first_present(row, ("DataEntrega", "Data Agendada", "HoraEntregaTxt", "Hora Agendada")))

    def _index_operational_row(self, row: dict[str, Any]) -> None:
        order = self._order_summary(row)
        if order.order_number or order.customer_name:
            self._operational_orders.append(order)

    def _find_operational_order(self, charge: ChargeEvent) -> MogoOrderSummary | None:
        best: tuple[int, MogoOrderSummary] | None = None
        for order in self._operational_orders:
            score = self._operational_match_score(charge, order)
            if score and (best is None or score > best[0]):
                best = (score, order)
        if best and best[0] >= 60:
            return best[1]
        return None

    def _operational_match_score(self, charge: ChargeEvent, order: MogoOrderSummary) -> int:
        score = 0
        charge_phone = only_digits(charge.customer_phone)
        order_phone = only_digits(order.phone)
        if charge_phone and order_phone and (charge_phone == order_phone or charge_phone.endswith(order_phone) or order_phone.endswith(charge_phone)):
            score += 100
        if normalize_text(charge.customer_email) and normalize_text(charge.customer_email) == normalize_text(order.email):
            score += 90
        if only_digits(charge.customer_document) and only_digits(charge.customer_document) == only_digits(order.document):
            score += 100
        if names_compatible(charge.customer_name, order.customer_name):
            score += 30
        order_amount = _parse_brlish_number(order.amount)
        if order_amount is not None and abs(order_amount - (charge.amount / 100)) < 0.01:
            score += 30
        return score

    def _index_row(self, row: dict[str, Any]) -> None:
        order = self._order_summary(row)
        purchase_id = self._purchase_id(order)
        if purchase_id not in self._valid_order_ids:
            self._valid_orders.append(order)
            self._valid_order_ids.add(purchase_id)

        document = only_digits(_first_present(row, ("document", "documento", "cpf", "cnpj", "customer_document")))
        if document:
            self._index_purchase(document, order, self._documents, self._document_orders, self._document_order_ids)

        email = normalize_text(_first_present(row, ("email", "customer_email", "cliente_email")))
        if email:
            self._index_purchase(email, order, self._emails, self._email_orders, self._email_order_ids)

        phone = only_digits(_first_present(row, ("telefone", "phone", "whatsapp", "celular", "customer_phone")))
        if phone:
            self._index_purchase(phone, order, self._phones, self._phone_orders, self._phone_order_ids)

        name = normalize_text(_first_present(row, ("cliente", "A5", "nome", "NomeCliente", "customer_name")))
        if name:
            self._index_purchase(name, order, self._names, self._name_orders, self._name_order_ids)

    def _index_purchase(
        self,
        key: str,
        order: MogoOrderSummary,
        keys: set[str],
        orders: dict[str, MogoOrderSummary],
        order_ids: dict[str, set[str]],
    ) -> None:
        keys.add(key)
        orders.setdefault(key, order)
        order_ids.setdefault(key, set()).add(self._purchase_id(order))

    def _purchase_id(self, order: MogoOrderSummary) -> str:
        if order.order_number:
            return order.order_number
        fallback = "|".join((order.date, order.amount, order.customer_name, order.item))
        return fallback or "-"

    def _valid_count(self, order_ids: dict[str, set[str]], key: str) -> int:
        return len(order_ids.get(key, set()))

    def _order_summary(self, row: dict[str, Any]) -> MogoOrderSummary:
        address, neighborhood = _format_address_parts(row)
        return MogoOrderSummary(
            order_number=_clean_mogo_value(_first_present(row, ("A13", "NumeroPedido", "Nº Pedido", "numPed", "pedido", "numero_pedido"))),
            status=_clean_mogo_value(_first_present(row, ("A10", "StatusEntrega", "StatusPedido", "StatusPago", "status", "situacao", "posicao", "Pago"))),
            customer_name=_clean_mogo_value(_first_present(row, ("cliente", "A5", "nome", "NomeCliente", "Cliente_Nome", "customer_name", "Cliente"))),
            date=_clean_mogo_value(_first_present(row, ("A0", "data", "Data Pedido", "DataPedido", "dataPed", "dataPag", "ultimo_pedido", "Último Pedido", "Ultimo Pedido"))),
            delivery_date=_clean_mogo_value(_first_present(row, ("Data Entrega", "DataEntrega", "Data Agendada"))),
            delivery_time=_clean_mogo_value(_first_present(row, ("Hora Entrega", "HoraEntregaTxt", "Hora Agendada"))),
            fulfillment=_clean_mogo_value(_first_present(row, ("A6", "Forma Entrega", "FormaEntrega", "ObsEntrega_Descricao", "forma_entrega", "delivery_type"))),
            address=address,
            neighborhood=neighborhood,
            amount=_clean_mogo_value(_first_present(row, ("ValorTotal", "ValorPago", "Valor Final", "ValorFinal", "totalped", "A4", "valor", "total_delivery", "Total Delivery", "total_pedidos", "Total Pedidos"))),
            origin=_clean_mogo_value(_first_present(row, ("Origem", "OrigemPedido", "origem"))),
            item=_clean_mogo_value(_first_present(row, ("A2", "Produto", "produto", "item"))),
            phone=_clean_mogo_value(_first_present(row, ("TelefoneCliente", "CelularCliente", "telefone", "phone", "whatsapp", "celular", "customer_phone"))),
            document=_clean_mogo_value(_first_present(row, ("Documento", "CPF", "document", "documento", "cpf", "cnpj", "customer_document"))),
            email=_clean_mogo_value(_first_present(row, ("Email", "email", "customer_email", "cliente_email"))),
        )


class LiveMogoOperationalOrderChecker:
    """Fetch current delivery/pickup context directly from Mogo open orders."""

    def __init__(self, scripts_dir: str | Path | None = None):
        self.scripts_dir = Path(scripts_dir or Path(__file__).resolve().parents[1] / "scripts")

    def lookup(self, charge: ChargeEvent) -> CustomerHistoryResult:
        try:
            order = self._lookup_order(charge)
        except BaseException:
            order = None
        return CustomerHistoryResult(False, None, "not_found", None, None, 0, order)

    def _lookup_order(self, charge: ChargeEvent) -> MogoOrderSummary | None:
        if str(self.scripts_dir) not in sys.path:
            sys.path.insert(0, str(self.scripts_dir))
        from mogo_login import MOGO_URL, mogo_login  # type: ignore

        session = mogo_login(verbose=False)
        rows = self._fetch_pending_rows(session, MOGO_URL) + self._fetch_delivery_rows(session, MOGO_URL)
        best: tuple[int, MogoOrderSummary] | None = None
        local = LocalMogoHistoryChecker(Path("/nonexistent"))
        for row in rows:
            order = local._order_summary(row)
            score = local._operational_match_score(charge, order)
            if score and (best is None or score > best[0]):
                best = (score, order)
        if best and best[0] >= 60:
            return best[1]
        return None

    def _fetch_pending_rows(self, session: Any, mogo_url: str) -> list[dict[str, Any]]:
        response = session.get(
            f"{mogo_url}/Pedido/ListPedidosParaEntrega",
            params={
                "_search": "true",
                "rows": "1000",
                "page": "1",
                "sidx": "DataEntrega",
                "sord": "asc",
                "filters": json.dumps({
                    "groupOp": "AND",
                    "rules": [{"field": "StatusEntrega", "op": "eq", "data": "Pendente"}],
                }),
            },
            timeout=30,
        )
        if response.status_code != 200:
            return []
        payload = response.json()
        rows = payload.get("rows", [])
        return [row for row in rows if isinstance(row, dict)]

    def _fetch_delivery_rows(self, session: Any, mogo_url: str) -> list[dict[str, Any]]:
        response = session.post(
            f"{mogo_url}/Pedido/ListPedidosParaEntrega",
            params={"cFiltroTipoEntrega": "1"},
            data={
                "_search": "false",
                "nd": "1",
                "rows": "1000",
                "page": "1",
                "sidx": "HoraInclusao",
                "sord": "desc",
            },
            timeout=30,
        )
        if response.status_code != 200:
            return []
        payload = response.json()
        rows = payload.get("rows", [])
        return [row for row in rows if isinstance(row, dict)]


@dataclass(frozen=True)
class SameDayPurchase:
    charge_id: str
    created_at: datetime
    amount: int


@dataclass(frozen=True)
class SameDayRepeatContext:
    kind: str
    sequence: int
    purchases: tuple[SameDayPurchase, ...]


@dataclass(frozen=True)
class RiskResult:
    alert: bool
    score: int
    reasons: list[str]
    charge: ChargeEvent
    customer_history: CustomerHistoryResult | None = None
    first_purchase_alert: bool = False
    same_day_repeat: SameDayRepeatContext | None = None


def extract_charge(payload: dict[str, Any]) -> ChargeEvent:
    data = payload.get("data") or {}
    tx = data.get("last_transaction") or {}
    card = tx.get("card") or {}
    customer = data.get("customer") or tx.get("customer") or {}
    return ChargeEvent(
        hook_id=str(payload.get("id") or ""),
        event_type=str(payload.get("type") or payload.get("event") or ""),
        charge_id=str(data.get("id") or ""),
        status=str(data.get("status") or tx.get("status") or ""),
        created_at=_parse_dt(data.get("created_at") or payload.get("created_at")),
        amount=int(data.get("amount") or tx.get("amount") or 0),
        customer_name=str(customer.get("name") or ""),
        customer_email=str(customer.get("email") or ""),
        customer_document=str(customer.get("document") or ""),
        customer_phone=extract_customer_phone(customer),
        card_brand=str(card.get("brand") or ""),
        card_last4=str(card.get("last_four_digits") or ""),
        holder_name=str(card.get("holder_name") or ""),
        holder_document=str(card.get("holder_document") or ""),
        card_billing_address=_format_pagarme_card_billing_address(card),
        acquirer_message=str(tx.get("acquirer_message") or ""),
        acquirer_return_code=str(tx.get("acquirer_return_code") or ""),
        payment_method=str(data.get("payment_method") or tx.get("payment_method") or ""),
        raw=payload,
    )


class RiskEngine:
    def __init__(self, db_path: str, history_checker: CustomerHistoryChecker | None = None, hotlist: FraudHotlist | None = None):
        self.db_path = db_path
        self.history_checker = history_checker or NoopCustomerHistoryChecker()
        self.hotlist = hotlist if hotlist is not None else FraudHotlist.from_path(DEFAULT_HOTLIST_PATH)
        Path(db_path).parent.mkdir(parents=True, exist_ok=True) if str(Path(db_path).parent) != "." else None
        self._init_db()

    def _connect(self):
        return sqlite3.connect(self.db_path)

    def _init_db(self) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS charge_events (
                    hook_id TEXT,
                    event_type TEXT NOT NULL,
                    charge_id TEXT PRIMARY KEY,
                    status TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    amount INTEGER NOT NULL,
                    identity_key TEXT NOT NULL,
                    customer_name TEXT,
                    customer_email TEXT,
                    customer_document TEXT,
                    card_key TEXT,
                    card_brand TEXT,
                    card_last4 TEXT,
                    holder_name TEXT,
                    acquirer_message TEXT,
                    acquirer_return_code TEXT,
                    raw_json TEXT NOT NULL
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_events_identity_time ON charge_events(identity_key, created_at)")

    def handle_event(self, payload: dict[str, Any]) -> RiskResult:
        charge = extract_charge(payload)
        if charge.is_pix:
            return RiskResult(False, 0, [], charge)
        self._store(charge)
        if not charge.is_paid:
            return RiskResult(False, 0, [], charge)
        score, reasons, history = self._score_paid_charge(charge)
        antifraud_alert = score >= ALERT_THRESHOLD
        first_purchase_alert = self._should_alert_first_purchase(charge, history, antifraud_alert)
        same_day_repeat = self._same_day_repeat_context(charge, history)
        return RiskResult(antifraud_alert, score, reasons, charge, history, first_purchase_alert, same_day_repeat)

    def _store(self, charge: ChargeEvent) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO charge_events (
                    hook_id, event_type, charge_id, status, created_at, amount, identity_key,
                    customer_name, customer_email, customer_document, card_key, card_brand,
                    card_last4, holder_name, acquirer_message, acquirer_return_code, raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    charge.hook_id,
                    charge.event_type,
                    charge.charge_id,
                    charge.status,
                    charge.created_at.isoformat(),
                    charge.amount,
                    charge.identity_key,
                    charge.customer_name,
                    charge.customer_email,
                    charge.customer_document,
                    charge.card_key,
                    charge.card_brand,
                    charge.card_last4,
                    charge.holder_name,
                    charge.acquirer_message,
                    charge.acquirer_return_code,
                    json.dumps(charge.raw, ensure_ascii=False),
                ),
            )

    def _recent_events(self, charge: ChargeEvent) -> list[sqlite3.Row]:
        since = (charge.created_at - timedelta(minutes=WINDOW_MINUTES)).isoformat()
        with self._connect() as conn:
            conn.row_factory = sqlite3.Row
            return list(
                conn.execute(
                    """
                    SELECT * FROM charge_events
                    WHERE identity_key = ? AND created_at >= ? AND charge_id != ?
                    ORDER BY created_at ASC
                    """,
                    (charge.identity_key, since, charge.charge_id),
                )
            )

    def _same_day_repeat_context(
        self,
        charge: ChargeEvent,
        history: CustomerHistoryResult,
    ) -> SameDayRepeatContext | None:
        if not charge.identity_key:
            return None
        brt = timezone(timedelta(hours=-3))
        local_date = charge.created_at.astimezone(brt).date()
        day_start = datetime(local_date.year, local_date.month, local_date.day, tzinfo=brt).astimezone(timezone.utc)
        with self._connect() as conn:
            conn.row_factory = sqlite3.Row
            try:
                rows = list(conn.execute(
                    """
                    SELECT e.charge_id, e.event_type, e.status, e.created_at, e.amount,
                           e.customer_name, COALESCE(r.decision, '') AS review_decision
                    FROM charge_events e
                    LEFT JOIN antifraud_reviews r ON r.charge_id = e.charge_id
                    WHERE e.identity_key = ?
                      AND e.charge_id != ?
                      AND e.created_at < ?
                    ORDER BY e.created_at ASC
                    """,
                    (charge.identity_key, charge.charge_id, charge.created_at.isoformat()),
                ))
            except sqlite3.OperationalError:
                rows = list(conn.execute(
                    """
                    SELECT charge_id, event_type, status, created_at, amount, customer_name,
                           '' AS review_decision
                    FROM charge_events
                    WHERE identity_key = ?
                      AND charge_id != ?
                      AND created_at < ?
                    ORDER BY created_at ASC
                    """,
                    (charge.identity_key, charge.charge_id, charge.created_at.isoformat()),
                ))

        valid_prior = self._valid_prior_paid_rows(rows)
        paid_today = [row for row in valid_prior if _parse_dt(row["created_at"]) >= day_start]
        if not paid_today:
            return None

        paid_before_today = any(_parse_dt(row["created_at"]) < day_start for row in valid_prior)
        external_history_before_today = self._history_predates_local_day(history, local_date)
        kind = (
            "informational_returning"
            if paid_before_today or external_history_before_today
            else "critical_first_day"
        )
        purchases = tuple(
            SameDayPurchase(
                charge_id=str(row["charge_id"] or ""),
                created_at=_parse_dt(row["created_at"]),
                amount=int(row["amount"] or 0),
            )
            for row in paid_today
        ) + (SameDayPurchase(charge.charge_id, charge.created_at, charge.amount),)
        return SameDayRepeatContext(kind, len(purchases), purchases)

    def _history_predates_local_day(
        self,
        history: CustomerHistoryResult,
        local_date,
    ) -> bool:
        if not history.has_prior_valid_purchase:
            return False
        order_date = (history.order.date if history.order else "").strip()
        if not order_date:
            return history.matched_by != "pagarme_prior_charge"
        candidates = (order_date, order_date[:10])
        for candidate in candidates:
            try:
                parsed = datetime.fromisoformat(candidate.replace("Z", "+00:00"))
                return parsed.date() < local_date
            except ValueError:
                pass
            try:
                return datetime.strptime(candidate, "%d/%m/%Y").date() < local_date
            except ValueError:
                pass
        return history.matched_by != "pagarme_prior_charge"

    def _same_day_card_prefix_matches(self, charge: ChargeEvent) -> list[sqlite3.Row]:
        first_six = _extract_card_first_six_from_raw(charge.raw)
        if not charge.is_card or not first_six or not normalize_text(charge.card_brand):
            return []

        brt = timezone(timedelta(hours=-3))
        local_date = charge.created_at.astimezone(brt).date()
        start = datetime(local_date.year, local_date.month, local_date.day, tzinfo=brt).astimezone(timezone.utc)
        end = start + timedelta(days=1)
        with self._connect() as conn:
            conn.row_factory = sqlite3.Row
            rows = list(
                conn.execute(
                    """
                    SELECT charge_id, event_type, status, created_at, identity_key,
                           customer_name, card_brand, card_last4, raw_json
                    FROM charge_events
                    WHERE charge_id != ?
                      AND identity_key != ?
                      AND created_at >= ?
                      AND created_at < ?
                      AND card_brand != ''
                    ORDER BY created_at DESC
                    """,
                    (charge.charge_id, charge.identity_key, start.isoformat(), end.isoformat()),
                )
            )

        matches: list[sqlite3.Row] = []
        for row in rows:
            if normalize_text(str(row["card_brand"] or "")) != normalize_text(charge.card_brand):
                continue
            try:
                raw = json.loads(str(row["raw_json"] or "{}"))
            except json.JSONDecodeError:
                continue
            if _extract_card_first_six_from_raw(raw) == first_six:
                matches.append(row)
        return matches

    def _exact_card_matches(self, charge: ChargeEvent) -> list[sqlite3.Row]:
        charge_card = _extract_card_identity_from_raw(charge.raw)
        if not charge.is_card or not normalize_text(charge.card_brand):
            return []
        if not all(charge_card):
            return []

        with self._connect() as conn:
            conn.row_factory = sqlite3.Row
            rows = list(
                conn.execute(
                    """
                    SELECT charge_id, event_type, status, created_at, identity_key,
                           customer_name, card_brand, card_last4, raw_json
                    FROM charge_events
                    WHERE charge_id != ?
                      AND identity_key != ?
                      AND card_brand != ''
                    ORDER BY created_at DESC
                    """,
                    (charge.charge_id, charge.identity_key),
                )
            )

        matches: list[sqlite3.Row] = []
        for row in rows:
            if normalize_text(str(row["card_brand"] or "")) != normalize_text(charge.card_brand):
                continue
            try:
                raw = json.loads(str(row["raw_json"] or "{}"))
            except json.JSONDecodeError:
                continue
            if _extract_card_identity_from_raw(raw) == charge_card:
                matches.append(row)
        return matches

    def _customer_history(self, charge: ChargeEvent) -> CustomerHistoryResult:
        try:
            history = self.history_checker.lookup(charge)
        except Exception as exc:
            return CustomerHistoryResult(False, None, "error", exc.__class__.__name__)
        if history.has_prior_valid_purchase or history.status == "error":
            return history
        prior_pagarme = self._prior_paid_charge_history(charge, history.operational_order)
        return prior_pagarme or history

    def _prior_paid_charge_history(self, charge: ChargeEvent, operational_order: MogoOrderSummary | None = None) -> CustomerHistoryResult | None:
        if not charge.identity_key:
            return None
        with self._connect() as conn:
            conn.row_factory = sqlite3.Row
            try:
                rows = list(
                    conn.execute(
                        """
                        SELECT e.charge_id, e.event_type, e.status, e.created_at, e.amount, e.customer_name,
                               COALESCE(r.decision, '') AS review_decision
                        FROM charge_events e
                        LEFT JOIN antifraud_reviews r ON r.charge_id = e.charge_id
                        WHERE e.identity_key = ?
                          AND e.charge_id != ?
                          AND e.created_at < ?
                        ORDER BY e.created_at ASC
                        """,
                        (charge.identity_key, charge.charge_id, charge.created_at.isoformat()),
                    )
                )
            except sqlite3.OperationalError:
                rows = list(
                    conn.execute(
                        """
                        SELECT charge_id, event_type, status, created_at, amount, customer_name,
                               '' AS review_decision
                        FROM charge_events
                        WHERE identity_key = ?
                          AND charge_id != ?
                          AND created_at < ?
                        ORDER BY created_at ASC
                        """,
                        (charge.identity_key, charge.charge_id, charge.created_at.isoformat()),
                    )
                )
        valid_paid_rows = self._valid_prior_paid_rows(rows)
        if not valid_paid_rows:
            return None
        latest = valid_paid_rows[-1]
        return CustomerHistoryResult(
            True,
            "pagarme_prior_charge",
            "valid_purchase",
            None,
            MogoOrderSummary(
                order_number=str(latest["charge_id"] or ""),
                status="paid",
                customer_name=str(latest["customer_name"] or ""),
                date=str(latest["created_at"] or ""),
                amount=str(latest["amount"] or ""),
                origin="Pagar.me",
            ),
            len(valid_paid_rows),
            operational_order,
        )

    def _valid_prior_paid_rows(self, rows: list[sqlite3.Row]) -> list[sqlite3.Row]:
        valid_paid_rows: list[sqlite3.Row] = []
        for row in rows:
            if self._is_reversal_history_row(row):
                self._remove_latest_reversed_paid_row(valid_paid_rows, row)
            elif self._is_paid_history_row(row):
                valid_paid_rows.append(row)
        return valid_paid_rows

    def _remove_latest_reversed_paid_row(self, valid_paid_rows: list[sqlite3.Row], reversal_row: sqlite3.Row) -> None:
        reversal_amount = int(reversal_row["amount"] or 0)
        for index in range(len(valid_paid_rows) - 1, -1, -1):
            paid_amount = int(valid_paid_rows[index]["amount"] or 0)
            if reversal_amount == 0 or paid_amount == reversal_amount:
                del valid_paid_rows[index]
                return

    def _is_paid_history_row(self, row: sqlite3.Row) -> bool:
        return str(row["event_type"] or "") == "charge.paid" or normalize_text(str(row["status"] or "")) == "paid"

    def _is_reversal_history_row(self, row: sqlite3.Row) -> bool:
        review_decision = normalize_text(str(row["review_decision"] or ""))
        if review_decision == "canceled":
            return True
        event_type = normalize_text(str(row["event_type"] or "")).replace(" ", ".")
        status = normalize_text(str(row["status"] or ""))
        reversal_events = {
            "order.canceled",
            "order.cancelled",
            "charge.canceled",
            "charge.cancelled",
            "charge.refunded",
            "charge.refund_succeeded",
        }
        reversal_statuses = {"canceled", "cancelled", "refunded", "voided", "estornado", "cancelado"}
        return event_type in reversal_events or status in reversal_statuses

    def _should_alert_first_purchase(
        self,
        charge: ChargeEvent,
        history: CustomerHistoryResult,
        antifraud_alert: bool,
    ) -> bool:
        if antifraud_alert:
            return False
        if not charge.is_card:
            return False
        if history.status != "not_found":
            return False
        if history.has_prior_valid_purchase:
            return False
        return True

    def _history_can_suppress_alerts(self, charge: ChargeEvent, history: CustomerHistoryResult) -> bool:
        if not history.has_prior_valid_purchase:
            return False
        if history.matched_by == "name_address":
            return True
        if history.matched_by in {"document", "email", "phone"}:
            return True
        if history.matched_by == "pagarme_prior_charge":
            return True
        if history.matched_by == "name" and history.valid_purchase_count >= 2:
            return True
        has_stronger_identity = bool(
            only_digits(charge.customer_document)
            or normalize_text(charge.customer_email)
            or only_digits(charge.customer_phone)
        )
        return not has_stronger_identity

    def _score_paid_charge(self, charge: ChargeEvent) -> tuple[int, list[str], CustomerHistoryResult]:
        hotlist_score = 0
        strong_score = 0
        weak_score = 0
        hotlist_reasons: list[str] = []
        strong_reasons: list[str] = []
        weak_reasons: list[str] = []
        operational_reasons: list[str] = []
        recent = self._recent_events(charge)
        failed = [row for row in recent if row["event_type"] == "charge.payment_failed" or row["status"] in ("failed", "not_authorized")]
        cards = {row["card_key"] for row in recent if row["card_key"]}
        if charge.card_key:
            cards.add(charge.card_key)

        history = self._customer_history(charge)
        suppress_weak = self._history_can_suppress_alerts(charge, history)
        suppress_checkout_retry = suppress_weak
        if history.status == "error":
            detail = f": {history.error}" if history.error else ""
            operational_reasons.append(f"Histórico Mogo não validado{detail}")

        exact_card_matches = self._exact_card_matches(charge)
        if exact_card_matches:
            identities = {str(row["identity_key"] or "") for row in exact_card_matches}
            count = len({identity for identity in identities if identity})
            suffix = f" ({count} cadastro(s) distinto(s))" if count else ""
            strong_score += 50
            strong_reasons.append(
                "Dados do cartão, bandeira e vencimento iguais em outro cadastro"
                f"{suffix}"
            )

        same_day_prefix_matches = [] if exact_card_matches else self._same_day_card_prefix_matches(charge)
        if same_day_prefix_matches:
            identities = {str(row["identity_key"] or "") for row in same_day_prefix_matches}
            count = len({identity for identity in identities if identity})
            suffix = f" ({count} cadastro(s) distinto(s))" if count else ""
            operational_reasons.append(
                "Aviso operacional: mesmos 6 primeiros dígitos e bandeira do cartão "
                f"já apareceram hoje em outro cadastro{suffix}; sinal auxiliar, não bloqueia sozinho"
            )

        if self.hotlist.matches(charge):
            hotlist_score += 50
            hotlist_reasons.append("Dado em lista quente de chargeback/fraude")

        known_fraud_address = _known_fraud_delivery_address(_context_order(history))
        if not known_fraud_address:
            known_fraud_address = _known_fraud_address_label(charge.card_billing_address)
        if known_fraud_address:
            strong_score += 50
            strong_reasons.append(f"Endereço com fraude anterior: {known_fraud_address}")

        customer_document = only_digits(charge.customer_document)
        holder_document = only_digits(charge.holder_document)
        if charge.is_card and not holder_document and not suppress_weak:
            strong_score += 50
            strong_reasons.append("CPF do titular do cartão ausente em pagamento de cartão")
        elif charge.is_card and customer_document and holder_document and customer_document != holder_document and not suppress_weak:
            strong_score += 50
            strong_reasons.append("CPF do cliente diferente do CPF do titular do cartão")

        if charge.holder_name and not names_compatible(charge.customer_name, charge.holder_name):
            if not customer_name_part_in_email_or_holder(charge.customer_name, charge.customer_email, charge.holder_name):
                weak_score += 50
                weak_reasons.append("Titular diferente do nome do cliente")

        if failed and not suppress_checkout_retry:
            strong_score += 50
            strong_reasons.append(f"Falha recente antes de pagamento aprovado ({len(failed)} em {WINDOW_LABEL})")

        if len(cards) >= 2 and not suppress_checkout_retry:
            strong_score += 50
            strong_reasons.append(f"2+ cartões diferentes no mesmo cliente/documento/email em {WINDOW_LABEL}")

        failed_amounts = [int(row["amount"] or 0) for row in failed]
        if failed_amounts and max(failed_amounts) > charge.amount and not suppress_checkout_retry:
            strong_score += 30
            strong_reasons.append("Valor maior falhou e valor menor foi aprovado")

        if charge.customer_email and charge.customer_name:
            email_user = normalize_text(charge.customer_email.split("@", 1)[0])
            name_tokens = set(normalize_text(charge.customer_name).split())
            if name_tokens and email_user and not any(token and token in email_user for token in name_tokens):
                weak_score += 20
                weak_reasons.append("Email pouco compatível com o nome do cliente")

        if history.matched_by == "name_address" and not strong_score:
            return hotlist_score, hotlist_reasons + operational_reasons, history
        if suppress_weak:
            return hotlist_score + strong_score, hotlist_reasons + strong_reasons + operational_reasons, history
        return hotlist_score + strong_score + weak_score, hotlist_reasons + strong_reasons + weak_reasons + operational_reasons, history


def _format_mogo_context(history: CustomerHistoryResult | None) -> str:
    if history is None:
        return "Mogo: histórico local não consultado"
    if history.status == "error":
        detail = f" ({history.error})" if history.error else ""
        return f"Mogo: histórico local não validado{detail}"
    if not history.has_prior_valid_purchase:
        return "Mogo: pedido/histórico local não localizado"

    order = history.order
    if order is None:
        return f"Mogo: histórico válido localizado por {history.matched_by or '-'}"

    lines = [f"Mogo: histórico válido localizado por {history.matched_by or '-'}"]
    fields = [
        ("Pedido", order.order_number),
        ("Status", order.status),
        ("Cliente Mogo", order.customer_name),
        ("Data", order.date),
        ("Entrega", " ".join(part for part in (order.delivery_date, order.delivery_time) if part)),
        ("Endereço", " - ".join(part for part in (order.address, order.neighborhood) if part)),
        ("Valor Mogo", order.amount),
        ("Origem", order.origin),
        ("Item", order.item),
    ]
    for label, value in fields:
        if value:
            lines.append(f"{label}: {value}")
    return "\n".join(lines)


def _format_brl(cents: int) -> str:
    return f"R$ {cents / 100:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _ensure_brl_text(value: str) -> str:
    value = value.strip()
    if not value:
        return value
    if value.lower().startswith("r$"):
        return value
    return f"R$ {value}"


def _format_brt_datetime(dt: datetime) -> str:
    brt = timezone(timedelta(hours=-3))
    return dt.astimezone(brt).strftime("%d/%m/%Y %H:%M")


def _alert_level(score: int) -> str:
    if score >= ALERT_THRESHOLD:
        return "FORTE"
    return "ATENÇÃO"


def _score_thermometer(score: int) -> str:
    if score >= ALERT_THRESHOLD:
        return "🔴 FORTE (50+ segura entrega)"
    if score >= 30:
        return "🟡 ATENÇÃO (30-49 revisar)"
    return "🟢 BAIXO (0-29 referência)"


def _score_header_line(score: int) -> str:
    return f"*Score antifraude: {score} — 🌡️ {_score_thermometer(score)}*"


def _format_document_tail(value: str | None) -> str:
    digits = only_digits(value)
    if not digits:
        return "-"
    return f"final {digits[-4:]}"


def _format_card_brand(value: str) -> str:
    value = (value or "").strip()
    return value[:1].upper() + value[1:].lower() if value else "-"


def _format_br_phone(value: str) -> str:
    digits = only_digits(value)
    if digits.startswith("55") and len(digits) in {12, 13}:
        digits = digits[2:]
    if len(digits) == 11:
        return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    return value or "-"


def _format_cpf(value: str) -> str:
    digits = only_digits(value)
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return value or "-"


def _has_hotlist_reason(reasons: list[str]) -> bool:
    return any("lista quente" in reason.lower() for reason in reasons)


def _mogo_order_header(history: CustomerHistoryResult | None) -> str:
    order = history.order if history else None
    if order and order.order_number:
        return f"🧾 HISTÓRICO MOGO: pedido #{order.order_number}"
    return "🧾 HISTÓRICO MOGO: não localizado"


def _mogo_customer_lines(history: CustomerHistoryResult | None, charge: ChargeEvent) -> list[str]:
    order = _context_order(history)
    return [
        "Cliente no Mogo",
        f"• Nome: {(order.customer_name if order and order.customer_name else charge.customer_name) or '-'}",
        f"• Telefone/WhatsApp: {_format_br_phone(order.phone if order and order.phone else charge.customer_phone)}",
        f"• CPF do pedido: {_format_cpf(order.document if order and order.document else charge.customer_document)}",
        f"• Email: {(order.email if order and order.email else charge.customer_email) or '-'}",
    ]


def _format_order_address(order: MogoOrderSummary) -> str:
    return " - ".join(part for part in (order.address, order.neighborhood) if part)


def _context_order(history: CustomerHistoryResult | None) -> MogoOrderSummary | None:
    if not history:
        return None
    return history.operational_order or history.order


def _order_modality(order: MogoOrderSummary | None) -> str:
    if not order:
        return "não localizada no alerta"
    raw = normalize_text(" ".join(part for part in (order.fulfillment, order.origin) if part))
    if "retir" in raw or "buscar" in raw or "balcao" in raw:
        return "Retirada"
    if "entreg" in raw or "delivery" in raw or "motoboy" in raw or _format_order_address(order):
        return "Entrega"
    return order.fulfillment or "não localizada no alerta"


def _order_schedule(order: MogoOrderSummary | None) -> str:
    if not order:
        return "não localizado no alerta"
    delivery_date = (order.delivery_date or "").strip()
    delivery_time = (order.delivery_time or "").strip()
    if delivery_date and delivery_time:
        return f"{delivery_date} {delivery_time}"
    if delivery_date:
        return f"{delivery_date} — sem hora agendada — ⚠️ tratar como para agora"
    if delivery_time:
        return f"sem data agendada, hora {delivery_time} — ⚠️ revisar antes de liberar"
    return "não informado — ⚠️ tratar como para agora"


def _fulfillment_lines(history: CustomerHistoryResult | None) -> list[str]:
    order = _context_order(history)
    modality = _order_modality(order)
    lines = [
        "Pedido",
        f"• Modalidade: {modality}",
        f"• Agendamento: {_order_schedule(order)}",
    ]
    if modality == "Entrega":
        address = _format_order_address(order) if order else ""
        lines.append(f"• Endereço de entrega: {address or 'não localizado no alerta'}")
    return lines


def _mogo_order_lines(history: CustomerHistoryResult | None) -> list[str]:
    if history is None:
        return ["Histórico Mogo", "• Histórico local: não consultado"]
    if history.status == "error":
        detail = f" ({history.error})" if history.error else ""
        return ["Histórico Mogo", f"• Histórico local: não validado{detail}"]
    if not history.has_prior_valid_purchase:
        lines = ["Histórico Mogo", "• Histórico local: pedido/histórico não localizado"]
        if history.operational_order:
            suffix = f" #{history.operational_order.order_number}" if history.operational_order.order_number else ""
            lines.append(f"• Pedido operacional:{suffix} localizado no Mogo")
        return lines

    order = history.order
    count_text = f" ({history.valid_purchase_count} compra(s) válida(s))" if history.valid_purchase_count else ""
    lines = ["Histórico Mogo", f"• Histórico local: localizado por {history.matched_by or '-'}{count_text}"]
    if not order:
        return lines

    fields = [
        ("Status Mogo", order.status),
        ("Data Mogo", order.date),
        ("Entrega", " ".join(part for part in (order.delivery_date, order.delivery_time) if part)),
        ("Endereço", " - ".join(part for part in (order.address, order.neighborhood) if part)),
        ("Valor Mogo", order.amount),
        ("Origem Mogo", order.origin),
        ("Item/produto", order.item),
    ]
    for label, value in fields:
        if value:
            lines.append(f"• {label}: {value}")
    return lines


def _related_profiles_lines(history: CustomerHistoryResult | None) -> list[str]:
    profiles = tuple(getattr(history, "related_profiles", ()) or ())[:3]
    if not profiles:
        return []
    lines = ["Cadastros possivelmente relacionados"]
    for index, profile in enumerate(profiles, start=1):
        lines.extend([
            f"• Cadastro {index}: {profile.match_reason}",
            f"  Nome: {profile.name or '-'}",
            f"  Telefone: {profile.phone or '-'}",
            f"  Email: {profile.email or '-'}",
            f"  Documento: {profile.document or '-'}",
            f"  Endereço: {profile.address or '-'}",
            f"  Última compra válida: {profile.last_purchase_date or '-'} — {profile.last_purchase_amount or '-'}",
            f"  Compras válidas: {profile.valid_purchase_count}",
        ])
    lines.append("• Informação auxiliar: não altera score nem decisão operacional.")
    return lines


def format_alert(result: RiskResult) -> str:
    charge = result.charge
    history = result.customer_history
    value_line = _format_brl(charge.amount)
    reasons = "\n".join(f"• {reason}" for reason in result.reasons) or "• Sem motivo detalhado"
    lines = [
        "🚨 POSSÍVEL FRAUDE — SEGURAR ENTREGA",
        "",
        *([
            "🚨🚨🚨 ATENÇÃO: JÁ CONSTA EM LISTA DE FRAUDADORES ANTERIORES 🚨🚨🚨",
            "⚠️ HISTÓRICO ANTERIOR DE CHARGEBACK/FRAUDE — NÃO VENDER / NÃO LIBERAR SEM AUTORIZAÇÃO EXPRESSA DO ZÃO",
            "",
        ] if _has_hotlist_reason(result.reasons) else []),
        _mogo_order_header(history),
        "Status operacional: SEGURAR / NÃO ENTREGAR",
        "",
        _score_header_line(result.score),
        "",
        *_fulfillment_lines(history),
        "",
        "Resumo",
        f"• Valor do pedido: {value_line}",
        f"• Data/hora: {_format_brt_datetime(charge.created_at)}",
        f"• Nível do alerta: {_alert_level(result.score)}",
        "",
        *_mogo_customer_lines(history, charge),
        "",
        *_mogo_order_lines(history),
        *(["", *_related_profiles_lines(history)] if _related_profiles_lines(history) else []),
        "",
        "Pagamento Pagar.me",
        f"• Cliente Pagar.me: {charge.customer_name or '-'}",
        f"• Email Pagar.me: {charge.customer_email or '-'}",
        f"• Documento do cliente Pagar.me: {charge.customer_document or '-'}",
        f"• Valor Pagar.me: {_format_brl(charge.amount)}",
        f"• Cartão: {charge.card_brand or '-'} final {charge.card_last4 or '-'}",
        f"• Titular do cartão: {charge.holder_name or '-'}",
        f"• Documento do titular do cartão: {_format_document_tail(charge.holder_document)}",
        f"• Score: {result.score}",
        "",
        "Motivos do alerta",
        reasons,
        "",
        (
            "Ação: NÃO VENDER / NÃO LIBERAR sem autorização expressa do Zão. Não acusar fraude."
            if _has_hotlist_reason(result.reasons)
            else "Ação: falar com o cliente antes de liberar entrega. Não acusar fraude."
        ),
    ]
    return "\n".join(lines)


def _same_day_purchase_lines(context: SameDayRepeatContext) -> list[str]:
    brt = timezone(timedelta(hours=-3))
    return [
        f"• {index}ª compra — {purchase.created_at.astimezone(brt).strftime('%H:%M')} — {_format_brl(purchase.amount)}"
        for index, purchase in enumerate(context.purchases, start=1)
    ]


def format_same_day_repeat_alert(result: RiskResult) -> str:
    context = result.same_day_repeat
    if context is None:
        raise ValueError("same-day repeat context is required")
    charge = result.charge
    return "\n".join([
        "🚨🚨 ALERTA MUITO CRÍTICO — RECOMPRA NO DIA DA PRIMEIRA COMPRA",
        "",
        "Status operacional: SEGURAR / NÃO ENTREGAR",
        "",
        f"• Cliente: {charge.customer_name or '-'}",
        f"• Data: {charge.created_at.astimezone(timezone(timedelta(hours=-3))).strftime('%d/%m/%Y')}",
        f"• Compra atual: {context.sequence}ª compra aprovada no dia",
        "• Situação: a primeira compra da vida ocorreu hoje e o cliente voltou a comprar no mesmo dia",
        "",
        "Compras aprovadas hoje",
        *_same_day_purchase_lines(context),
        "",
        "Ação: conferir todas as compras com o cliente antes de liberar. Não acusar fraude.",
    ])


def format_same_day_repeat_notice(result: RiskResult) -> str:
    context = result.same_day_repeat
    if context is None:
        raise ValueError("same-day repeat context is required")
    charge = result.charge
    return "\n".join([
        "ℹ️ AVISO INFORMATIVO — CLIENTE COM MAIS DE UMA COMPRA NO DIA",
        "",
        "Status operacional: NÃO SEGURA ENTREGA",
        "",
        f"• Cliente: {charge.customer_name or '-'}",
        f"• Data: {charge.created_at.astimezone(timezone(timedelta(hours=-3))).strftime('%d/%m/%Y')}",
        f"• Compra atual: {context.sequence}ª compra aprovada no dia",
        "• Situação: cliente antigo com múltiplas compras no mesmo dia",
        "",
        "Compras aprovadas hoje",
        *_same_day_purchase_lines(context),
        "",
        "Ação: aviso para conferência rápida de possível duplicidade ou erro do cliente. Não bloqueia a entrega.",
    ])


def format_first_purchase_alert(result: RiskResult) -> str:
    charge = result.charge
    history = result.customer_history
    order = _context_order(history)
    modality = _order_modality(order)
    is_pickup = modality == "Retirada"
    header_action = "CONFERIR NA RETIRADA" if is_pickup else "CONFERIR ANTES DE ENTREGAR"
    raw_address_line = _format_order_address(order) if order else ""
    address_line = "não aplicável — retirada na loja" if is_pickup else (raw_address_line or "não localizado no alerta")
    customer_name = (order.customer_name if order and order.customer_name else charge.customer_name) or "-"
    order_lines = [
        f"• Cliente: {customer_name}",
        f"• Modalidade: {modality}",
    ]
    if not is_pickup:
        order_lines.append(f"• Agendamento: {_order_schedule(order)}")
        order_lines.append(f"• Endereço de entrega: {address_line}")
    order_lines.extend([
        f"• Valor: {_format_brl(charge.amount)}",
        "• Pagamento: cartão online aprovado",
        "• Antifraude Pagar.me: sem alerta",
    ])
    advisory_reasons = list(getattr(result, "reasons", []) or [])
    advisory_lines = (
        [
            "",
            "Avisos operacionais",
            *[f"• {reason}" for reason in advisory_reasons],
        ]
        if advisory_reasons else []
    )
    lines = [
        f"🟡 PRIMEIRA COMPRA — {header_action}",
        "",
        "Status operacional: NÃO LIBERAR SEM CONFERÊNCIA",
        "",
        "Pedido",
        *order_lines,
        "",
        "Pagamento",
        f"• Cartão: {_format_card_brand(charge.card_brand)} final {charge.card_last4 or '-'}",
        f"• Titular do cartão: {charge.holder_name or '-'}",
        "• Status: aprovado",
        "",
        "Histórico Mogo",
        "• CPF: não localizado em compra anterior",
        "• Telefone: não localizado em compra anterior",
        "• Email: não localizado em compra anterior",
        "• Nome: não localizado em compra anterior confiável",
        f"• Endereço: {address_line}",
        *(["", *_related_profiles_lines(history)] if _related_profiles_lines(history) else []),
        "",
        "Resumo",
        "• Situação: possível primeira compra",
        "• Risco: baixo/médio",
        "• Motivo: pedido aprovado sem histórico confiável no Mogo",
        *advisory_lines,
        "",
        "Ação da equipe",
        "• Conferir documento do comprador na retirada." if is_pickup else "• Conferir documento do comprador antes de liberar.",
        "• Se outra pessoa for retirar, pedir autorização do comprador." if is_pickup else "• Se houver terceiro envolvido, pedir autorização do comprador.",
        "• Não acusar fraude. Tratar como procedimento padrão de primeira compra.",
    ]
    return "\n".join(lines)
