#!/usr/bin/env python3
"""
BigDog — Mogo "Finalizados • Média de Produtos por Pedido"

Gera um relatório customizado a partir de:
- PDV > Entregas > Finalizados
- Filtro de período
- Tipo de data configurável (padrão: Pedido)
- Origem configurável (padrão: Neemo)

Saídas:
  /root/workspaces/cake-brain/relatorios/Mogo/Finalizados Media Produtos/<arquivo>.xlsx
  /root/workspaces/cake-brain/relatorios/Mogo/Finalizados Media Produtos/<arquivo>.json

Exemplo:
  python3 mogo-finalizados-media-produtos.py \
    --from 01/12/2025 \
    --to 23/04/2026 \
    --origem Neemo \
    --tipo-data Pedido \
    --json
"""

from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
import threading
import time

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import requests

sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL


OUTPUT_DIR = Path('/root/workspaces/cake-brain/relatorios/Mogo/Finalizados Media Produtos')


def parse_date(value: str) -> datetime:
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    raise SystemExit(f'Data inválida: {value}. Use DD/MM/YYYY, DD-MM-YYYY ou YYYY-MM-DD.')


def daterange(start: datetime, end: datetime):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def clean_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def parse_decimal(value) -> float:
    if value in (None, ''):
        return 0.0
    text = str(value).strip().replace('.', '').replace(',', '.')
    try:
        return float(text)
    except ValueError:
        return 0.0


def format_currency(value: float) -> str:
    return f'R$ {value:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def extract_origin(html: str) -> str:
    if not html:
        return ''

    title_match = re.search(r"title=['\"]?([^'\"\s>]+)", str(html))
    if title_match:
        return title_match.group(1).strip()

    text = re.sub(r'<[^>]+>', ' ', str(html))
    return re.sub(r'\s+', ' ', text).strip()


def normalize(value: str) -> str:
    return clean_text(value).casefold()


def month_key(date_str: str) -> str:
    dt = parse_date(date_str)
    return dt.strftime('%Y-%m')


def month_label(key: str) -> str:
    dt = datetime.strptime(key, '%Y-%m')
    return dt.strftime('%m/%Y')


def fetch_orders_for_day(session, day: datetime, tipo_data: str) -> list[dict]:
    day_str = day.strftime('%d/%m/%Y')
    page = 1
    rows = 1000
    collected = []

    while True:
        response = session.post(
            f'{MOGO_URL}/Pedido/ListPedidosParaEntrega',
            params={
                'cFiltroTipoEntrega': '2',
                'dtDe': day_str,
                'dtAte': day_str,
                'tipoFiltroData': tipo_data,
            },
            data={
                '_search': 'true',
                'nd': '1',
                'rows': str(rows),
                'page': str(page),
                'sidx': 'HoraInclusao',
                'sord': 'desc',
                'totalrows': '',
            },
            timeout=60,
        )
        response.raise_for_status()
        data = response.json()
        batch = data.get('rows', []) or []
        collected.extend(batch)

        total_records = int(data.get('records') or 0)
        if len(batch) < rows or len(collected) >= total_records:
            break
        page += 1

    return collected


def fetch_order_items(session, order_id: int | str) -> list[dict]:
    response = session.get(
        f'{MOGO_URL}/Pedido/DadosItemPedido',
        params={'nId': order_id},
        timeout=60,
    )
    response.raise_for_status()
    data = response.json()
    return data.get('data', []) or []


def build_item_session_factory(source_session):
    base_headers = dict(source_session.headers)
    base_cookies = source_session.cookies.get_dict()
    local = threading.local()

    def get_session():
        session = getattr(local, 'session', None)
        if session is None:
            session = requests.Session()
            session.headers.update(base_headers)
            for key, value in base_cookies.items():
                session.cookies.set(key, value)
            local.session = session
        return session

    return get_session


def summarize_items(items: list[dict]) -> tuple[float, int, str, list[dict]]:
    total_quantity = 0.0
    item_lines = 0
    parts = []
    raw_items = []

    for item in items:
        product = clean_text(item.get('Produto_Nome') or item.get('Descricao'))
        qty = parse_decimal(item.get('Quantidade'))
        total_value = parse_decimal(item.get('ValorTotal'))
        total_quantity += qty
        item_lines += 1
        raw_items.append({
            'produto': product,
            'quantidade': qty,
            'valor_total': total_value,
        })
        qty_label = f'{qty:g}x ' if qty else ''
        parts.append(f'{qty_label}{product}')

    return total_quantity, item_lines, ' | '.join(parts), raw_items


def auto_fit(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width


def style_header(ws, headers: list[str], color: str = '1F4E78') -> None:
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def save_workbook(summary_rows: list[list], order_rows: list[list], item_rows: list[list], output_path: Path) -> None:
    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = 'Resumo Mensal'
    summary_headers = [
        'Mês', 'Pedidos', 'Produtos Totais', 'Média Produtos/Pedido',
        'Linhas de Itens', 'Valor Total', 'Ticket Médio'
    ]
    style_header(ws_summary, summary_headers, color='2E7D32')
    for row_idx, row in enumerate(summary_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_summary.cell(row=row_idx, column=col_idx, value=value)
    auto_fit(ws_summary, [12, 12, 16, 22, 16, 14, 14])

    ws_orders = wb.create_sheet('Pedidos')
    order_headers = [
        'Mês', 'Nº Pedido', 'Data Pedido', 'Data Entrega', 'Cliente', 'Origem',
        'Valor Final', 'Produtos no Pedido', 'Linhas de Itens', 'Itens'
    ]
    style_header(ws_orders, order_headers)
    for row_idx, row in enumerate(order_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_orders.cell(row=row_idx, column=col_idx, value=value)
    auto_fit(ws_orders, [12, 12, 14, 14, 28, 14, 12, 18, 16, 100])

    ws_items = wb.create_sheet('Itemizado')
    item_headers = [
        'Mês', 'Nº Pedido', 'Data Pedido', 'Cliente', 'Origem', 'Produto', 'Qtd', 'Valor Total'
    ]
    style_header(ws_items, item_headers)
    for row_idx, row in enumerate(item_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_items.cell(row=row_idx, column=col_idx, value=value)
    auto_fit(ws_items, [12, 12, 14, 28, 14, 40, 10, 12])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description='Analisa pedidos finalizados do Mogo e calcula média de produtos por pedido.')
    parser.add_argument('--from', dest='date_from', required=True, help='Data inicial. Aceita DD/MM/YYYY, DD-MM-YYYY ou YYYY-MM-DD.')
    parser.add_argument('--to', dest='date_to', required=True, help='Data final. Aceita DD/MM/YYYY, DD-MM-YYYY ou YYYY-MM-DD.')
    parser.add_argument('--origem', default='Neemo', help='Origem a filtrar. Padrão: Neemo. Use vazio para não filtrar.')
    parser.add_argument('--tipo-data', default='Pedido', help='Tipo de data do filtro Finalizados. Padrão: Pedido.')
    parser.add_argument('--workers', type=int, default=8, help='Número de workers para buscar itens dos pedidos. Padrão: 8.')
    parser.add_argument('--json', action='store_true', help='Salva também o JSON bruto.')
    args = parser.parse_args()

    start = parse_date(args.date_from)
    end = parse_date(args.date_to)
    if end < start:
        raise SystemExit('Data final não pode ser menor que a inicial.')

    origin_filter = normalize(args.origem)
    session = mogo_login()

    print(
        'Buscando Finalizados:',
        start.strftime('%d/%m/%Y'), '→', end.strftime('%d/%m/%Y'),
        f'| tipo-data={args.tipo_data}',
        f'| origem={args.origem or "(todas)"}'
    )

    filtered_orders = []

    for day in daterange(start, end):
        attempts = 0
        while True:
            try:
                orders = fetch_orders_for_day(session, day, args.tipo_data)
                break
            except Exception as exc:
                attempts += 1
                if attempts >= 4:
                    raise
                wait_seconds = attempts * 2
                print(
                    f'  {day.strftime("%d/%m/%Y")}: falha ao buscar ({exc.__class__.__name__}) '
                    f'| retry {attempts}/3 em {wait_seconds}s'
                )
                time.sleep(wait_seconds)
                session = mogo_login()

        selected = []
        for order in orders:
            origin = extract_origin(order.get('OrigemPedido', ''))
            if origin_filter and normalize(origin) != origin_filter:
                continue
            selected.append(order)
        filtered_orders.extend(selected)
        print(f'  {day.strftime("%d/%m/%Y")}: {len(orders)} finalizados | {len(selected)} filtrados')

    if not filtered_orders:
        print('Nenhum pedido encontrado com esse filtro.')
        return

    monthly = defaultdict(lambda: {
        'pedidos': 0,
        'produtos_total': 0.0,
        'linhas_itens': 0,
        'valor_total': 0.0,
    })

    order_rows = []
    item_rows = []
    raw_orders = []

    total_orders = len(filtered_orders)
    print(f'Pedidos após filtro: {total_orders}')

    item_session_factory = build_item_session_factory(session)

    def process_order(payload):
        idx, order = payload
        order_id = order.get('Id')
        number = order.get('NumeroPedido')
        data_pedido = clean_text(order.get('DataPedido'))
        data_entrega = clean_text(order.get('DataEntrega'))
        origem = extract_origin(order.get('OrigemPedido', ''))
        cliente = clean_text(order.get('NomeCliente'))
        valor_final = parse_decimal(order.get('ValorFinal'))
        key = month_key(data_pedido or data_entrega)

        attempts = 0
        item_session = None
        while True:
            try:
                item_session = item_session or item_session_factory()
                items = fetch_order_items(item_session, order_id)
                break
            except Exception as exc:
                attempts += 1
                if attempts >= 4:
                    raise
                wait_seconds = attempts * 2
                print(
                    f'  Pedido {number}: falha ao buscar itens ({exc.__class__.__name__}) '
                    f'| retry {attempts}/3 em {wait_seconds}s'
                )
                time.sleep(wait_seconds)
                item_session = mogo_login()
        produtos_total, linhas_itens, itens_resumo, raw_items = summarize_items(items)

        item_row_batch = []
        for item in raw_items:
            item_row_batch.append([
                month_label(key),
                number,
                data_pedido,
                cliente,
                origem,
                item['produto'],
                item['quantidade'],
                item['valor_total'],
            ])

        return idx, {
            'key': key,
            'number': number,
            'data_pedido': data_pedido,
            'data_entrega': data_entrega,
            'cliente': cliente,
            'origem': origem,
            'valor_final': valor_final,
            'produtos_total': produtos_total,
            'linhas_itens': linhas_itens,
            'itens_resumo': itens_resumo,
            'raw_items': raw_items,
            'item_rows': item_row_batch,
            'order_id': order_id,
        }

    results = [None] * total_orders
    workers = max(1, min(args.workers, total_orders))
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(process_order, payload): payload[0]
            for payload in enumerate(filtered_orders, start=1)
        }
        completed = 0
        for future in as_completed(future_map):
            idx, result = future.result()
            results[idx - 1] = result
            completed += 1
            if completed % 25 == 0 or completed == total_orders:
                print(f'  Itens carregados: {completed}/{total_orders}')

    for result in results:
        key = result['key']
        monthly[key]['pedidos'] += 1
        monthly[key]['produtos_total'] += result['produtos_total']
        monthly[key]['linhas_itens'] += result['linhas_itens']
        monthly[key]['valor_total'] += result['valor_final']

        order_rows.append([
            month_label(key),
            result['number'],
            result['data_pedido'],
            result['data_entrega'],
            result['cliente'],
            result['origem'],
            result['valor_final'],
            result['produtos_total'],
            result['linhas_itens'],
            result['itens_resumo'],
        ])

        item_rows.extend(result['item_rows'])

        raw_orders.append({
            'mes': month_label(key),
            'numero_pedido': result['number'],
            'pedido_id': result['order_id'],
            'data_pedido': result['data_pedido'],
            'data_entrega': result['data_entrega'],
            'cliente': result['cliente'],
            'origem': result['origem'],
            'valor_final': result['valor_final'],
            'produtos_no_pedido': result['produtos_total'],
            'linhas_itens': result['linhas_itens'],
            'itens': result['raw_items'],
        })

    summary_rows = []
    total_pedidos = 0
    total_produtos = 0.0
    total_linhas = 0
    total_valor = 0.0

    for key in sorted(monthly.keys()):
        pedidos = monthly[key]['pedidos']
        produtos = monthly[key]['produtos_total']
        linhas = monthly[key]['linhas_itens']
        valor = monthly[key]['valor_total']
        media = (produtos / pedidos) if pedidos else 0.0
        ticket = (valor / pedidos) if pedidos else 0.0
        summary_rows.append([
            month_label(key),
            pedidos,
            produtos,
            media,
            linhas,
            valor,
            ticket,
        ])
        total_pedidos += pedidos
        total_produtos += produtos
        total_linhas += linhas
        total_valor += valor

    media_total = (total_produtos / total_pedidos) if total_pedidos else 0.0
    ticket_total = (total_valor / total_pedidos) if total_pedidos else 0.0
    summary_rows.append([
        'TOTAL',
        total_pedidos,
        total_produtos,
        media_total,
        total_linhas,
        total_valor,
        ticket_total,
    ])

    slug_origin = re.sub(r'[^a-z0-9]+', '-', normalize(args.origem) or 'todas').strip('-') or 'todas'
    filename_base = f"{start.strftime('%d-%m-%Y')}_a_{end.strftime('%d-%m-%Y')}_{slug_origin}"

    output_xlsx = OUTPUT_DIR / f'{filename_base}.xlsx'
    save_workbook(summary_rows, order_rows, item_rows, output_xlsx)
    print(f'Excel salvo: {output_xlsx}')

    if args.json:
        output_json = OUTPUT_DIR / f'{filename_base}.json'
        output_json.write_text(json.dumps({
            'filtro': {
                'data_de': start.strftime('%d/%m/%Y'),
                'data_ate': end.strftime('%d/%m/%Y'),
                'tipo_data': args.tipo_data,
                'origem': args.origem,
            },
            'resumo_mensal': [
                {
                    'mes': row[0],
                    'pedidos': row[1],
                    'produtos_totais': row[2],
                    'media_produtos_por_pedido': row[3],
                    'linhas_itens': row[4],
                    'valor_total': row[5],
                    'ticket_medio': row[6],
                }
                for row in summary_rows
            ],
            'pedidos': raw_orders,
        }, ensure_ascii=False, indent=2))
        print(f'JSON salvo: {output_json}')

    print('\nResumo total')
    print('-----------')
    print(f'Pedidos: {total_pedidos}')
    print(f'Produtos totais: {total_produtos:g}')
    print(f'Média produtos/pedido: {media_total:.2f}')
    print(f'Valor total: {format_currency(total_valor)}')
    print(f'Ticket médio: {format_currency(ticket_total)}')


if __name__ == '__main__':
    main()
