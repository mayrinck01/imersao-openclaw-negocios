#!/usr/bin/env python3
"""
BigDog — Mogo "Venda em Nota Assinada" (Mensal)
Roda todo dia 1 de cada mês às 07:00 BRT.
Relatório referente ao mês anterior.

Endpoint: GET /Financeiro/BillsToReceiveJqGrid
Filtro: Descrição = "Venda em nota assinada", Data Emissão
Email: joao@cakeco.com.br
"""

import sys, os, subprocess, json, re, calendar, argparse
from datetime import datetime
from pathlib import Path

sys.path.insert(0, '/root/workspaces/cake-brain/automacoes/scripts')
from mogo_login import mogo_login, MOGO_URL
from filter_pt import month_year_pt
from mogo_excel import format_currency_cells

parser = argparse.ArgumentParser()
parser.add_argument('mes', nargs='?', type=int, default=None)
parser.add_argument('ano', nargs='?', type=int, default=None)
parser.add_argument('--no-email', action='store_true')
args = parser.parse_args()

# ── Período: mês anterior (ou args) ──────────────────────────────────────────

hoje = datetime.now()
default_mes  = hoje.month - 1 if hoje.month > 1 else 12
default_ano  = hoje.year if hoje.month > 1 else hoje.year - 1
mes  = args.mes or default_mes
ano  = args.ano or default_ano
ultimo_dia = calendar.monthrange(ano, mes)[1]

data_de_br  = f"01/{mes:02d}/{ano}"
data_ate_br = f"{ultimo_dia:02d}/{mes:02d}/{ano}"
mes_ref     = f"{ano:04d}-{mes:02d}"
mes_nome = month_year_pt(datetime(ano, mes, 1))

print(f"Venda em Nota Assinada: {data_de_br} → {data_ate_br}")

# ── Diretório de saída ────────────────────────────────────────────────────────

OUTPUT_DIR = Path('/root/workspaces/cake-brain/relatorios/Mogo/Venda em Nota Assinada')
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
xlsx_path = OUTPUT_DIR / f'{mes_ref}-venda-nota-assinada.xlsx'

# ── Login ─────────────────────────────────────────────────────────────────────

session = mogo_login()

# ── Buscar dados ──────────────────────────────────────────────────────────────

print("Buscando Contas a Receber...")
params = {
    '_search': 'true', 'nd': '1', 'rows': '5000', 'page': '1',
    'sidx': '', 'sord': 'asc',
    # BillsToReceiveJqGrid respeita os nomes usados na tela do Mogo.
    # Quando enviamos DataDe/DataAte/TipoData, ele cai no padrão por vencimento.
    'dataDe': data_de_br,
    'dataAte': data_ate_br,
    'selDate': 'emissao',
    'inpPesqCrD': 'venda em nota assinada',
    'chkVencidos': 'on',
    'chkReceber': 'on',
    'chkRecebido': 'on',
    'chkCredito': 'on',
    'chkDebito': 'on',
    'chkCaixaAVista': 'false',
    'chkTituloComp': 'false',
    'validaConciliacao': '-1',
}

r = session.get(f"{MOGO_URL}/Financeiro/BillsToReceiveJqGrid", params=params, timeout=120)
if r.status_code != 200:
    print(f"ERRO {r.status_code}: {r.text[:200]}")
    sys.exit(1)

all_rows = r.json().get('rows', [])
print(f"Total registros: {len(all_rows)}")

# ── Filtrar "Venda em nota assinada" ─────────────────────────────────────────

def strip_html(text):
    if not text: return ''
    return re.sub(r'<[^>]+>', '', str(text)).strip()


def parse_br_date(text):
    text = strip_html(text)
    if not text:
        return None
    try:
        return datetime.strptime(text, '%d/%m/%Y')
    except ValueError:
        return None

rows_descricao = [row for row in all_rows
                  if 'venda em nota assinada' in strip_html(row.get('Descricao', '')).lower()]

rows = []
rows_fora_periodo = []
for row in rows_descricao:
    dt_emissao = parse_br_date(row.get('EntradaCompetencia', ''))
    if dt_emissao and dt_emissao.month == mes and dt_emissao.year == ano:
        rows.append(row)
    else:
        rows_fora_periodo.append(row)

print(f"Venda em nota assinada (descrição): {len(rows_descricao)} registros")
if rows_fora_periodo:
    print(f"⚠️  Descartados {len(rows_fora_periodo)} registros fora do mês de emissão {mes:02d}/{ano}")
print(f"Venda em nota assinada (emissão validada): {len(rows)} registros")

if not rows:
    print("Nenhum registro encontrado após validar Data de Emissão.")
    sys.exit(1)

# ── Gerar Excel ───────────────────────────────────────────────────────────────

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Mesmo padrão do export bruto do Mogo (exemplo enviado pelo Zão)
COLUNAS = [
    ('Cliente_Nome', 'Cliente_Nome'),
    ('Descricao', 'Descricao'),
    ('Vencimento', 'Vencimento'),
    ('Valor', 'Valor'),
    ('Saldo', 'Saldo'),
    ('ValorRecebido', 'ValorRecebido'),
    ('EntradaCompetencia', 'EntradaCompetencia'),
    ('Historico', 'Historico'),
    ('PlanoDeContas_Descricao', 'PlanoDeContas_Descricao'),
    ('Solicitante', 'Solicitante'),
    ('ContaBancaria_Id', 'ContaBancaria_Id'),
    ('ContaBancaria_Descricao', 'ContaBancaria_Descricao'),
    ('DataDeRecebimento', 'DataDeRecebimento'),
    ('ReceberComCredito', 'ReceberComCredito'),
    ('RazaoSocial', 'RazaoSocial'),
    ('Observacao', 'Observacao'),
    ('Conciliado', 'Conciliado'),
    ('NotaFiscal', 'NotaFiscal'),
    ('NumeroNF', 'NumeroNF'),
    ('Itens', 'Itens'),
    ('IsService', 'IsService'),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Planilha'

hfill = PatternFill(start_color='1a6b3c', end_color='1a6b3c', fill_type='solid')
hfont = Font(color='FFFFFF', bold=True, size=9)
for c_idx, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c_idx, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, row in enumerate(rows, 2):
    for c_idx, (field, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=strip_html(row.get(field, '') or ''))

for idx, (_, header) in enumerate(COLUNAS, 1):
    max_len = len(header)
    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
        val = '' if row[0] is None else str(row[0])
        max_len = max(max_len, len(val))
    ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Enviar email (opcional) ──────────────────────────────────────────────────

if not args.no_email:
    def parse_val(v):
        try:
            return float(str(v).replace('.','').replace(',','.').strip())
        except:
            return 0.0

    def fmt_brl(v):
        return f"R$ {v:,.2f}".replace(',','X').replace('.',',').replace('X','.')

    total_valor = sum(parse_val(strip_html(r.get('Valor', ''))) for r in rows)
    total_recebido = sum(parse_val(strip_html(r.get('ValorRecebido', ''))) for r in rows)
    total_saldo = sum(parse_val(strip_html(r.get('Saldo', ''))) for r in rows)

    op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
    env = os.environ.copy()
    env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

    assunto = f"📋 Venda em Nota Assinada — {len(rows)} registros — {fmt_brl(total_valor)} — {mes_nome}"
    corpo = f"""Relatório Venda em Nota Assinada — {mes_nome}

📅 Período: {data_de_br} → {data_ate_br}
🗂️ Tipo de data: Emissão
📌 Filtro: Descrição = Venda em nota assinada
📦 Registros: {len(rows)}

💰 Valor total:     {fmt_brl(total_valor)}
✅ Valor recebido:  {fmt_brl(total_recebido)}
⏳ Saldo em aberto: {fmt_brl(total_saldo)}

(BigDog 🐕)"""

    cmd = [
        'bash', '-c',
        f'GOG_KEYRING_PASSWORD="" gog gmail send '
        f'--account cakebigdog@gmail.com '
        f'--client cakebigdog '
        f'--to joao@cakeco.com.br '
        f'--subject "{assunto}" '
        f'--body "{corpo}" '
        f'--attach "{xlsx_path}"'
    ]
    res = subprocess.run(cmd, capture_output=True, text=True, env=env)
    if 'message_id' in res.stdout:
        print(f"✅ Email enviado para joao@cakeco.com.br")
    else:
        print(f"ERRO email: {res.stderr[:200]}")

print("\n✅ Relatório concluído")
