#!/usr/bin/env python3
"""
BigDog — Mogo Contas a Receber (Vendas em Nota Assinada)
Relatório: Financeiro > Contas a Receber
Filtro: Descrição = "venda em nota assinada"
Período: mês anterior (emissão)
Agendamento: dia 1º de cada mês, 07:00 BRT
Email: joao@cakeco.com.br
"""

import os, sys, subprocess, time, json, calendar
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, '/root/workspaces/cake-brain/automacoes/scripts')
from mogo_login import mogo_login
from mogo_excel import format_currency_cells
import requests

# ── Configuração ─────────────────────────────────────────────────────────────

MOGO_BASE_URL = "https://app3.mogogourmet.com.br"
OUTPUT_DIR = Path("/root/workspaces/cake-brain/relatorios/Mogo/ContasAssinada")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Período: mês anterior ────────────────────────────────────────────────────

hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]

data_de = f"{ano_ant}-{mes_ant:02d}-01"
data_ate = f"{ano_ant}-{mes_ant:02d}-{ultimo_dia:02d}"

# Formato DD/MM/YYYY para envio ao Mogo
data_de_br = f"01/{mes_ant:02d}/{ano_ant}"
data_ate_br = f"{ultimo_dia:02d}/{mes_ant:02d}/{ano_ant}"

print(f"Relatório Contas Assinada: {data_de_br} → {data_ate_br}")

# ── Login ────────────────────────────────────────────────────────────────────

try:
    session = mogo_login()
    print("✅ Autenticado no Mogo")
except Exception as e:
    print(f"❌ ERRO de login: {e}")
    sys.exit(1)

# ── Buscar dados: Contas a Receber ───────────────────────────────────────────

url_contas = f"{MOGO_BASE_URL}/Sistema/RelatorioFinanceiro/RelatorioContasAReceber"

params = {
    "chkVencidos": "on",
    "chkReceber": "on",
    "chkRecebido": "on",
    "chkCredito": "on",
    "chkDebito": "on",
    "chkCaixaAVista": "false",
    "inpPesqCrD": "venda em nota assinada",
    "dataDe": data_de_br,
    "dataAte": data_ate_br,
    "selDate": "emissao",
    "chkTituloComp": "false",
    "validaConciliacao": "-1",
}

print(f"\nBuscando {url_contas}")
print(f"Filtros: descrição='venda em nota assinada', período={data_de_br} a {data_ate_br}")

try:
    resp = session.get(
        url_contas,
        params=params,
        timeout=60,
        allow_redirects=True
    )
    
    if resp.status_code != 200:
        print(f"❌ ERRO {resp.status_code}: {resp.text[:300]}")
        sys.exit(1)
    
    print(f"✅ Resposta recebida ({len(resp.text)} bytes)")
    
    # Salvar HTML bruto
    html_file = OUTPUT_DIR / f"{ano_ant:04d}-{mes_ant:02d}-contas-assinada.html"
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(resp.text)
    print(f"✅ HTML salvo: {html_file}")
    
except Exception as e:
    print(f"❌ ERRO ao buscar: {e}")
    sys.exit(1)

# ── Extrair e processar tabela (BeautifulSoup) ──────────────────────────────

xlsx_file = None
try:
    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    
    soup = BeautifulSoup(resp.text, 'html.parser')
    table = soup.find('table')
    
    if table:
        # Extrair headers
        headers = []
        for th in table.find_all('th'):
            headers.append(th.get_text(strip=True))
        
        # Extrair linhas
        rows = []
        for tr in table.find_all('tr')[1:]:  # Pular header
            row = []
            for td in tr.find_all('td'):
                row.append(td.get_text(strip=True))
            if row:
                rows.append(row)
        
        print(f"\n📊 Tabela extraída: {len(rows)} linhas, {len(headers)} colunas")
        
        # Salvar XLSX com openpyxl
        xlsx_file = OUTPUT_DIR / f"{ano_ant:04d}-{mes_ant:02d}-contas-assinada.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Contas Assinada"
        
        # Header
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Dados
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        format_currency_cells(wb)
        wb.save(xlsx_file)
        print(f"✅ XLSX salvo: {xlsx_file}")
        
        # Salvar JSON
        json_data = []
        for row in rows:
            json_data.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
        
        json_file = OUTPUT_DIR / f"{ano_ant:04d}-{mes_ant:02d}-contas-assinada.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON salvo: {json_file}")
        
except Exception as e:
    print(f"⚠️  Não foi possível extrair tabela: {e}")

# ── Enviar email ─────────────────────────────────────────────────────────────

print("\n📧 Enviando email...")

assunto = f"Contas a Receber - Vendas Assinadas {data_de_br} a {data_ate_br}"
corpo = f"""Relatório Mogo — Contas a Receber

Período: {data_de_br} a {data_ate_br}
Filtro: Vendas em Nota Assinada
Data de referência: Emissão

---
Arquivos anexados:
- XLSX com tabela extraída
- HTML com layout original do Mogo
"""

try:
    cmd = ['gog', 'gmail', 'send',
           '--account', 'cakebigdog@gmail.com',
           '--client', 'cakebigdog',
           '--to', 'joao@cakeco.com.br',
           '--subject', assunto,
           '--body', corpo]
    
    if xlsx_file and os.path.exists(xlsx_file):
        cmd.extend(['--attach', str(xlsx_file)])
    cmd.extend(['--attach', str(html_file)])
    
    result = subprocess.run(cmd, capture_output=True, text=True, env=os.environ)
    
    if result.returncode == 0:
        print(f"✅ Email enviado")
    else:
        print(f"⚠️  Resultado: {result.stdout[:200]}")
        if result.stderr:
            print(f"Erro: {result.stderr[:200]}")
        
except Exception as e:
    print(f"❌ ERRO ao enviar email: {e}")
    sys.exit(1)

print("\n✅ Relatório concluído")
