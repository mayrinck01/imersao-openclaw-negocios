#!/usr/bin/env python3
"""
BigDog — Mogo "Vendas Sintético" (Mensal)
Roda todo dia 1 e 2 de cada mês, qualquer horário.
Relatório sempre referente ao mês anterior.

Replica: Relatórios > Vendas Produtos > Vendas Sintético
  - Período: primeiro ao último dia do mês anterior
  - Filtro Data: Pedido (TipoFiltroData=1)
  - Sem filtros adicionais

Endpoint: GET /relatorios/BuscaDadosRelatorioDinamico
  idGeradorRelatorios=0, codRelatorio=27

Campos: A0=Qtde, A1=Produto, A2=Valor Médio, A3=Total, A6=Cód, A7=Grupo, A8=SubGrupo
"""

import sys, os, subprocess, json, calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login
session = mogo_login()

# Período: mês anterior
hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]
data_ini = datetime(ano_ant, mes_ant, 1).strftime('%d/%m/%Y')
data_fim = datetime(ano_ant, mes_ant, ultimo_dia).strftime('%d/%m/%Y')
mes_ref  = datetime(ano_ant, mes_ant, 1).strftime('%m-%Y')
mes_nome = month_year_pt(datetime(ano_ant, mes_ant, 1))

print(f"Vendas Sintético: {data_ini} a {data_fim}...")

filtro = (
    f"TipoFiltroData{{1"
    f"|DataDe{{{data_ini}"
    f"|DataAte{{{data_fim}"
    f"|SubGrupo{{"
    f"|Grupo{{"
    f"|ProdutoEstabelecimento{{"
    f"|ComboProdutos{{"
    f"|LocalEstoque{{"
    f"|TurnoTrabalho{{"
    f"|Estabelecimento{{"
    f"|OrigemPedido{{"
    f"|Cliente{{"
    f"|TipoPagamento{{"
    f"|Atendente{{"
    f"|TipoPedido{{"
)

r = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
    'idGeradorRelatorios': '0',
    'codRelatorio': '27',
    'filtro': filtro,
    'gridparamns': json.dumps({
        "Searching": True, "RecordsCount": 9999,
        "PageIndex": 0, "SortingName": "", "SortingOrder": "ASC"
    }),
    'colunas': '[]',
    'dbNameFranquia': ''
}, timeout=60)

if r.status_code != 200:
    print(f"ERRO ({r.status_code})")
    sys.exit(1)

d = r.json()
registros = d.get('rows') or []
print(f"Registros: {len(registros)}")

if not registros:
    print("Nenhum dado encontrado.")
    sys.exit(0)

# Calcular total geral
total_valor = 0.0
for reg in registros:
    try:
        v = str(reg.get('A3', '0')).replace('.','').replace(',','.').strip()
        total_valor += float(v)
    except:
        pass
total_fmt = f"R$ {total_valor:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

# ── Salvar Excel ───────────────────────────────────────────────────────────────
COLUNAS = [
    ('A1', 'Produto'),
    ('A0', 'Qtde'),
    ('A2', 'Valor Médio'),
    ('A3', 'Total'),
    ('A7', 'Grupo'),
    ('A8', 'SubGrupo'),
    ('A6', 'Cód Produto'),
    ('A5', 'Info Extra'),
]


COLUNAS = order_columns_by_first_record(registros[0] if registros else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Vendas Sintetico'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{mes_ref}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = pt_title("Vendas Sintético")

hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, reg in enumerate(registros, 2):
    for c_idx, (col_key, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=reg.get(col_key, '') or '')

# Linha de total
ws.append([])
ws.append(['TOTAL', '', '', total_fmt, '', '', '', ''])

for i, w in enumerate([35, 10, 14, 14, 20, 20, 12, 12], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Salvar JSON ────────────────────────────────────────────────────────────────
registros_nomeados = [
    {
        'produto':      r.get('A1',''),
        'qtde':         r.get('A0',''),
        'valor_medio':  r.get('A2',''),
        'total':        r.get('A3',''),
        'grupo':        r.get('A7',''),
        'subgrupo':     r.get('A8',''),
        'cod_produto':  r.get('A6',''),
    }
    for r in registros
]
json_path = f"{pasta}/{mes_ref}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({
        "periodo": {"de": data_ini, "ate": data_fim},
        "tipo_data": "Pedido",
        "total_produtos": len(registros),
        "faturamento_total": total_fmt,
        "registros": registros_nomeados
    }, f, ensure_ascii=False, indent=2)
print(f"JSON salvo: {json_path}")

# ── Enviar email ───────────────────────────────────────────────────────────────
corpo = f"📊 Vendas Sintético — {mes_nome}\n"
corpo += "=" * 50 + "\n\n"
corpo += f"📅 Período: {data_ini} a {data_fim} (Filtro Data: Pedido)\n"
corpo += f"📦 Total de produtos: {len(registros)}\n"
corpo += f"💰 Faturamento total: {total_fmt}\n\n"
corpo += "(BigDog 🐕)"

cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "📊 Mogo Vendas Sintético — {mes_nome} — {total_fmt}" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}" '
    f'--attach "{json_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
