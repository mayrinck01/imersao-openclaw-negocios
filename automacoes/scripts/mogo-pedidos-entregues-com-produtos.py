#!/usr/bin/env python3
"""
BigDog — Mogo "Pedidos Entregues com Produtos"

Gera um relatório complementar com os produtos de cada pedido entregue,
sem alterar o relatório original de "Pedidos Entregues".

Saída padrão:
  /root/workspaces/cake-brain/relatorios/Mogo/Pedidos Entregues Itens/DD-MM-YYYY.xlsx

Uso:
  python3 scripts/mogo-pedidos-entregues-com-produtos.py
  python3 scripts/mogo-pedidos-entregues-com-produtos.py --date 16/04/2026
  python3 scripts/mogo-pedidos-entregues-com-produtos.py --date 16-04-2026 --json
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL


OUTPUT_DIR = Path('/root/workspaces/cake-brain/relatorios/Mogo/Pedidos Entregues Itens')


def parse_date(value: str | None) -> datetime:
    if not value:
        return datetime.now() - timedelta(days=1)

    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass

    raise SystemExit('Data inválida. Use DD/MM/YYYY, DD-MM-YYYY ou YYYY-MM-DD.')


def to_api_date(dt: datetime) -> str:
    return dt.strftime('%d/%m/%Y')


def to_file_date(dt: datetime) -> str:
    return dt.strftime('%d-%m-%Y')


def clean_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def parse_decimal(value) -> float:
    if value in (None, ''):
        return 0.0
    text = str(value).strip().replace('.', '').replace(',', '.')
    try:
        return float(text)
    except ValueError:
        return 0.0


def extract_origin(html: str) -> str:
    if not html:
        return ''
    match = re.search(r"title='?\"?([^'\">\s]+)'?\"?", str(html))
    if match:
        return match.group(1)
    return re.sub(r'<[^>]+>', '', str(html)).strip()


def pick_names(items, keys=('Descricao', 'Nome', 'Produto_Produto')) -> list[str]:
    names = []
    for item in items or []:
        for key in keys:
            value = item.get(key)
            if value:
                names.append(str(value).strip())
                break
    return names


def build_item_summary(item: dict) -> str:
    product = clean_text(item.get('Produto_Nome') or item.get('Descricao'))
    qty = parse_decimal(item.get('Quantidade'))
    adicionais = pick_names(item.get('Adicionais'))
    componentes = pick_names(item.get('Componentes'))
    sabores = pick_names(item.get('Sabores'))
    observacoes = pick_names(item.get('Observacoes'))
    nota = clean_text(item.get('Observacao'))

    extras = []
    if adicionais:
        extras.append('Adic: ' + ', '.join(adicionais))
    if componentes:
        extras.append('Comp: ' + ', '.join(componentes))
    if sabores:
        extras.append('Sabores: ' + ', '.join(sabores))
    if observacoes:
        extras.append('Obs item: ' + ', '.join(observacoes))
    if nota:
        extras.append('Nota: ' + nota)

    summary = f'{qty:g}x {product}' if qty else product
    if extras:
        summary += ' (' + ' | '.join(extras) + ')'
    return summary


def fetch_delivered_orders(session, api_date: str) -> list[dict]:
    response = session.post(
        f'{MOGO_URL}/Pedido/ListPedidosParaEntrega',
        params={
            'cFiltroTipoEntrega': '2',
            'dtDe': api_date,
            'dtAte': api_date,
            'tipoFiltroData': 'Entrega',
        },
        data={
            '_search': 'true',
            'nd': '1',
            'rows': '1000',
            'page': '1',
            'sidx': 'HoraInclusao',
            'sord': 'desc',
            'totalrows': '',
        },
        timeout=30,
    )
    response.raise_for_status()
    return response.json().get('rows', [])


def fetch_order_items(session, order_id: int | str) -> list[dict]:
    response = session.get(
        f'{MOGO_URL}/Pedido/DadosItemPedido',
        params={'nId': order_id},
        timeout=30,
    )
    response.raise_for_status()
    data = response.json()
    return data.get('data', []) or []


def auto_fit(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width


def style_header(ws, headers: list[str], color: str = '1F4E78') -> None:
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def build_report_rows(orders: list[dict], session) -> tuple[list[list], list[list], list[dict]]:
    summary_rows = []
    item_rows = []
    raw_rows = []

    for order in orders:
        order_id = order.get('Id')
        items = fetch_order_items(session, order_id)

        item_summaries = []
        raw_items = []

        for item in items:
            product = clean_text(item.get('Produto_Nome') or item.get('Descricao'))
            qty = parse_decimal(item.get('Quantidade'))
            total_value = clean_text(item.get('ValorTotal'))
            adicionais = pick_names(item.get('Adicionais'))
            componentes = pick_names(item.get('Componentes'))
            sabores = pick_names(item.get('Sabores'))
            observacoes = pick_names(item.get('Observacoes'))
            nota = clean_text(item.get('Observacao'))

            item_summaries.append(build_item_summary(item))
            raw_items.append({
                'produto': product,
                'quantidade': qty,
                'valor_total': total_value,
                'adicionais': adicionais,
                'componentes': componentes,
                'sabores': sabores,
                'observacoes': observacoes,
                'nota': nota,
            })

            item_rows.append([
                order.get('NumeroPedido'),
                order.get('NomeCliente'),
                extract_origin(order.get('OrigemPedido', '')),
                order.get('DataEntrega'),
                order.get('HoraEntregaTxt'),
                product,
                qty,
                total_value,
                ', '.join(adicionais),
                ', '.join(componentes),
                ', '.join(sabores),
                '; '.join([x for x in [nota, ', '.join(observacoes)] if x]),
            ])

        summary_rows.append([
            order.get('NumeroPedido'),
            order.get('NomeCliente'),
            extract_origin(order.get('OrigemPedido', '')),
            order.get('DataEntrega'),
            order.get('HoraEntregaTxt'),
            order.get('Bairro'),
            order.get('ObsEntrega_Descricao'),
            clean_text(order.get('ValorFinal')),
            ' | '.join(item_summaries),
        ])

        raw_rows.append({
            'pedido_id': order_id,
            'numero_pedido': order.get('NumeroPedido'),
            'cliente': order.get('NomeCliente'),
            'origem': extract_origin(order.get('OrigemPedido', '')),
            'data_entrega': order.get('DataEntrega'),
            'hora_entrega': order.get('HoraEntregaTxt'),
            'bairro': order.get('Bairro'),
            'forma_entrega': order.get('ObsEntrega_Descricao'),
            'valor_final': clean_text(order.get('ValorFinal')),
            'itens': raw_items,
        })

    return summary_rows, item_rows, raw_rows


def save_workbook(summary_rows: list[list], item_rows: list[list], output_path: Path) -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Resumo por Pedido'
    summary_headers = [
        'Nº Pedido', 'Cliente', 'Origem', 'Data Entrega', 'Hora',
        'Bairro', 'Forma Entrega', 'Valor Final', 'Itens'
    ]
    style_header(ws, summary_headers)
    for row_idx, row in enumerate(summary_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    auto_fit(ws, [12, 28, 12, 14, 8, 18, 22, 12, 120])

    ws2 = wb.create_sheet('Itemizado')
    item_headers = [
        'Nº Pedido', 'Cliente', 'Origem', 'Data Entrega', 'Hora', 'Produto',
        'Qtd', 'Valor Total', 'Adicionais', 'Componentes', 'Sabores', 'Observações'
    ]
    style_header(ws2, item_headers)
    for row_idx, row in enumerate(item_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    auto_fit(ws2, [12, 28, 12, 14, 8, 32, 8, 12, 24, 24, 24, 40])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description='Gera relatório complementar de pedidos entregues com produtos.')
    parser.add_argument('--date', help='Data alvo. Aceita DD/MM/YYYY, DD-MM-YYYY ou YYYY-MM-DD.')
    parser.add_argument('--json', action='store_true', help='Salva também um JSON bruto complementar.')
    args = parser.parse_args()

    target_date = parse_date(args.date)
    api_date = to_api_date(target_date)
    file_date = to_file_date(target_date)

    print(f'Buscando pedidos entregues com produtos em {api_date}...')
    session = mogo_login()
    orders = fetch_delivered_orders(session, api_date)
    print(f'Pedidos encontrados: {len(orders)}')

    if not orders:
        print('Nenhum pedido encontrado para a data informada.')
        return

    summary_rows, item_rows, raw_rows = build_report_rows(orders, session)

    output_xlsx = OUTPUT_DIR / f'{file_date}.xlsx'
    save_workbook(summary_rows, item_rows, output_xlsx)
    print(f'Excel salvo: {output_xlsx}')

    if args.json:
        output_json = OUTPUT_DIR / f'{file_date}.json'
        output_json.write_text(json.dumps(raw_rows, ensure_ascii=False, indent=2))
        print(f'JSON salvo: {output_json}')

    print(f'Concluído: {len(summary_rows)} pedidos | {len(item_rows)} itens')


if __name__ == '__main__':
    main()
