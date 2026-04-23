#!/usr/bin/env python3
"""
BigDog — Mogo Contas a Receber (Vendas em Nota Assinada) — VERSÃO EXCEL DIRETO
Relatório: Financeiro > Contas a Receber
Filtro: Descrição = "venda em nota assinada"
Período: mês anterior (emissão)
Método: Clica no botão Excel para exportar XLSX direto
Agendamento: dia 1º de cada mês, 07:00 BRT
Email: joao@cakeco.com.br
"""

import os, sys, subprocess, time, json, calendar
from datetime import datetime, timedelta
from pathlib import Path
import requests

sys.path.insert(0, '/root/workspaces/cake-brain/automacoes/scripts')
from mogo_login import mogo_login
from mogo_excel import format_currency_cells

# ── Configuração ─────────────────────────────────────────────────────────────

MOGO_BASE_URL = "https://app3.mogogourmet.com.br"
OUTPUT_DIR = Path("/root/workspaces/cake-brain/relatorios/Mogo/ContasAssinada")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Período: mês anterior ────────────────────────────────────────────────────

hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]

data_de_br = f"01/{mes_ant:02d}/{ano_ant}"
data_ate_br = f"{ultimo_dia:02d}/{mes_ant:02d}/{ano_ant}"

print(f"Relatório Contas Assinada: {data_de_br} → {data_ate_br}")

# ── Login ────────────────────────────────────────────────────────────────────

try:
    session = mogo_login()
    print("✅ Autenticado no Mogo")
except Exception as e:
    print(f"❌ ERRO de login: {e}")
    sys.exit(1)

# ── Opção 1: Tentar export direto via parâmetro de download ────────────────

print("\n📥 Tentando download direto do XLSX...")

# O Mogo geralmente oferece export via parâmetro ?export=xlsx ou similar
urls_export = [
    f"{MOGO_BASE_URL}/Sistema/RelatorioFinanceiro/RelatorioContasAReceber/ExportarXlsx",
    f"{MOGO_BASE_URL}/Sistema/RelatorioFinanceiro/RelatorioContasAReceber?export=xlsx",
    f"{MOGO_BASE_URL}/Sistema/RelatorioFinanceiro/RelatorioContasAReceber/Exportar",
]

params = {
    "chkVencidos": "on",
    "chkReceber": "on",
    "chkRecebido": "on",
    "chkCredito": "on",
    "chkDebito": "on",
    "chkCaixaAVista": "false",
    "inpPesqCrD": "venda em nota assinada",
    "dataDe": data_de_br,
    "dataAte": data_ate_br,
    "selDate": "emissao",
    "chkTituloComp": "false",
    "validaConciliacao": "-1",
}

xlsx_file = None

for url_export in urls_export:
    try:
        resp = session.get(url_export, params=params, timeout=60, allow_redirects=True)
        
        # Verificar se é um arquivo Excel (começa com bytes de XLSX)
        if resp.content[:4] == b'PK\x03\x04':  # Assinatura de arquivo ZIP/XLSX
            xlsx_file = OUTPUT_DIR / f"{ano_ant:04d}-{mes_ant:02d}-contas-assinada.xlsx"
            with open(xlsx_file, 'wb') as f:
                f.write(resp.content)
            print(f"✅ XLSX baixado direto: {xlsx_file} ({len(resp.content)} bytes)")
            break
        elif resp.status_code == 200 and len(resp.content) > 1000:
            print(f"⚠️  Resposta recebida mas pode não ser XLSX: {url_export}")
    except Exception as e:
        print(f"  {url_export}: {str(e)[:60]}")

# ── Se não conseguir export direto, fazer scraping do HTML anterior ──────────

if not xlsx_file:
    print("\n⚠️  Não consegui export direto, usando scraping de HTML...")
    
    url_contas = f"{MOGO_BASE_URL}/Sistema/RelatorioFinanceiro/RelatorioContasAReceber"
    
    try:
        resp = session.get(url_contas, params=params, timeout=60)
        
        if resp.status_code == 200:
            from bs4 import BeautifulSoup
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
            
            soup = BeautifulSoup(resp.text, 'html.parser')
            table = soup.find('table')
            
            if table:
                headers = [th.get_text(strip=True) for th in table.find_all('th')]
                rows = []
                for tr in table.find_all('tr')[1:]:
                    row = [td.get_text(strip=True) for td in tr.find_all('td')]
                    if row:
                        rows.append(row)
                
                print(f"📊 Tabela extraída: {len(rows)} linhas")
                
                xlsx_file = OUTPUT_DIR / f"{ano_ant:04d}-{mes_ant:02d}-contas-assinada.xlsx"
                wb = Workbook()
                ws = wb.active
                ws.title = "Contas Assinada"
                
                # Header
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                # Dados
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                format_currency_cells(wb)
                wb.save(xlsx_file)
                print(f"✅ XLSX criado via scraping: {xlsx_file}")
    except Exception as e:
        print(f"❌ Erro ao fazer scraping: {e}")

# ── Salvar HTML bruto para referência ────────────────────────────────────────

print("\n💾 Salvando HTML de referência...")
html_file = OUTPUT_DIR / f"{ano_ant:04d}-{mes_ant:02d}-contas-assinada.html"
try:
    resp = session.get(
        f"{MOGO_BASE_URL}/Sistema/RelatorioFinanceiro/RelatorioContasAReceber",
        params=params, timeout=60
    )
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(resp.text)
    print(f"✅ HTML salvo: {html_file}")
except Exception as e:
    print(f"⚠️  Erro ao salvar HTML: {e}")

# ── Enviar email ─────────────────────────────────────────────────────────────

print("\n📧 Enviando email...")

assunto = f"Contas a Receber - Vendas Assinadas {data_de_br} a {data_ate_br}"
corpo = f"""Relatório Mogo — Contas a Receber

Período: {data_de_br} a {data_ate_br}
Filtro: Vendas em Nota Assinada
Data de referência: Emissão

---
Arquivos anexados:
- XLSX com os dados
- HTML com layout original
"""

try:
    cmd = ['gog', 'gmail', 'send',
           '--account', 'cakebigdog@gmail.com',
           '--client', 'cakebigdog',
           '--to', 'joao@cakeco.com.br',
           '--subject', assunto,
           '--body', corpo]
    
    if xlsx_file and os.path.exists(xlsx_file):
        cmd.extend(['--attach', str(xlsx_file)])
    
    if os.path.exists(html_file):
        cmd.extend(['--attach', str(html_file)])
    
    result = subprocess.run(cmd, capture_output=True, text=True, env=os.environ)
    
    if result.returncode == 0:
        print(f"✅ Email enviado")
    else:
        print(f"⚠️  Resultado: {result.stdout[:200]}")
        
except Exception as e:
    print(f"❌ ERRO ao enviar: {e}")
    sys.exit(1)

print("\n✅ Relatório concluído")
