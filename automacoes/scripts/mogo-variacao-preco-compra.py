#!/usr/bin/env python3
"""
BigDog — Mogo "Variação do Preço de Compra" (Mensal)
Roda todo dia 1 e 2 de cada mês, qualquer horário.
Relatório sempre referente ao mês anterior.

Replica: Relatórios > Compras > Variação do Preço de Compra
  - Período: primeiro ao último dia do mês anterior

Endpoint: GET /relatorios/BuscaDadosRelatorioDinamico
  idGeradorRelatorios=0, codRelatorio=87

Campos: prodN=Produto, ValorNovo, ValorAntigo, diferenca,
        ValorMenor, ValorMaior, custoMedio, PrecoCompra,
        medida, sub=SubCategoria, cat=Categoria, pid=ID
"""

import sys, os, subprocess, json, calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login
session = mogo_login()

# Período: mês anterior
hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]
data_ini = datetime(ano_ant, mes_ant, 1).strftime('%d/%m/%Y')
data_fim = datetime(ano_ant, mes_ant, ultimo_dia).strftime('%d/%m/%Y')
mes_ref  = datetime(ano_ant, mes_ant, 1).strftime('%m-%Y')
mes_nome = month_year_pt(datetime(ano_ant, mes_ant, 1))

print(f"Variação do Preço de Compra: {data_ini} a {data_fim}...")

filtro = (
    f"DataDe{{{data_ini}|DataAte{{{data_fim}"
    f"|Grupo{{|SubGrupo{{|diferenca{{0|percent{{"
    f"|ClassificacaoProduto{{|ProdutoEstabelecimento{{"
    f"|Categoria{{|SubCategoria{{"
)

r = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
    'idGeradorRelatorios': '0', 'codRelatorio': '87',
    'filtro': filtro,
    'gridparamns': json.dumps({
        "Searching": True, "RecordsCount": 9999,
        "PageIndex": 0, "SortingName": "", "SortingOrder": "ASC"
    }),
    'colunas': '[]', 'dbNameFranquia': ''
}, timeout=60)

if r.status_code != 200:
    print(f"ERRO ({r.status_code})")
    sys.exit(1)

d = r.json()
registros = d.get('rows') or []
print(f"Registros: {len(registros)}")

if not registros:
    print("Nenhum dado encontrado.")
    sys.exit(0)

# Contar variações
com_variacao = [reg for reg in registros if reg.get('diferenca','0') not in ('0,00','0')]

# ── Salvar Excel ───────────────────────────────────────────────────────────────
COLUNAS = [
    ('prodN',       'Produto'),
    ('cat',         'Categoria'),
    ('sub',         'SubCategoria'),
    ('medida',      'Unidade'),
    ('ValorAntigo', 'Preço Anterior'),
    ('ValorNovo',   'Preço Atual'),
    ('diferenca',   'Diferença'),
    ('ValorMenor',  'Menor Preço'),
    ('ValorMaior',  'Maior Preço'),
    ('custoMedio',  'Custo Médio'),
    ('PrecoCompra', 'Preço Compra'),
]


COLUNAS = order_columns_by_first_record(registros[0] if registros else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Variacao Preco Compra'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{mes_ref}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Variação Preço Compra"

hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

# Colorir linhas com variação em amarelo
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for r_idx, reg in enumerate(registros, 2):
    tem_variacao = reg.get('diferenca','0') not in ('0,00','0')
    for c_idx, (col_key, _) in enumerate(COLUNAS, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=reg.get(col_key, '') or '')
        if tem_variacao:
            cell.fill = yellow_fill

for i, w in enumerate([40, 15, 18, 20, 14, 14, 12, 12, 12, 12, 14], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Salvar JSON ────────────────────────────────────────────────────────────────
registros_nomeados = [
    {
        'produto':       r.get('prodN',''),
        'categoria':     r.get('cat',''),
        'subcategoria':  r.get('sub',''),
        'unidade':       r.get('medida',''),
        'preco_anterior':r.get('ValorAntigo',''),
        'preco_atual':   r.get('ValorNovo',''),
        'diferenca':     r.get('diferenca',''),
        'menor_preco':   r.get('ValorMenor',''),
        'maior_preco':   r.get('ValorMaior',''),
        'custo_medio':   r.get('custoMedio',''),
        'preco_compra':  r.get('PrecoCompra',''),
    }
    for r in registros
]
json_path = f"{pasta}/{mes_ref}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({
        "periodo": {"de": data_ini, "ate": data_fim},
        "total_produtos": len(registros),
        "produtos_com_variacao": len(com_variacao),
        "registros": registros_nomeados
    }, f, ensure_ascii=False, indent=2)
print(f"JSON salvo: {json_path}")

# ── Enviar email ───────────────────────────────────────────────────────────────
corpo = f"📈 Variação do Preço de Compra — {mes_nome}\n"
corpo += "=" * 50 + "\n\n"
corpo += f"📅 Período: {data_ini} a {data_fim}\n"
corpo += f"📦 Total de produtos: {len(registros)}\n"
corpo += f"⚠️  Com variação de preço: {len(com_variacao)}\n\n"
if com_variacao:
    corpo += "Produtos com variação:\n"
    for reg in com_variacao[:20]:
        corpo += f"  • {reg.get('prodN','')[:35]} | Anterior: R${reg.get('ValorAntigo','')} → Atual: R${reg.get('ValorNovo','')} | Δ {reg.get('diferenca','')}\n"
    if len(com_variacao) > 20:
        corpo += f"  ... (+{len(com_variacao)-20} no Excel)\n"
corpo += "\n(BigDog 🐕)"

cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "📈 Mogo Variação Preço Compra — {mes_nome} — {len(com_variacao)} variação(ões)" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}" '
    f'--attach "{json_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
