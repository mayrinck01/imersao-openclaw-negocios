#!/usr/bin/env python3
"""
BigDog — Mogo "Ticket Médio" (Mensal)
Roda todo dia 1 e 2 de cada mês, qualquer horário.
Relatório sempre referente ao mês anterior.

Replica: Relatórios > Faturamento > Ticket Médio
  - Período: primeiro ao último dia do mês anterior
  - Tipo de Data: Pedido (TipoData=1)

Endpoint: GET /relatorios/BuscaDadosRelatorioDinamico
  idGeradorRelatorios=0, codRelatorio=37

Campos retornados:
  A0=Tipo Pedido, A1=Quantidade, A2=Total, A3=Ticket Médio
"""

import sys, os, subprocess, json, calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login
session = mogo_login()

# Período: mês anterior
hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]
data_ini = datetime(ano_ant, mes_ant, 1).strftime('%d/%m/%Y')
data_fim = datetime(ano_ant, mes_ant, ultimo_dia).strftime('%d/%m/%Y')
mes_ref  = datetime(ano_ant, mes_ant, 1).strftime('%m-%Y')
mes_nome = month_year_pt(datetime(ano_ant, mes_ant, 1))

print(f"Ticket Médio: {data_ini} a {data_fim}...")

filtro = f"DataDe{{{data_ini}|DataAte{{{data_fim}|TipoData{{1"
gridparamns = json.dumps({"Searching": True, "RecordsCount": 9999, "PageIndex": 0,
                           "SortingName": "", "SortingOrder": "ASC"})

r = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
    'idGeradorRelatorios': '0',
    'codRelatorio': '37',
    'filtro': filtro,
    'gridparamns': gridparamns,
    'colunas': '[]',
    'dbNameFranquia': ''
}, timeout=30)

if r.status_code != 200:
    print(f"ERRO ({r.status_code})")
    sys.exit(1)

d = r.json()
registros = d.get('rows') or []
print(f"Registros: {len(registros)}")

if not registros:
    print("Nenhum dado encontrado.")
    sys.exit(0)

# Mapear colunas (A0-A3 → nomes legíveis)
COLUNAS = [
    ('A0', 'Tipo Pedido'),
    ('A1', 'Quantidade'),
    ('A2', 'Total (R$)'),
    ('A3', 'Ticket Médio (R$)'),
]

# ── Salvar Excel ───────────────────────────────────────────────────────────────

COLUNAS = order_columns_by_first_record(registros[0] if registros else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Ticket Medio'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{mes_ref}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = pt_title("Ticket Médio")

hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=10)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, reg in enumerate(registros, 2):
    for c_idx, (col_key, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=reg.get(col_key, '') or '')

# Linha de totais
if registros:
    ws.append([])
    total_qtd = sum(float(str(r.get('A1','0')).replace('.','').replace(',','.')) for r in registros)
    total_val = sum(float(str(r.get('A2','0')).replace('.','').replace(',','.')) for r in registros)
    ticket_geral = total_val / total_qtd if total_qtd > 0 else 0
    ws.append([
        'TOTAL',
        f"{total_qtd:,.0f}".replace(',','.'),
        f"{total_val:,.2f}".replace(',','X').replace('.', ',').replace('X','.'),
        f"{ticket_geral:,.2f}".replace(',','X').replace('.', ',').replace('X','.')
    ])

for i, w in enumerate([18, 16, 18, 18], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Salvar JSON ────────────────────────────────────────────────────────────────
# Converter para nomes legíveis no JSON
registros_nomeados = [
    {
        'tipo_pedido': r.get('A0',''),
        'quantidade':  r.get('A1',''),
        'total':       r.get('A2',''),
        'ticket_medio':r.get('A3','')
    }
    for r in registros
]
json_path = f"{pasta}/{mes_ref}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({
        "periodo": {"de": data_ini, "ate": data_fim},
        "tipo_data": "Pedido",
        "registros": registros_nomeados
    }, f, ensure_ascii=False, indent=2)
print(f"JSON salvo: {json_path}")

# ── Montar email ───────────────────────────────────────────────────────────────
corpo = f"🎯 Ticket Médio — {mes_nome}\n"
corpo += "=" * 45 + "\n\n"
corpo += f"📅 Período: {data_ini} a {data_fim} (Tipo Data: Pedido)\n\n"
for reg in registros:
    corpo += f"  • {reg.get('A0',''):10} | Qtd: {reg.get('A1',''):>10} | Total: R$ {reg.get('A2',''):>12} | Ticket Médio: R$ {reg.get('A3','')}\n"
corpo += "\n(BigDog 🐕)"

# ── Enviar email ───────────────────────────────────────────────────────────────
cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "🎯 Mogo Ticket Médio — {mes_nome}" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}" '
    f'--attach "{json_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
