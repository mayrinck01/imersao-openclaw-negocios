#!/usr/bin/env python3
"""
BigDog — Mogo "Saldo Crédito Carteira" (Mensal)
Roda todo dia 1 e 2 de cada mês, qualquer horário.
Não tem filtro de data — mostra o saldo atual de crédito em carteira dos clientes.

Replica: Relatórios > Financeiro > Saldo Crédito Carteira
  - Clicar em Gerar Relatório (sem filtros)

Endpoint: GET /relatorios/BuscaDadosRelatorioDinamico
  idGeradorRelatorios=0, codRelatorio=130

Campos: nome, cpf, Total=Saldo, subRel=ID Cliente
"""

import sys, os, subprocess, json, calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login
session = mogo_login()

hoje = datetime.now()
mes_ref  = hoje.strftime('%m-%Y')
mes_nome = month_year_pt(hoje)
data_geracao = hoje.strftime('%d/%m/%Y %H:%M')

print(f"Saldo Crédito Carteira — gerado em {data_geracao}...")

r = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
    'idGeradorRelatorios': '0', 'codRelatorio': '130',
    'filtro': 'Cliente{',
    'gridparamns': json.dumps({
        "Searching": True, "RecordsCount": 9999,
        "PageIndex": 0, "SortingName": "", "SortingOrder": "ASC"
    }),
    'colunas': '[]', 'dbNameFranquia': ''
}, timeout=30)

if r.status_code != 200:
    print(f"ERRO ({r.status_code})")
    sys.exit(1)

d = r.json()
registros = d.get('rows') or []
print(f"Clientes com saldo: {len(registros)}")

if not registros:
    print("Nenhum saldo encontrado.")
    sys.exit(0)

# Calcular total
total_saldo = 0.0
for reg in registros:
    try:
        v = str(reg.get('Total', '0')).replace('.','').replace(',','.').strip()
        total_saldo += float(v)
    except:
        pass
total_fmt = f"R$ {total_saldo:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

# ── Salvar Excel ───────────────────────────────────────────────────────────────
COLUNAS = [
    ('nome',   'nome'),
    ('cpf',    'cpf'),
    ('Total',  'Total'),
    ('subRel', 'subRel'),
]


COLUNAS = order_columns_by_first_record(registros[0] if registros else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Saldo Credito Carteira'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{mes_ref}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Planilha"

hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, reg in enumerate(registros, 2):
    for c_idx, (col_key, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=reg.get(col_key, '') or '')

# Total
ws.append([])
ws.append(['TOTAL', '', total_fmt, ''])

for i, w in enumerate([35, 16, 14, 12], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Salvar JSON ────────────────────────────────────────────────────────────────
registros_nomeados = [
    {
        'cliente':    r.get('nome','').strip(),
        'cpf':        r.get('cpf',''),
        'saldo':      r.get('Total',''),
        'id_cliente': r.get('subRel',''),
    }
    for r in registros
]
json_path = f"{pasta}/{mes_ref}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({
        "gerado_em": data_geracao,
        "total_clientes": len(registros),
        "total_saldo": total_fmt,
        "registros": registros_nomeados
    }, f, ensure_ascii=False, indent=2)
print(f"JSON salvo: {json_path}")

# ── Enviar email ───────────────────────────────────────────────────────────────
corpo = f"💳 Saldo Crédito Carteira — {mes_nome}\n"
corpo += "=" * 50 + "\n\n"
corpo += f"📅 Gerado em: {data_geracao}\n"
corpo += f"👥 Clientes com saldo: {len(registros)}\n"
corpo += f"💰 Total em carteira: {total_fmt}\n\n"
corpo += "(BigDog 🐕)"

cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "💳 Mogo Saldo Crédito Carteira — {mes_nome} — {total_fmt}" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}" '
    f'--attach "{json_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
