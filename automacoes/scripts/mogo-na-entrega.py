#!/usr/bin/env python3
"""
BigDog — Mogo "Na Entrega" Monitor
Roda às 08:28 BRT todos os dias.
Se houver pedidos com status "Entregando" às 08:28, gera alerta por email.

Endpoint: POST /Pedido/ListPedidosParaEntrega
  Query:  cFiltroTipoEntrega=1
  Form:   _search=false, rows=1000, page=1, sidx=HoraInclusao, sord=desc
"""

import sys, os, subprocess, json, re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login com novo fluxo de 3 etapas
session = mogo_login()
print("Login OK. Buscando pedidos Na Entrega...")

# POST /Pedido/ListPedidosParaEntrega com cFiltroTipoEntrega=1 (Em Entrega/Entregando)
r = session.post(
    f"{MOGO_URL}/Pedido/ListPedidosParaEntrega",
    params={'cFiltroTipoEntrega': '1'},
    data={
        '_search': 'false',
        'nd': '1',
        'rows': '1000',
        'page': '1',
        'sidx': 'HoraInclusao',
        'sord': 'desc'
    },
    timeout=30
)

if r.status_code != 200:
    print(f"ERRO ao buscar pedidos ({r.status_code})")
    sys.exit(1)

try:
    data = r.json()
    pedidos = data.get('rows', [])
except Exception as e:
    print(f"ERRO ao parsear JSON: {e}")
    sys.exit(1)

hoje = datetime.now().strftime('%d-%m-%Y')
hoje_br = datetime.now().strftime('%d/%m/%Y')
print(f"Pedidos Na Entrega: {len(pedidos)}")

if not pedidos:
    print("✅ Nenhum pedido na entrega. Tudo ok.")
    sys.exit(0)

# Tem pedidos em rota às 08:28 — ALERTA!
COLUNAS = [
    ('NumeroPedido',         'Nº Pedido'),
    ('NomeCliente',          'Cliente'),
    ('DataEntrega',          'Data Entrega'),
    ('HoraEntregaTxt',       'Hora Entrega'),
    ('Logradouro',           'Logradouro'),
    ('Bairro',               'Bairro'),
    ('ValorFinal',           'Valor Final'),
    ('FormaPagtoDelivery',   'Forma Pagto'),
    ('HoraSaidaEntregadorTxt', 'Hora Saída'),
]


COLUNAS = order_columns_by_first_record(pedidos[0] if pedidos else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Na Entrega'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{hoje}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = pt_title("Na Entrega")

hfill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, pedido in enumerate(pedidos, 2):
    for c_idx, (col_name, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=pedido.get(col_name, '') or '')

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# Email de alerta
linhas = f"🚨 ALERTA — {len(pedidos)} pedido(s) ainda na entrega às 08:28 de {hoje_br}\n\n"
for p in pedidos:
    linhas += (
        f"• #{p.get('NumeroPedido','')} | {p.get('NomeCliente','')[:25]} | "
        f"{p.get('Bairro','')} | R${p.get('ValorFinal','')} | "
        f"Saída: {p.get('HoraSaidaEntregadorTxt','') or '?'}\n"
    )
linhas += "\nVerifique se esses pedidos foram entregues e baixados corretamente."

cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "🚨 ALERTA Mogo — {len(pedidos)} pedido(s) na entrega — {hoje_br}" '
    f'--body "{linhas}" '
    f'--attach "{xlsx_path}"'
]
result = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in result.stdout:
    print(f"✅ Email de alerta enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {result.stderr[:200]}")

print("Concluído.")
