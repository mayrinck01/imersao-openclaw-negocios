#!/usr/bin/env python3
"""
BigDog — Mogo "Análise de Cortesias" (Mensal)
Roda todo dia 1 e 2 de cada mês, qualquer horário.
Relatório sempre referente ao mês anterior.

Replica: Relatórios > Auditoria > Análise de Cortesias
  - Período: primeiro ao último dia do mês anterior

Endpoint: GET /relatorios/BuscaDadosRelatorioDinamico
  idGeradorRelatorios=0, codRelatorio=23

Campos: A0=Data, A1=Hora, A2=Funcionário, A3=Valor, A4=Cliente,
        A5=Telefone, A8=Motivo, A9=Custo, A10=ID
"""

import sys, os, subprocess, json, calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login
session = mogo_login()

# Período: mês anterior
hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]
data_ini = datetime(ano_ant, mes_ant, 1).strftime('%d/%m/%Y')
data_fim = datetime(ano_ant, mes_ant, ultimo_dia).strftime('%d/%m/%Y')
mes_ref  = datetime(ano_ant, mes_ant, 1).strftime('%m-%Y')
mes_nome = month_year_pt(datetime(ano_ant, mes_ant, 1))

print(f"Análise de Cortesias: {data_ini} a {data_fim}...")

filtro = f"Funcionario{{|DataDe{{{data_ini}|DataAte{{{data_fim}"

r = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
    'idGeradorRelatorios': '0', 'codRelatorio': '23',
    'filtro': filtro,
    'gridparamns': json.dumps({
        "Searching": True, "RecordsCount": 9999,
        "PageIndex": 0, "SortingName": "", "SortingOrder": "ASC"
    }),
    'colunas': '[]', 'dbNameFranquia': ''
}, timeout=30)

if r.status_code != 200:
    print(f"ERRO ({r.status_code})")
    sys.exit(1)

d = r.json()
registros = d.get('rows') or []
print(f"Registros: {len(registros)}")

if not registros:
    print("Nenhuma cortesia encontrada.")
    sys.exit(0)

# Calcular totais
total_valor = 0.0
total_custo = 0.0
for reg in registros:
    try:
        total_valor += float(str(reg.get('A3','0')).replace('.','').replace(',','.').strip())
        total_custo += float(str(reg.get('A9','0')).replace('.','').replace(',','.').strip())
    except:
        pass

def fmt(v):
    return f"R$ {v:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

# ── Salvar Excel ───────────────────────────────────────────────────────────────
COLUNAS = [
    ('A0', 'Data'),
    ('A1', 'Hora'),
    ('A2', 'Funcionário'),
    ('A4', 'Cliente'),
    ('A5', 'Telefone'),
    ('A3', 'Valor'),
    ('A9', 'Custo'),
    ('A8', 'Motivo'),
    ('A10','ID'),
]


COLUNAS = order_columns_by_first_record(registros[0] if registros else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Analise Cortesias'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{mes_ref}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = pt_title("Análise de Cortesias")

hfill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, reg in enumerate(registros, 2):
    for c_idx, (col_key, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=reg.get(col_key, '') or '')

ws.append([])
ws.append(['TOTAL', '', '', '', '', fmt(total_valor), fmt(total_custo), '', ''])

for i, w in enumerate([12, 8, 18, 30, 16, 12, 12, 50, 10], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Salvar JSON ────────────────────────────────────────────────────────────────
registros_nomeados = [
    {
        'data':        r.get('A0',''),
        'hora':        r.get('A1',''),
        'funcionario': r.get('A2',''),
        'cliente':     r.get('A4',''),
        'telefone':    r.get('A5',''),
        'valor':       r.get('A3',''),
        'custo':       r.get('A9',''),
        'motivo':      r.get('A8',''),
        'id':          r.get('A10',''),
    }
    for r in registros
]
json_path = f"{pasta}/{mes_ref}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({
        "periodo": {"de": data_ini, "ate": data_fim},
        "total_registros": len(registros),
        "total_valor": fmt(total_valor),
        "total_custo": fmt(total_custo),
        "registros": registros_nomeados
    }, f, ensure_ascii=False, indent=2)
print(f"JSON salvo: {json_path}")

# ── Enviar email ───────────────────────────────────────────────────────────────
corpo = f"🎁 Análise de Cortesias — {mes_nome}\n"
corpo += "=" * 50 + "\n\n"
corpo += f"📅 Período: {data_ini} a {data_fim}\n"
corpo += f"📦 Total de cortesias: {len(registros)}\n"
corpo += f"💰 Valor total: {fmt(total_valor)}\n"
corpo += f"💸 Custo total: {fmt(total_custo)}\n\n"
for reg in registros:
    corpo += f"• {reg.get('A0','')} | {reg.get('A2','')} | {reg.get('A4','')[:25]} | R${reg.get('A3','')} | {reg.get('A8','')[:40]}\n"
corpo += "\n(BigDog 🐕)"

cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "🎁 Mogo Análise de Cortesias — {mes_nome} — {len(registros)} cortesia(s)" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}" '
    f'--attach "{json_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
