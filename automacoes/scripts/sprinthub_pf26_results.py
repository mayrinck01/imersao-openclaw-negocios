#!/usr/bin/env python3
"""Local BigDog automation for Pesquisa Fidelidade 2026 results.

Read-only by design: this module classifies SprintHub lead snapshots and writes
local reports. It does not create/update SprintHub leads, tags, or fields.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote
from urllib.request import Request, urlopen


CAMPAIGN_TAGS = {
    "pesquisa_fidelidade_2026",
    "pesquisa_fidelidade_2026_iniciado",
    "pesquisa_fidelidade_2026_concluido",
    "pesquisa_fidelidade_2026_detrator",
    "cupom_vip12_entregue",
    "brinde_vip12_resgatado",
}

FIELD_ALIASES = {
    "cupom": {
        "pesq_fid_cupom",
        "pesquisa fidelidade cupom",
        "pesquisa fidelidade - cupom",
        "cupom fidelidade codigo",
        "cupom_fidelidade_codigo",
    },
    "status": {
        "pesq_fid_status",
        "pesquisa fidelidade status",
        "pesquisa fidelidade - status",
        "pesquisa_fidelidade_status",
    },
    "nps": {
        "nps",
        "pesq_fid_nps",
        "pesquisa fidelidade nps",
        "pesquisa fidelidade - nps",
        "pesquisa_fidelidade_nps",
    },
    "comentario_nps": {
        "comentario_nps",
        "pesq_fid_comentario_nps",
        "pesquisa fidelidade comentario nps",
        "pesquisa fidelidade - comentario nps",
        "pesquisa fidelidade - comentário nps",
    },
    "aniversario": {
        "aniversario",
        "aniversario_raw",
        "data_aniversario_raw",
        "pesq_fid_aniversario",
        "pesquisa fidelidade aniversario",
        "pesquisa fidelidade - aniversario",
        "pesquisa fidelidade - aniversário",
    },
    "frequencia": {
        "frequencia",
        "frequência",
        "pesq_fid_frequencia",
        "pesquisa fidelidade frequencia",
        "pesquisa fidelidade - frequencia",
        "pesquisa fidelidade - frequência",
    },
    "motivacao": {
        "motivacao",
        "motivação",
        "pesq_fid_motivo",
        "pesquisa fidelidade motivo",
        "pesquisa fidelidade - motivo",
        "pesquisa fidelidade - motivacao",
        "pesquisa fidelidade - motivação",
    },
    "beneficio": {
        "beneficio",
        "benefício",
        "pesq_fid_beneficio",
        "pesquisa fidelidade beneficio",
        "pesquisa fidelidade - beneficio",
        "pesquisa fidelidade - benefício",
    },
    "ideia_livre": {
        "ideia_livre",
        "pesq_fid_ideia_livre",
        "pesquisa fidelidade ideia livre",
        "pesquisa fidelidade - ideia livre",
    },
    "brinde_resgatado": {
        "pesq_fid_brinde_resgatado",
        "pesquisa fidelidade brinde resgatado",
        "pesquisa fidelidade - brinde resgatado",
        "brinde_resgatado",
    },
}


def normalize_key(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_tag(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def only_digits(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def classify_nps(value: Any) -> str:
    try:
        score = int(str(value).strip())
    except (TypeError, ValueError):
        return "invalido"
    if score < 0 or score > 10:
        return "invalido"
    if score <= 6:
        return "detrator"
    if score <= 8:
        return "neutro"
    return "promotor"


def normalize_birthday(value: Any) -> tuple[str, bool]:
    raw = str(value or "").strip()
    if not raw:
        return "", True

    digits = only_digits(raw)
    if len(digits) == 8:
        raw = f"{digits[:2]}/{digits[2:4]}/{digits[4:]}"
    elif len(digits) == 4:
        raw = f"{digits[:2]}/{digits[2:]}"

    for fmt in ("%d/%m/%Y", "%d/%m"):
        try:
            dt = datetime.strptime(raw, fmt)
        except ValueError:
            continue
        if fmt == "%d/%m/%Y":
            current_year = datetime.now().year
            if dt.year < 1900 or dt.year > current_year:
                return str(value or "").strip(), False
            return dt.strftime("%d/%m/%Y"), True
        return dt.strftime("%d/%m"), True

    return str(value or "").strip(), False


def _stringify_field_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value).strip()
    if isinstance(value, list):
        parts = [_stringify_field_value(item) for item in value]
        return "; ".join(part for part in parts if part)
    if isinstance(value, dict):
        for key in ("value", "text", "label", "name", "answer", "response"):
            result = _stringify_field_value(value.get(key))
            if result:
                return result
    return ""


def _collect_field_pairs(obj: Any) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    if isinstance(obj, dict):
        possible_name = (
            obj.get("name")
            or obj.get("label")
            or obj.get("key")
            or obj.get("field")
            or obj.get("alias")
            or obj.get("title")
        )
        possible_value = (
            obj.get("value")
            if "value" in obj
            else obj.get("text")
            if "text" in obj
            else obj.get("answer")
            if "answer" in obj
            else obj.get("response")
            if "response" in obj
            else None
        )
        if possible_name is not None and possible_value is not None:
            pairs.append((str(possible_name), _stringify_field_value(possible_value)))

        for key, value in obj.items():
            if isinstance(value, (dict, list)):
                pairs.extend(_collect_field_pairs(value))
            elif value not in (None, ""):
                pairs.append((str(key), _stringify_field_value(value)))
    elif isinstance(obj, list):
        for item in obj:
            pairs.extend(_collect_field_pairs(item))
    return pairs


def extract_custom_fields(lead: dict[str, Any]) -> dict[str, str]:
    fields: dict[str, str] = {}
    search_roots = [
        lead,
        lead.get("customFields"),
        lead.get("custom_fields"),
        lead.get("fields"),
        lead.get("allFields"),
        lead.get("extra"),
    ]
    for root in search_roots:
        for key, value in _collect_field_pairs(root):
            normalized = normalize_key(key)
            if normalized and value and normalized not in fields:
                fields[normalized] = value
    return fields


def get_field(fields: dict[str, str], logical_name: str) -> str:
    aliases = {normalize_key(alias) for alias in FIELD_ALIASES[logical_name]}
    for key, value in fields.items():
        if key in aliases:
            return value
    # Accept loose prefix/suffix matches for SprintHub labels that include IDs.
    for key, value in fields.items():
        if any(len(alias) >= 4 and alias in key for alias in aliases):
            return value
    return ""


def extract_tags(lead: dict[str, Any]) -> set[str]:
    tags: set[str] = set()
    raw_tags = lead.get("tags") or lead.get("tag") or []
    if isinstance(raw_tags, str):
        raw_tags = [raw_tags]
    if isinstance(raw_tags, dict):
        raw_tags = [raw_tags]
    if isinstance(raw_tags, list):
        for item in raw_tags:
            if isinstance(item, dict):
                tag = item.get("tag") or item.get("name") or item.get("label")
            else:
                tag = item
            normalized = normalize_tag(tag)
            if normalized:
                tags.add(normalized)
    return tags


def _lead_id(lead: dict[str, Any]) -> str:
    return str(lead.get("id") or lead.get("lead_id") or lead.get("contact_id") or "").strip()


def _lead_name(lead: dict[str, Any]) -> str:
    full = str(lead.get("fullname") or lead.get("lead_name") or lead.get("name") or "").strip()
    if full:
        return full
    return " ".join(
        part
        for part in [str(lead.get("firstname") or "").strip(), str(lead.get("lastname") or "").strip()]
        if part
    )


def classify_lead(lead: dict[str, Any]) -> dict[str, Any]:
    fields = extract_custom_fields(lead)
    tags = extract_tags(lead)

    cupom = get_field(fields, "cupom")
    nps = get_field(fields, "nps")
    aniversario_raw = get_field(fields, "aniversario")
    aniversario, aniversario_ok = normalize_birthday(aniversario_raw)
    nps_categoria = classify_nps(nps) if nps else ""

    has_iniciado = "pesquisa_fidelidade_2026_iniciado" in tags
    has_concluido = "pesquisa_fidelidade_2026_concluido" in tags
    has_vip12 = "cupom_vip12_entregue" in tags or cupom.upper() == "VIP12"
    has_detrator_tag = "pesquisa_fidelidade_2026_detrator" in tags
    legacy_pesquisa15 = cupom.upper() == "PESQUISA15"

    response_values = {
        "frequencia": get_field(fields, "frequencia"),
        "motivacao": get_field(fields, "motivacao"),
        "beneficio": get_field(fields, "beneficio"),
        "comentario_nps": get_field(fields, "comentario_nps"),
        "ideia_livre": get_field(fields, "ideia_livre"),
        "brinde_resgatado": get_field(fields, "brinde_resgatado"),
    }
    has_response_field = any(response_values.values()) or bool(nps) or bool(aniversario_raw)
    current_campaign = bool((tags & CAMPAIGN_TAGS) or has_response_field or has_vip12)

    diagnosticos: list[str] = []
    needs_alert = False
    needs_followup = False

    if legacy_pesquisa15 and not (tags & CAMPAIGN_TAGS) and not has_response_field:
        status = "legado_pesquisa15"
        current_campaign = False
    elif has_concluido:
        status = "concluido"
    elif has_iniciado or has_response_field:
        status = "incompleto"
        needs_followup = True
        diagnosticos.append("iniciado_sem_concluir")
    else:
        status = "fora_da_campanha"
        current_campaign = False

    if nps_categoria == "detrator" or has_detrator_tag:
        needs_alert = True
        diagnosticos.append("nps_detrator")
    elif nps_categoria == "invalido":
        needs_alert = True
        diagnosticos.append("nps_invalido")

    if has_detrator_tag and nps_categoria and nps_categoria != "detrator":
        needs_alert = True
        diagnosticos.append("tag_detrator_inconsistente_com_nps")

    if not aniversario_ok:
        needs_alert = True
        diagnosticos.append("aniversario_invalido")

    if has_concluido and not has_vip12:
        diagnosticos.append("concluido_sem_vip12")
        needs_alert = True

    if legacy_pesquisa15:
        diagnosticos.append("legado_pesquisa15")

    return {
        "lead_id": _lead_id(lead),
        "lead_name": _lead_name(lead),
        "whatsapp": str(lead.get("whatsapp") or lead.get("phone") or lead.get("mobile") or "").strip(),
        "status": status,
        "current_campaign": current_campaign,
        "legacy_pesquisa15": legacy_pesquisa15,
        "tags": ",".join(sorted(tags)),
        "cupom": cupom,
        "nps": nps,
        "nps_categoria": nps_categoria,
        "aniversario": aniversario,
        "aniversario_valido": aniversario_ok,
        "frequencia": response_values["frequencia"],
        "motivacao": response_values["motivacao"],
        "beneficio": response_values["beneficio"],
        "comentario_nps": response_values["comentario_nps"],
        "ideia_livre": response_values["ideia_livre"],
        "brinde_resgatado": response_values["brinde_resgatado"],
        "vip12_entregue": has_vip12,
        "needs_alert": needs_alert,
        "needs_followup": needs_followup,
        "diagnosticos": diagnosticos,
    }


def build_campaign_report(leads: list[dict[str, Any]]) -> dict[str, Any]:
    rows = [classify_lead(lead) for lead in leads]
    current_rows = [row for row in rows if row["current_campaign"] and not row["legacy_pesquisa15"]]
    legacy_rows = [row for row in rows if row["legacy_pesquisa15"]]
    diagnostics_rows = [row for row in rows if row["diagnosticos"]]

    status_counts = Counter(row["status"] for row in rows)
    nps_counts = Counter(row["nps_categoria"] for row in current_rows if row["nps_categoria"])

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
        "current_rows": current_rows,
        "legacy_rows": legacy_rows,
        "diagnostics_rows": diagnostics_rows,
        "totals": {
            "leads": len(rows),
            "atuais": len(current_rows),
            "concluidos": status_counts["concluido"],
            "incompletos": status_counts["incompleto"],
            "legado_pesquisa15": len(legacy_rows),
            "vip12_entregue": sum(1 for row in rows if row["vip12_entregue"]),
            "alertas": sum(1 for row in rows if row["needs_alert"]),
            "followups": sum(1 for row in rows if row["needs_followup"]),
            "nps_detrator": nps_counts["detrator"],
            "nps_neutro": nps_counts["neutro"],
            "nps_promotor": nps_counts["promotor"],
            "nps_invalido": nps_counts["invalido"],
        },
    }


CSV_COLUMNS = [
    "lead_id",
    "lead_name",
    "whatsapp",
    "status",
    "cupom",
    "nps",
    "nps_categoria",
    "aniversario",
    "frequencia",
    "motivacao",
    "beneficio",
    "comentario_nps",
    "ideia_livre",
    "brinde_resgatado",
    "vip12_entregue",
    "needs_alert",
    "needs_followup",
    "diagnosticos",
    "tags",
]


def _csv_value(row: dict[str, Any], column: str) -> str:
    value = row.get(column, "")
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    if isinstance(value, bool):
        return "sim" if value else "nao"
    return str(value or "")


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str] = CSV_COLUMNS) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _csv_value(row, column) for column in columns})


def write_summary(path: Path, report: dict[str, Any], as_of: str) -> None:
    totals = report["totals"]
    lines = [
        f"# Pesquisa Fidelidade 2026 - BigDog",
        "",
        f"Gerado em: {report['generated_at']}",
        f"Data-base: {as_of}",
        "",
        "## Totais",
        "",
        f"- Leads analisados: {totals['leads']}",
        f"- Atuais PF26: {totals['atuais']}",
        f"- Concluidos: {totals['concluidos']}",
        f"- Incompletos: {totals['incompletos']}",
        f"- VIP12 entregue: {totals['vip12_entregue']}",
        f"- Legado PESQUISA15: {totals['legado_pesquisa15']}",
        f"- Alertas BigDog: {totals['alertas']}",
        f"- Follow-ups: {totals['followups']}",
        "",
        "## NPS",
        "",
        f"- Promotores: {totals['nps_promotor']}",
        f"- Neutros: {totals['nps_neutro']}",
        f"- Detratores: {totals['nps_detrator']}",
        f"- Invalidos: {totals['nps_invalido']}",
        "",
        "## Regra operacional",
        "",
        "- Este relatorio e read-only: nao escreve no SprintHub.",
        "- PESQUISA15 fica separado como legado e nao entra no resultado atual.",
        "- Alertas apontam casos para revisao humana antes de qualquer acao externa.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def write_xlsx_if_available(path: Path, report: dict[str, Any]) -> Path | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Respostas atuais"
    ws.append(CSV_COLUMNS)
    for row in report["current_rows"]:
        ws.append([_csv_value(row, column) for column in CSV_COLUMNS])

    ws = wb.create_sheet("Diagnostico")
    ws.append(CSV_COLUMNS)
    for row in report["diagnostics_rows"]:
        ws.append([_csv_value(row, column) for column in CSV_COLUMNS])

    ws = wb.create_sheet("Legado PESQUISA15")
    ws.append(CSV_COLUMNS)
    for row in report["legacy_rows"]:
        ws.append([_csv_value(row, column) for column in CSV_COLUMNS])

    ws = wb.create_sheet("Totais")
    ws.append(["metrica", "valor"])
    for key, value in report["totals"].items():
        ws.append([key, value])

    wb.save(path)
    return path


def export_report(report: dict[str, Any], output_dir: Path, as_of: str | None = None) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    as_of = as_of or datetime.now().date().isoformat()

    paths = {
        "summary_md": output_dir / f"pf26-resumo-{as_of}.md",
        "responses_csv": output_dir / f"pf26-respostas-{as_of}.csv",
        "diagnostics_csv": output_dir / f"pf26-diagnostico-{as_of}.csv",
        "legacy_csv": output_dir / f"pf26-legado-pesquisa15-{as_of}.csv",
    }

    write_summary(paths["summary_md"], report, as_of)
    write_csv(paths["responses_csv"], report["current_rows"])
    write_csv(paths["diagnostics_csv"], report["diagnostics_rows"])
    write_csv(paths["legacy_csv"], report["legacy_rows"])

    xlsx_path = write_xlsx_if_available(output_dir / f"pf26-resultados-{as_of}.xlsx", report)
    if xlsx_path:
        paths["xlsx"] = xlsx_path
    return paths


@dataclass(frozen=True)
class SprintHubReadOnlyConfig:
    api_base: str
    token: str
    origin: str = "cakeco"
    limit: int = 120


class SprintHubReadOnlyClient:
    """Small read-only SprintHub client.

    The client only performs GET requests. It intentionally has no methods for
    lead updates, tag changes, or opportunity changes.
    """

    def __init__(self, config: SprintHubReadOnlyConfig):
        self.config = config

    def _get_json(self, path: str) -> dict[str, Any]:
        request = Request(
            f"{self.config.api_base.rstrip('/')}{path}",
            headers={
                "Accept": "application/json",
                "apitoken": self.config.token,
                "origin": self.config.origin,
            },
            method="GET",
        )
        with urlopen(request, timeout=30) as response:
            payload = response.read().decode("utf-8")
        data = json.loads(payload)
        if not isinstance(data, dict):
            raise ValueError("SprintHub response must be a JSON object")
        return data

    def list_leads_page(self, page: int) -> dict[str, Any]:
        # SprintHub does not expose Lead.tags in the public leads list query.
        # Tag/campaign classification is done after fetching each candidate
        # with allFields=1.
        query = quote("{total,leads{id,fullname,whatsapp,updatedDate,lastActive}}")
        return self._get_json(f"/leads?query={query}&limit={self.config.limit}&page={page}")

    def get_lead_all_fields(self, lead_id: str) -> dict[str, Any]:
        data = self._get_json(f"/leads/{quote(str(lead_id))}?allFields=1")
        lead = data.get("lead") or data.get("data", {}).get("lead") or data
        if not isinstance(lead, dict):
            return {}
        lead["tags"] = self.get_lead_tags(lead_id)
        return lead

    def get_lead_tags(self, lead_id: str) -> list[dict[str, Any]]:
        query = quote("{tags{id,tag,color}}")
        data = self._get_json(f"/leads/{quote(str(lead_id))}?query={query}")
        tags = data.get("tags") or data.get("data", {}).get("tags") or []
        return tags if isinstance(tags, list) else []

    def collect_campaign_candidates(self, max_pages: int = 50) -> list[dict[str, Any]]:
        candidates: list[dict[str, Any]] = []
        seen: set[str] = set()
        for page in range(1, max_pages + 1):
            data = self.list_leads_page(page)
            leads = data.get("leads") or data.get("data", {}).get("leads") or []
            if not leads:
                break
            for lead in leads:
                if not isinstance(lead, dict):
                    continue
                lead_id = _lead_id(lead)
                if not lead_id or lead_id in seen:
                    continue
                full = self.get_lead_all_fields(lead_id)
                if isinstance(full, dict):
                    row = classify_lead(full)
                    if row["current_campaign"] or row["legacy_pesquisa15"]:
                        candidates.append(full)
                    seen.add(lead_id)
        return candidates

    def collect_leads_by_id(self, lead_ids: list[str]) -> list[dict[str, Any]]:
        leads: list[dict[str, Any]] = []
        for lead_id in lead_ids:
            full = self.get_lead_all_fields(lead_id)
            if isinstance(full, dict):
                leads.append(full)
        return leads


def load_leads_from_json(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        leads = data.get("leads") or data.get("rows") or data.get("data")
        if isinstance(leads, list):
            return [item for item in leads if isinstance(item, dict)]
    raise ValueError("input JSON must be a list of leads or an object with leads")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="BigDog PF26 local read-only report")
    parser.add_argument("--input-json", type=Path, help="Local SprintHub lead snapshot JSON")
    parser.add_argument("--output-dir", type=Path, default=Path("/root/.openclaw/media/outbound/pf26-bigdog"))
    parser.add_argument("--as-of", default=datetime.now().date().isoformat())
    parser.add_argument("--fetch-sprinthub", action="store_true", help="Read campaign leads from SprintHub API")
    parser.add_argument("--lead-id", action="append", default=[], help="Fetch a specific SprintHub lead id with allFields=1")
    parser.add_argument("--api-base", default="https://sprinthub-api-master.sprinthub.app")
    parser.add_argument("--api-token-env", default="SPRINTHUB_API_TOKEN")
    parser.add_argument("--origin", default="cakeco")
    parser.add_argument("--max-pages", type=int, default=1)
    args = parser.parse_args(argv)

    if args.fetch_sprinthub:
        import os

        token = os.environ.get(args.api_token_env, "")
        if not token:
            raise SystemExit(f"{args.api_token_env} is required for --fetch-sprinthub")
        client = SprintHubReadOnlyClient(SprintHubReadOnlyConfig(args.api_base, token, args.origin))
        if args.lead_id:
            leads = client.collect_leads_by_id(args.lead_id)
        else:
            leads = client.collect_campaign_candidates(max_pages=args.max_pages)
    elif args.input_json:
        leads = load_leads_from_json(args.input_json)
    else:
        raise SystemExit("Provide --input-json or --fetch-sprinthub")

    report = build_campaign_report(leads)
    paths = export_report(report, args.output_dir, args.as_of)
    print(json.dumps({key: str(value) for key, value in paths.items()}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
