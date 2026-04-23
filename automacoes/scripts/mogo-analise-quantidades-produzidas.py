#!/usr/bin/env python3
"""
BigDog — Mogo "Análise de Quantidades Produzidas" (Mensal)
Roda todo dia 1 e 2 de cada mês, qualquer horário.
Relatório sempre referente ao mês anterior.

Replica: Relatórios > Produção / KDS > Análise de Quantidades Produzidas
  - Período: primeiro ao último dia do mês anterior

Endpoint: GET /relatorios/BuscaDadosRelatorioDinamico
  idGeradorRelatorios=0, codRelatorio=98

Campos: A0=Produto, A1=Qtde, A2=Custo Unit., A3=Custo Total,
        A4=Data, A5=Hora, A6=Turno, A7=Chave
"""

import sys, os, subprocess, json, calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login
session = mogo_login()

# Período: mês anterior
hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]
data_ini = datetime(ano_ant, mes_ant, 1).strftime('%d/%m/%Y')
data_fim = datetime(ano_ant, mes_ant, ultimo_dia).strftime('%d/%m/%Y')
mes_ref  = datetime(ano_ant, mes_ant, 1).strftime('%m-%Y')
mes_nome = month_year_pt(datetime(ano_ant, mes_ant, 1))

print(f"Análise de Quantidades Produzidas: {data_ini} a {data_fim}...")

filtro = f"DataDe{{{data_ini}|DataAte{{{data_fim}|ProdutoEstabelecimento{{|Categoria{{|SubCategoria{{"

r_test = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
    'idGeradorRelatorios': '0', 'codRelatorio': '98',
    'filtro': filtro,
    'gridparamns': json.dumps({"Searching": True, "RecordsCount": 1, "PageIndex": 0, "SortingName": "", "SortingOrder": "ASC"}),
    'colunas': '[]', 'dbNameFranquia': ''
}, timeout=60)
total_records = int(r_test.json().get('records', 0))
print(f"Total de registros: {total_records}")

todos_registros = []
lote = 2000
pagina = 0

while True:
    r = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
        'idGeradorRelatorios': '0', 'codRelatorio': '98',
        'filtro': filtro,
        'gridparamns': json.dumps({
            "Searching": True, "RecordsCount": lote,
            "PageIndex": pagina, "SortingName": "", "SortingOrder": "ASC"
        }),
        'colunas': '[]', 'dbNameFranquia': ''
    }, timeout=90)

    rows = r.json().get('rows') or []
    todos_registros.extend(rows)
    print(f"  Lote {pagina+1}: {len(rows)} (total: {len(todos_registros)})")

    if len(rows) < lote or len(todos_registros) >= total_records:
        break
    pagina += 1

print(f"Total carregado: {len(todos_registros)}")

if not todos_registros:
    print("Nenhum dado encontrado.")
    sys.exit(0)

# Calcular total custo
total_custo = 0.0
for reg in todos_registros:
    try:
        v = float(str(reg.get('A3','0')).replace('.','').replace(',','.').strip())
        total_custo += v
    except:
        pass

def fmt(v):
    return f"R$ {v:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

# ── Salvar Excel ───────────────────────────────────────────────────────────────
COLUNAS = [
    ('A4', 'Data'),
    ('A5', 'Hora'),
    ('A0', 'Produto'),
    ('A1', 'Qtde'),
    ('A2', 'Custo Unit.'),
    ('A3', 'Custo Total'),
    ('A6', 'Turno'),
]


COLUNAS = order_columns_by_first_record(todos_registros[0] if todos_registros else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Analise Quantidades Produzidas'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{mes_ref}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Qtdes Produzidas"

hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, reg in enumerate(todos_registros, 2):
    for c_idx, (col_key, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=reg.get(col_key, '') or '')

ws.append([])
ws.append(['', '', 'TOTAL', '', '', fmt(total_custo), ''])

for i, w in enumerate([12, 10, 40, 10, 12, 14, 12], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Salvar JSON ────────────────────────────────────────────────────────────────
registros_nomeados = [
    {
        'data':        r.get('A4',''),
        'hora':        r.get('A5',''),
        'produto':     r.get('A0',''),
        'qtde':        r.get('A1',''),
        'custo_unit':  r.get('A2',''),
        'custo_total': r.get('A3',''),
        'turno':       r.get('A6',''),
    }
    for r in todos_registros
]
json_path = f"{pasta}/{mes_ref}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({
        "periodo": {"de": data_ini, "ate": data_fim},
        "total_registros": len(todos_registros),
        "custo_total": fmt(total_custo),
        "registros": registros_nomeados
    }, f, ensure_ascii=False, indent=2)
print(f"JSON salvo: {json_path}")

# ── Enviar email ───────────────────────────────────────────────────────────────
corpo = f"🏭 Análise de Quantidades Produzidas — {mes_nome}\n"
corpo += "=" * 50 + "\n\n"
corpo += f"📅 Período: {data_ini} a {data_fim}\n"
corpo += f"📦 Total de registros: {len(todos_registros)}\n"
corpo += f"💰 Custo total de produção: {fmt(total_custo)}\n\n"
corpo += "(BigDog 🐕)"

cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "🏭 Mogo Qtdes Produzidas — {mes_nome} — {fmt(total_custo)}" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}" '
    f'--attach "{json_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
