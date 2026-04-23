#!/usr/bin/env python3
"""
BigDog — Mogo "Taxa de Serviço" (Mensal)
Roda todo dia 1 e 2 de cada mês, qualquer horário.
Relatório sempre referente ao mês anterior.

Replica: Relatórios > Auditoria > Taxa de Serviço
  - Período: primeiro ao último dia do mês anterior

Endpoint: GET /relatorios/BuscaDadosRelatorioDinamico
  idGeradorRelatorios=0, codRelatorio=14

Campos: A0=Funcionário, A1=Taxa%, A2=Total Taxa, A3=Produto,
        A4=Comanda, A5=Nº Pedido, A6=Data, A7=Hora,
        A8=Qtde, A9=Tipo, A10=Turno
"""

import sys, os, subprocess, json, calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login
session = mogo_login()

# Período: mês anterior
hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]
data_ini = datetime(ano_ant, mes_ant, 1).strftime('%d/%m/%Y')
data_fim = datetime(ano_ant, mes_ant, ultimo_dia).strftime('%d/%m/%Y')
mes_ref  = datetime(ano_ant, mes_ant, 1).strftime('%m-%Y')
mes_nome = month_year_pt(datetime(ano_ant, mes_ant, 1))

print(f"Taxa de Serviço: {data_ini} a {data_fim}...")

filtro = (
    f"DataDe{{{data_ini}"
    f"|DataAte{{{data_fim}"
    f"|Funcionario{{"
    f"|TurnoTrabalho{{"
    f"|Mesa{{"
    f"|Cartao{{"
    f"|Taxa{{0"
    f"|ComboProdutos{{"
    f"|ProdutoEstabelecimento{{"
)

# Total
r_test = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
    'idGeradorRelatorios': '0', 'codRelatorio': '14',
    'filtro': filtro,
    'gridparamns': json.dumps({"Searching": True, "RecordsCount": 1, "PageIndex": 0, "SortingName": "", "SortingOrder": "ASC"}),
    'colunas': '[]', 'dbNameFranquia': ''
}, timeout=30)
total_records = int(r_test.json().get('records', 0))
print(f"Total de registros: {total_records}")

# Buscar em lotes
todos_registros = []
lote = 2000
pagina = 0

while True:
    r = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
        'idGeradorRelatorios': '0', 'codRelatorio': '14',
        'filtro': filtro,
        'gridparamns': json.dumps({
            "Searching": True, "RecordsCount": lote,
            "PageIndex": pagina, "SortingName": "", "SortingOrder": "ASC"
        }),
        'colunas': '[]', 'dbNameFranquia': ''
    }, timeout=90)

    rows = r.json().get('rows') or []
    todos_registros.extend(rows)
    print(f"  Lote {pagina+1}: {len(rows)} registros (total: {len(todos_registros)})")

    if len(rows) < lote or len(todos_registros) >= total_records:
        break
    pagina += 1

print(f"Total carregado: {len(todos_registros)}")

if not todos_registros:
    print("Nenhum dado encontrado.")
    sys.exit(0)

# Calcular total
total_taxa = 0.0
for reg in todos_registros:
    try:
        v = float(str(reg.get('A2','0')).replace('.','').replace(',','.').strip())
        total_taxa += v
    except:
        pass

def fmt(v):
    return f"R$ {v:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

# ── Salvar Excel ───────────────────────────────────────────────────────────────
COLUNAS = [
    ('A0',  'A0'),
    ('A1',  'A1'),
    ('A2',  'A2'),
    ('A3',  'A3'),
    ('A4',  'A4'),
    ('A5',  'A5'),
    ('A6',  'A6'),
    ('A7',  'A7'),
    ('A8',  'A8'),
    ('A9',  'A9'),
    ('A10', 'A10'),
]


COLUNAS = order_columns_by_first_record(todos_registros[0] if todos_registros else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Taxa Servico'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{mes_ref}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Planilha"

hfill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, reg in enumerate(todos_registros, 2):
    for c_idx, (col_key, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=reg.get(col_key, '') or '')

for i, w in enumerate([12, 8, 12, 35, 8, 8, 12, 20, 10, 12, 12], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Salvar JSON ────────────────────────────────────────────────────────────────
registros_nomeados = [
    {
        'data':        r.get('A6',''),
        'hora':        r.get('A7',''),
        'num_pedido':  r.get('A5',''),
        'produto':     r.get('A3',''),
        'qtde':        r.get('A8',''),
        'taxa_pct':    r.get('A1',''),
        'total_taxa':  r.get('A2',''),
        'funcionario': r.get('A0',''),
        'comanda':     r.get('A4',''),
        'tipo':        r.get('A9',''),
        'turno':       r.get('A10',''),
    }
    for r in todos_registros
]
json_path = f"{pasta}/{mes_ref}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({
        "periodo": {"de": data_ini, "ate": data_fim},
        "total_registros": len(todos_registros),
        "total_taxa": fmt(total_taxa),
        "registros": registros_nomeados
    }, f, ensure_ascii=False, indent=2)
print(f"JSON salvo: {json_path}")

# ── Enviar email ───────────────────────────────────────────────────────────────
corpo = f"🔍 Taxa de Serviço — {mes_nome}\n"
corpo += "=" * 50 + "\n\n"
corpo += f"📅 Período: {data_ini} a {data_fim}\n"
corpo += f"📦 Total de registros: {len(todos_registros)}\n"
corpo += f"💰 Total de taxa: {fmt(total_taxa)}\n\n"
corpo += "(BigDog 🐕)"

cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "🔍 Mogo Taxa de Serviço — {mes_nome} — {fmt(total_taxa)}" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}" '
    f'--attach "{json_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
