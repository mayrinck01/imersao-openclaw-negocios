#!/usr/bin/env python3
"""
BigDog — Mogo "Pendentes" Report
Roda todo dia às 08:08 BRT.
Gera resumo de pedidos agendados (status Pendente) por data/hora e envia por email.
Envia ao Cake Atendimento um alerta dos pedidos com data operacional vencida.
"""

import sys, os, json
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import excel_safe_value, order_columns_by_records, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from gog_mail import send_gmail
from mogo_login import mogo_login, MOGO_URL
from mogo_pendentes_alerts import (
    CAKE_ATENDIMENTO_GROUP,
    build_overdue_alert_message,
    build_pending_email_body,
    build_pending_email_subject,
    overdue_pending_orders,
    send_whatsapp_group_alerts,
    sort_pending_orders,
)

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login com novo fluxo de 3 etapas
session = mogo_login()

# Buscar pedidos com status Pendente via API jqGrid
print("Buscando Pendentes...")
r = session.get(f"{MOGO_URL}/Pedido/ListPedidosParaEntrega", params={
    '_search': 'true',
    'rows': '500',
    'page': '1',
    'sidx': 'DataEntrega',
    'sord': 'asc',
    'filters': json.dumps({
        "groupOp": "AND",
        "rules": [{"field": "StatusEntrega", "op": "eq", "data": "Pendente"}]
    })
}, timeout=30)

if r.status_code != 200:
    print(f"ERRO ao buscar pedidos ({r.status_code})")
    sys.exit(1)

try:
    data = r.json()
    todos_pedidos = data.get('rows', [])
except Exception as e:
    print(f"ERRO ao parsear JSON: {e}")
    sys.exit(1)

# Filtrar só os pendentes (redundante, mas garante)
pedidos = [p for p in todos_pedidos if p.get('StatusEntrega','').lower() == 'pendente']
pedidos = sort_pending_orders(pedidos)

hoje = datetime.now().strftime('%d-%m-%Y')
hoje_br = datetime.now().strftime('%d/%m/%Y')
hoje_date = datetime.now().date()
print(f"Pendentes encontrados: {len(pedidos)}")

if not pedidos:
    print("Nenhum pedido pendente.")
    sys.exit(0)

# Agrupar por data e hora
por_data_hora = defaultdict(lambda: defaultdict(list))
for p in pedidos:
    data_ent = p.get('DataEntrega', 'Sem data')
    hora     = p.get('HoraEntregaTxt', 'Sem hora') or 'Sem hora'
    try:
        h = hora.split(':')[0].zfill(2)
        hora_grp = f"{h}:00"
    except:
        hora_grp = hora
    por_data_hora[data_ent][hora_grp].append(p)

pedidos_atrasados = overdue_pending_orders(pedidos, today=hoje_date)
if pedidos_atrasados:
    print(f"ALERTA: {len(pedidos_atrasados)} pendente(s) com data vencida")

# Salvar Excel
COLUNAS = [
    ('NumeroPedido',   'Nº Pedido'),
    ('NomeCliente',    'Cliente'),
    ('DataEntrega',    'Data Entrega'),
    ('HoraEntregaTxt', 'Hora Entrega'),
    ('Logradouro',     'Logradouro'),
    ('Bairro',         'Bairro'),
    ('Numero',         'Número'),
    ('ValorFinal',     'Valor Final'),
    ('StatusPago',     'Pago'),
    ('OrigemPedido',   'Origem'),
]


COLUNAS = order_columns_by_records(pedidos, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Pendentes'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{hoje}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = pt_title("Pendentes")

hfill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, pedido in enumerate(pedidos, 2):
    for c_idx, (col_name, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=excel_safe_value(pedido.get(col_name, '')))

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel: {xlsx_path}")

# Montar email
total = len(pedidos)
corpo = build_pending_email_body(
    pedidos=pedidos,
    grouped_by_date_hour=por_data_hora,
    today_label=hoje,
    overdue_orders=pedidos_atrasados,
)
subject = build_pending_email_subject(
    total=total,
    today_label=hoje,
    overdue_count=len(pedidos_atrasados),
)

# Enviar email
res = send_gmail(
    account="cakebigdog@gmail.com",
    client="cakebigdog",
    to="joao@cakeco.com.br",
    subject=subject,
    body=corpo,
    attach=xlsx_path,
    env=env,
)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")

if pedidos_atrasados:
    alert = build_overdue_alert_message(pedidos_atrasados, today_label=hoje_br)
    whatsapp_results = send_whatsapp_group_alerts(
        alert,
        targets=[CAKE_ATENDIMENTO_GROUP],
    )
    whatsapp_ok = [result for result in whatsapp_results if result.get("ok")]
    whatsapp_failed = [result for result in whatsapp_results if not result.get("ok")]
    if whatsapp_ok:
        print(f"✅ Alerta WhatsApp enviado para {len(whatsapp_ok)} grupo(s) operacional(is)")
    if whatsapp_failed:
        failed_targets = ", ".join(str(result.get("target")) for result in whatsapp_failed)
        print(f"AVISO WhatsApp grupos não entregues: {failed_targets}")
