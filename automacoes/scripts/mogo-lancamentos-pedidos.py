#!/usr/bin/env python3
"""
BigDog — Mogo "Lançamentos de Pedidos" (Mensal)
Roda todo dia 1 e 2 de cada mês, qualquer horário.
Relatório sempre referente ao mês anterior.

Replica: Relatórios > Auditoria > Lançamentos de Pedidos
  - Período: primeiro ao último dia do mês anterior
  - Estratégia: busca semana a semana (relatório pesado, ~700 registros/dia)

Endpoint: GET /relatorios/BuscaDadosRelatorioDinamico
  idGeradorRelatorios=0, codRelatorio=115

Campos: A0=Data, A1=Hora, A2=Produto, A3=Qtde, A4=Valor, A5=Cliente,
        A6=Tipo Pedido, A8=Funcionário, A10=Status, A13=Nº Pedido,
        OrigemPedido, A12=Desconto, A11=Acréscimo, A16=ID
"""

import sys, os, subprocess, json, calendar
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login
session = mogo_login()

# Período: mês anterior
hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]
primeiro = datetime(ano_ant, mes_ant, 1)
ultimo   = datetime(ano_ant, mes_ant, ultimo_dia)
mes_ref  = primeiro.strftime('%m-%Y')
mes_nome = month_year_pt(primeiro)

print(f"Lançamentos de Pedidos: {primeiro.strftime('%d/%m/%Y')} a {ultimo.strftime('%d/%m/%Y')}...")

# Buscar semana a semana para evitar timeout
todos_registros = []
cursor = primeiro

while cursor <= ultimo:
    fim_semana = min(cursor + timedelta(days=6), ultimo)
    data_ini_str = cursor.strftime('%d/%m/%Y')
    data_fim_str = fim_semana.strftime('%d/%m/%Y')

    filtro = (
        f"idped{{|DataDe{{{data_ini_str}"
        f"|DataAte{{{data_fim_str}"
        f"|TipoPedido{{|Cartao{{|Mesa{{"
        f"|ProdutoEstabelecimento{{|OrigemPedido{{"
        f"|ComboProdutos{{|ClienteCr{{|HoraDe{{|HoraAte{{"
    )

    pagina = 0
    lote = 2000

    while True:
        try:
            r = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
                'idGeradorRelatorios': '0', 'codRelatorio': '115',
                'filtro': filtro,
                'gridparamns': json.dumps({
                    "Searching": True, "RecordsCount": lote,
                    "PageIndex": pagina, "SortingName": "", "SortingOrder": "ASC"
                }),
                'colunas': '[]', 'dbNameFranquia': ''
            }, timeout=180)

            d = r.json()
            rows = d.get('rows') or []
            todos_registros.extend(rows)

            total_semana = int(d.get('records', 0))
            print(f"  {data_ini_str}–{data_fim_str} lote {pagina+1}: {len(rows)} | total semana: {total_semana}")

            if len(rows) < lote or len([x for x in todos_registros if x.get('A0','') >= data_ini_str]) >= total_semana:
                break
            pagina += 1

        except Exception as e:
            print(f"  ERRO semana {data_ini_str}: {e}")
            break

    cursor = fim_semana + timedelta(days=1)

print(f"Total carregado: {len(todos_registros)}")

if not todos_registros:
    print("Nenhum dado encontrado.")
    sys.exit(0)

# Calcular total
total_valor = 0.0
for reg in todos_registros:
    try:
        v = float(str(reg.get('A4','0')).replace('.','').replace(',','.').strip())
        total_valor += v
    except:
        pass

def fmt(v):
    return f"R$ {v:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

# ── Salvar Excel ───────────────────────────────────────────────────────────────
COLUNAS = [
    ('A0',          'Data'),
    ('A1',          'Hora'),
    ('A13',         'Nº Pedido'),
    ('A5',          'Cliente'),
    ('A2',          'Produto'),
    ('A3',          'Qtde'),
    ('A4',          'Valor'),
    ('A12',         'Desconto'),
    ('A11',         'Acréscimo'),
    ('OrigemPedido','Origem'),
    ('A6',          'Tipo Pedido'),
    ('A10',         'Status'),
    ('A8',          'Funcionário'),
]


COLUNAS = order_columns_by_first_record(todos_registros[0] if todos_registros else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Lancamentos Pedidos'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{mes_ref}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = pt_title("Lançamentos de Pedidos")

hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, reg in enumerate(todos_registros, 2):
    for c_idx, (col_key, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=reg.get(col_key, '') or '')

for i, w in enumerate([12, 8, 12, 28, 30, 8, 12, 12, 12, 12, 14, 10, 18], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Salvar JSON ────────────────────────────────────────────────────────────────
json_path = f"{pasta}/{mes_ref}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({
        "periodo": {"de": primeiro.strftime('%d/%m/%Y'), "ate": ultimo.strftime('%d/%m/%Y')},
        "total_registros": len(todos_registros),
        "faturamento_total": fmt(total_valor),
        "registros": todos_registros
    }, f, ensure_ascii=False, indent=2)
print(f"JSON salvo: {json_path}")

# ── Enviar email ───────────────────────────────────────────────────────────────
corpo = f"📋 Lançamentos de Pedidos — {mes_nome}\n"
corpo += "=" * 50 + "\n\n"
corpo += f"📅 Período: {primeiro.strftime('%d/%m/%Y')} a {ultimo.strftime('%d/%m/%Y')}\n"
corpo += f"📦 Total de lançamentos: {len(todos_registros)}\n"
corpo += f"💰 Valor total: {fmt(total_valor)}\n\n"
corpo += "(BigDog 🐕)"

cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "📋 Mogo Lançamentos de Pedidos — {mes_nome} — {fmt(total_valor)}" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}" '
    f'--attach "{json_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
