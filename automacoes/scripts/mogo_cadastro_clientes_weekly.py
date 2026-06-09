#!/usr/bin/env python3
"""Snapshot semanal do cadastro de clientes do Mogo e análise de qualidade."""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from mogo_login import MOGO_URL, mogo_login  # noqa: E402


DEFAULT_OUTPUT_DIR = Path("/root/workspaces/cake-brain/relatorios/Mogo/Cadastro Clientes Semanal")
DEFAULT_TELEGRAM_TARGET = "968564677"
CUSTOMER_TYPE = "kardapio.Suprimentos.Cliente"

FIELD_MAP = {
    "Id": "id",
    "Nome": "nome",
    "RazaoSocial": "razao_social",
    "Fantasia": "fantasia",
    "Pessoa_Nome": "pessoa_nome",
    "CNPJ_CPF": "documento",
    "Email": "email",
    "TelefoneCelular": "telefone_celular",
    "TelefoneResidencial": "telefone_residencial",
    "TelefoneComercial": "telefone_comercial",
    "Logradouro": "logradouro",
    "Numero": "numero",
    "Complemento": "complemento",
    "Bairro": "bairro",
    "CEP": "cep",
    "Cidade": "cidade",
    "DescricaoCidade": "descricao_cidade",
    "UFCidade": "uf",
    "Nascimento": "nascimento",
    "DataCadastro": "data_cadastro",
    "DataCadastroCli": "data_cadastro_cliente",
    "DataUltimoPedido": "data_ultimo_pedido",
    "QuantidadeDelivery": "quantidade_delivery",
    "QuantidadeVisitas": "quantidade_visitas",
    "ValorConsumido": "valor_consumido",
    "OrigemCadastro": "origem_cadastro",
    "ComoConheceu": "como_conheceu",
    "Ativo": "ativo",
    "TipoCliente": "tipo_cliente",
    "TipoPessoa": "tipo_pessoa",
    "TagsCRMCliente": "tags_crm_cliente",
}

BLOCKED_FIELD_PATTERNS = [
    "senha",
    "password",
    "token",
    "cartao",
    "cartoes",
    "providerkey",
    "provider_key",
    "firebase",
    "malga",
]

TRACKED_CHANGE_FIELDS = [
    "nome",
    "razao_social",
    "documento",
    "email",
    "telefone_celular",
    "telefone_residencial",
    "telefone_comercial",
    "logradouro",
    "numero",
    "complemento",
    "bairro",
    "cep",
    "cidade",
    "uf",
    "nascimento",
    "ativo",
    "tipo_cliente",
]

CONTACT_FIELDS = [
    "email",
    "telefone_celular",
    "telefone_residencial",
    "telefone_comercial",
]


def camel_to_snake(name: str) -> str:
    name = re.sub(r"[^0-9A-Za-z]+", "_", clean_text(name))
    name = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", name)
    name = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", name)
    normalized = re.sub(r"_+", "_", name).strip("_").lower()
    return (
        normalized.replace("_i_food", "_ifood")
        .replace("_a_i_qfome", "_aiqfome")
        .replace("_c_e_ofood", "_ceofood")
        .replace("_r_g", "_rg")
    )


def is_blocked_field(name: str) -> bool:
    key = camel_to_snake(name).replace("_", "")
    raw = clean_text(name).lower().replace("_", "")
    return any(pattern.replace("_", "") in key or pattern.replace("_", "") in raw for pattern in BLOCKED_FIELD_PATTERNS)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def only_digits(value: Any) -> str:
    return re.sub(r"\D+", "", clean_text(value))


def sanitize_customer(raw: dict[str, Any]) -> dict[str, Any]:
    sanitized = {out_key: clean_text(raw.get(in_key)) for in_key, out_key in FIELD_MAP.items()}
    mapped_inputs = set(FIELD_MAP)
    for raw_key, value in raw.items():
        if raw_key in mapped_inputs or is_blocked_field(raw_key):
            continue
        out_key = camel_to_snake(raw_key)
        if not out_key or out_key in sanitized:
            continue
        if isinstance(value, (dict, list)):
            sanitized[out_key] = json.dumps(value, ensure_ascii=False)
        else:
            sanitized[out_key] = clean_text(value)
    sanitized["id"] = clean_text(raw.get("Id") or raw.get("id"))
    sanitized["ativo"] = raw.get("Ativo") if isinstance(raw.get("Ativo"), bool) else clean_text(raw.get("Ativo"))
    return sanitized


def valid_email(email: str) -> bool:
    if not email:
        return False
    return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email) is not None


def valid_phone(phone: str) -> bool:
    digits = only_digits(phone)
    if len(digits) < 10 or len(digits) > 13:
        return False
    return len(set(digits)) > 2


def valid_cpf(digits: str) -> bool:
    if len(digits) != 11 or len(set(digits)) == 1:
        return False
    first = sum(int(digits[i]) * (10 - i) for i in range(9))
    d1 = 11 - (first % 11)
    d1 = 0 if d1 >= 10 else d1
    second = sum(int(digits[i]) * (11 - i) for i in range(10))
    d2 = 11 - (second % 11)
    d2 = 0 if d2 >= 10 else d2
    return digits[-2:] == f"{d1}{d2}"


def valid_cnpj(digits: str) -> bool:
    if len(digits) != 14 or len(set(digits)) == 1:
        return False
    weights_1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    weights_2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    total_1 = sum(int(digits[i]) * weights_1[i] for i in range(12))
    d1 = 11 - (total_1 % 11)
    d1 = 0 if d1 >= 10 else d1
    total_2 = sum(int(digits[i]) * weights_2[i] for i in range(13))
    d2 = 11 - (total_2 % 11)
    d2 = 0 if d2 >= 10 else d2
    return digits[-2:] == f"{d1}{d2}"


def valid_document(document: str) -> bool:
    digits = only_digits(document)
    if not digits:
        return False
    if len(digits) == 11:
        return valid_cpf(digits)
    if len(digits) == 14:
        return valid_cnpj(digits)
    return False


def normalize_name(name: str) -> str:
    cleaned = re.sub(r"[^a-z0-9 ]+", "", clean_text(name).lower())
    return re.sub(r"\s+", " ", cleaned).strip()


def redacted_email(email: str) -> str:
    email = clean_text(email)
    if "@" not in email:
        return "***"
    local, domain = email.split("@", 1)
    local_part = f"{local[:1]}***" if len(local) > 1 else local[:1]
    pieces = domain.split(".")
    if len(pieces) >= 2:
        domain_part = f"{pieces[0][:1]}***.{pieces[-1]}"
    else:
        domain_part = f"{domain[:1]}***"
    return f"{local_part}@{domain_part}"


def redacted_phone(phone: str) -> str:
    digits = only_digits(phone)
    if not digits:
        return "***"
    return f"***{digits[-4:]}"


def redacted_document(document: str) -> str:
    return "***" if clean_text(document) else ""


def redact_value(field: str, value: Any) -> Any:
    text = clean_text(value)
    if field == "email":
        return redacted_email(text)
    if field.startswith("telefone"):
        return redacted_phone(text)
    if field in {"documento", "rg"}:
        return redacted_document(text)
    return text


def public_record(record: dict[str, Any]) -> dict[str, Any]:
    public = dict(record)
    for field in ["email", "telefone_celular", "telefone_residencial", "telefone_comercial", "documento"]:
        if field in public:
            public[field] = redact_value(field, public[field])
    return public


def dirty_reasons(record: dict[str, Any]) -> list[str]:
    reasons: list[str] = []
    name = clean_text(record.get("nome") or record.get("razao_social") or record.get("pessoa_nome"))
    phones = [
        clean_text(record.get("telefone_celular")),
        clean_text(record.get("telefone_residencial")),
        clean_text(record.get("telefone_comercial")),
    ]
    email = clean_text(record.get("email"))
    document = clean_text(record.get("documento"))

    if len(normalize_name(name)) < 3 or normalize_name(name) in {"teste", "cliente", "sem nome", "nao informado"}:
        reasons.append("nome_suspeito")
    if not any(phones):
        reasons.append("sem_telefone")
    elif not any(valid_phone(phone) for phone in phones if phone):
        reasons.append("telefone_invalido")
    if email and not valid_email(email):
        reasons.append("email_invalido")
    if not document:
        reasons.append("sem_documento")
    elif not valid_document(document):
        reasons.append("documento_invalido")
    if not all(clean_text(record.get(field)) for field in ["logradouro", "bairro", "numero"]):
        reasons.append("endereco_incompleto")
    return reasons


def improvement_reasons(record: dict[str, Any]) -> list[str]:
    reasons: list[str] = []
    if not clean_text(record.get("nascimento")):
        reasons.append("sem_aniversario")
    if not clean_text(record.get("email")):
        reasons.append("sem_email")
    if not clean_text(record.get("bairro")):
        reasons.append("sem_bairro")
    return reasons


def duplicate_groups(records: list[dict[str, Any]], field_name: str, normalizer) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        key = normalizer(record.get(field_name))
        if key:
            groups[key].append(record)
    duplicates = []
    for key, items in sorted(groups.items()):
        if len(items) > 1:
            duplicates.append(
                {
                    "key": "***" if field_name in {"documento", "telefone_celular"} else key,
                    "ids": [clean_text(item.get("id")) for item in items],
                    "nomes": [clean_text(item.get("nome")) for item in items],
                }
            )
    return duplicates


def analyze_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    dirty_counts: dict[str, int] = defaultdict(int)
    improvement_counts: dict[str, int] = defaultdict(int)
    dirty_records = []
    improvement_records = []

    for record in records:
        reasons = dirty_reasons(record)
        if reasons:
            for reason in reasons:
                dirty_counts[reason] += 1
            item = public_record(record)
            item["motivos"] = reasons
            dirty_records.append(item)

        improvements = improvement_reasons(record)
        if improvements:
            for reason in improvements:
                improvement_counts[reason] += 1
            item = public_record(record)
            item["melhorias"] = improvements
            improvement_records.append(item)

    duplicates = {
        "telefone": duplicate_groups(records, "telefone_celular", only_digits),
        "documento": duplicate_groups(records, "documento", only_digits),
    }

    return {
        "total_clientes": len(records),
        "dirty_counts": dict(dirty_counts),
        "dirty_records_count": len(dirty_records),
        "dirty_records": dirty_records,
        "improvement_counts": dict(improvement_counts),
        "improvement_records_count": len(improvement_records),
        "improvement_records": improvement_records,
        "duplicate_counts": {key: len(value) for key, value in duplicates.items()},
        "duplicates": duplicates,
    }


def record_index(records: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {clean_text(record.get("id")): record for record in records if clean_text(record.get("id"))}


def compare_snapshots(previous: list[dict[str, Any]], current: list[dict[str, Any]]) -> dict[str, Any]:
    if not previous:
        return {
            "initial_baseline": True,
            "new_count": 0,
            "removed_count": 0,
            "changed_count": 0,
            "new_records": [],
            "removed_records": [],
            "changed_records": [],
        }

    previous_by_id = record_index(previous)
    current_by_id = record_index(current)
    previous_ids = set(previous_by_id)
    current_ids = set(current_by_id)

    new_records = [public_record(current_by_id[item_id]) for item_id in sorted(current_ids - previous_ids)]
    removed_records = [public_record(previous_by_id[item_id]) for item_id in sorted(previous_ids - current_ids)]
    changed_records = []

    for item_id in sorted(previous_ids & current_ids):
        old = previous_by_id[item_id]
        new = current_by_id[item_id]
        changes = []
        for field in TRACKED_CHANGE_FIELDS:
            old_value = old.get(field)
            new_value = new.get(field)
            if clean_text(old_value) != clean_text(new_value):
                changes.append(
                    {
                        "field": field,
                        "old": redact_value(field, old_value),
                        "new": redact_value(field, new_value),
                    }
                )
        if changes:
            changed_records.append({"id": item_id, "nome": clean_text(new.get("nome")), "changes": changes})

    return {
        "initial_baseline": False,
        "new_count": len(new_records),
        "removed_count": len(removed_records),
        "changed_count": len(changed_records),
        "new_records": new_records,
        "removed_records": removed_records,
        "changed_records": changed_records,
    }


def fetch_customer_records(limit: int | None = None) -> list[dict[str, Any]]:
    session = mogo_login()
    page_size = min(limit or 2000, 2000)
    records: list[dict[str, Any]] = []
    page = 1
    total_records: int | None = None

    while True:
        params = {
            "cTipo": CUSTOMER_TYPE,
            "_search": "false",
            "rows": str(page_size),
            "page": str(page),
            "sidx": "Id",
            "sord": "asc",
        }
        response = session.get(f"{MOGO_URL}/Cadastros/GenerioDadosJqGrid", params=params, timeout=90)
        response.raise_for_status()
        data = response.json()
        if total_records is None:
            total_records = int(data.get("records") or 0)
        rows = data.get("rows") or []
        records.extend(sanitize_customer(row) for row in rows)
        if limit and len(records) >= limit:
            return records[:limit]
        if not rows or len(records) >= total_records:
            return records
        page += 1


def load_snapshot(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return data.get("records", [])


def previous_snapshot_path(output_dir: Path, current_path: Path) -> Path | None:
    candidates = sorted(output_dir.glob("*-snapshot.json"))
    candidates = [path for path in candidates if path != current_path]
    return candidates[-1] if candidates else None


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_markdown(path: Path, run_date: date, analysis: dict[str, Any], comparison: dict[str, Any]) -> None:
    lines = [
        f"# Cadastro de Clientes Mogo — {run_date:%d/%m/%Y}",
        "",
        f"- Total de clientes: {analysis['total_clientes']}",
        f"- Cadastros sujos: {analysis['dirty_records_count']}",
        f"- Oportunidades de melhoria: {analysis['improvement_records_count']}",
        f"- Novos desde snapshot anterior: {comparison['new_count']}",
        f"- Removidos/inativados desde snapshot anterior: {comparison['removed_count']}",
        f"- Alterados desde snapshot anterior: {comparison['changed_count']}",
        "",
        "## Principais Sujeiras",
    ]
    for reason, count in sorted(analysis["dirty_counts"].items(), key=lambda item: item[1], reverse=True):
        lines.append(f"- {reason}: {count}")
    lines.append("")
    lines.append("## Melhorias")
    for reason, count in sorted(analysis["improvement_counts"].items(), key=lambda item: item[1], reverse=True):
        lines.append(f"- {reason}: {count}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def add_sheet(workbook: openpyxl.Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws = workbook.create_sheet(title=title[:31])
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for index, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=index, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_index, row in enumerate(rows, 2):
        for col_index, column in enumerate(columns, 1):
            value = row.get(column, "")
            if isinstance(value, list):
                value = ", ".join(str(item) for item in value)
            ws.cell(row=row_index, column=col_index, value=value)
    for col_index, column in enumerate(columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = min(max(len(column) + 4, 14), 45)


def rows_with_full_contact_fields(
    rows: list[dict[str, Any]],
    source_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    source_by_id = record_index(source_records)
    enriched = []
    for row in rows:
        item = dict(row)
        source = source_by_id.get(clean_text(item.get("id")))
        if source:
            for field in CONTACT_FIELDS:
                if field in item:
                    item[field] = clean_text(source.get(field))
        enriched.append(item)
    return enriched


def changed_rows(
    changed_records: list[dict[str, Any]],
    previous_records: list[dict[str, Any]] | None = None,
    current_records: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    previous_by_id = record_index(previous_records or [])
    current_by_id = record_index(current_records or [])
    rows = []
    for record in changed_records:
        for change in record["changes"]:
            field = change["field"]
            old_value = change["old"]
            new_value = change["new"]
            if field in CONTACT_FIELDS:
                previous = previous_by_id.get(record["id"], {})
                current = current_by_id.get(record["id"], {})
                old_value = clean_text(previous.get(field))
                new_value = clean_text(current.get(field))
            rows.append(
                {
                    "id": record["id"],
                    "nome": record["nome"],
                    "campo": field,
                    "antes": old_value,
                    "depois": new_value,
                }
            )
    return rows


def all_columns(records: list[dict[str, Any]]) -> list[str]:
    preferred = [
        "id",
        "nome",
        "razao_social",
        "fantasia",
        "documento",
        "email",
        "telefone_celular",
        "telefone_residencial",
        "telefone_comercial",
        "logradouro",
        "numero",
        "complemento",
        "bairro",
        "cep",
        "cidade",
        "uf",
        "nascimento",
        "data_cadastro",
        "data_ultimo_pedido",
        "ativo",
    ]
    present = set()
    for record in records:
        present.update(record.keys())
    ordered = [column for column in preferred if column in present]
    ordered.extend(sorted(present - set(ordered)))
    return ordered


def write_excel(
    path: Path,
    records: list[dict[str, Any]],
    analysis: dict[str, Any],
    comparison: dict[str, Any],
    previous_records: list[dict[str, Any]] | None = None,
) -> None:
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary_rows = [
        ("total_clientes", analysis["total_clientes"]),
        ("cadastros_sujos", analysis["dirty_records_count"]),
        ("melhorias", analysis["improvement_records_count"]),
        ("novos", comparison["new_count"]),
        ("removidos", comparison["removed_count"]),
        ("alterados", comparison["changed_count"]),
        ("duplicidades_telefone", analysis["duplicate_counts"]["telefone"]),
        ("duplicidades_documento", analysis["duplicate_counts"]["documento"]),
    ]
    for row_index, (metric, value) in enumerate(summary_rows, 1):
        summary.cell(row=row_index, column=1, value=metric)
        summary.cell(row=row_index, column=2, value=value)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 16

    add_sheet(workbook, "Cadastro Completo", records, all_columns(records))

    public_columns = [
        "id",
        "nome",
        "razao_social",
        "email",
        "telefone_celular",
        "telefone_residencial",
        "telefone_comercial",
        "documento",
        "logradouro",
        "numero",
        "bairro",
        "nascimento",
    ]
    previous_records = previous_records or []
    add_sheet(
        workbook,
        "Cadastro Sujo",
        rows_with_full_contact_fields(analysis["dirty_records"], records),
        public_columns + ["motivos"],
    )
    add_sheet(
        workbook,
        "Melhorias",
        rows_with_full_contact_fields(analysis["improvement_records"], records),
        public_columns + ["melhorias"],
    )
    add_sheet(
        workbook,
        "Alteracoes",
        changed_rows(comparison["changed_records"], previous_records, records),
        ["id", "nome", "campo", "antes", "depois"],
    )
    add_sheet(workbook, "Novos", rows_with_full_contact_fields(comparison["new_records"], records), public_columns)
    add_sheet(
        workbook,
        "Removidos",
        rows_with_full_contact_fields(comparison["removed_records"], previous_records),
        public_columns,
    )
    add_sheet(workbook, "Duplicidades Telefone", analysis["duplicates"]["telefone"], ["key", "ids", "nomes"])
    add_sheet(workbook, "Duplicidades Documento", analysis["duplicates"]["documento"], ["key", "ids", "nomes"])

    workbook.save(path)


def build_summary_message(run_date: date, analysis: dict[str, Any], comparison: dict[str, Any], xlsx_path: str) -> str:
    lines = [
        f"Cadastro de clientes Mogo — {run_date:%d/%m/%Y}",
        "",
        f"- {analysis['total_clientes']} clientes no snapshot",
        f"- {analysis['dirty_records_count']} com cadastro sujo",
        f"- {analysis['improvement_records_count']} com melhoria possível",
    ]
    if comparison.get("initial_baseline"):
        lines.append("- baseline inicial criado; comparação começa no próximo ciclo")
    else:
        lines.extend(
            [
                f"- {comparison['new_count']} novos",
                f"- {comparison['removed_count']} removidos/inativados",
                f"- {comparison['changed_count']} alterados",
            ]
        )
    lines.extend(["", f"Arquivo: {Path(xlsx_path).name}"])
    return "\n".join(lines)


def send_telegram(message: str, media_path: Path | None, target: str = DEFAULT_TELEGRAM_TARGET) -> None:
    cmd = [
        "openclaw",
        "message",
        "send",
        "--channel",
        "telegram",
        "--target",
        target,
        "--message",
        message,
    ]
    if media_path:
        cmd.extend(["--media", str(media_path), "--force-document"])
    subprocess.run(cmd, check=True)


def run_weekly(output_dir: Path, run_date: date, *, limit: int | None = None, send: bool = True) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = run_date.isoformat()
    snapshot_path = output_dir / f"{stamp}-snapshot.json"
    analysis_path = output_dir / f"{stamp}-analise.json"
    md_path = output_dir / f"{stamp}-resumo.md"
    xlsx_path = output_dir / f"{stamp}-analise.xlsx"

    records = fetch_customer_records(limit=limit)
    previous_path = previous_snapshot_path(output_dir, snapshot_path)
    previous_records = load_snapshot(previous_path) if previous_path else []
    analysis = analyze_records(records)
    comparison = compare_snapshots(previous_records, records)

    write_json(
        snapshot_path,
        {
            "run_date": stamp,
            "source": "Cadastros/GenerioDadosJqGrid",
            "customer_type": CUSTOMER_TYPE,
            "records": records,
        },
    )
    write_json(
        analysis_path,
        {
            "run_date": stamp,
            "previous_snapshot": str(previous_path) if previous_path else None,
            "analysis": analysis,
            "comparison": comparison,
        },
    )
    write_markdown(md_path, run_date, analysis, comparison)
    write_excel(xlsx_path, records, analysis, comparison, previous_records=previous_records)

    message = build_summary_message(run_date, analysis, comparison, str(xlsx_path))
    if send:
        send_telegram(message, xlsx_path)
    return {
        "snapshot_path": str(snapshot_path),
        "analysis_path": str(analysis_path),
        "markdown_path": str(md_path),
        "xlsx_path": str(xlsx_path),
        "analysis": analysis,
        "comparison": comparison,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analisa semanalmente o cadastro de clientes do Mogo.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--run-date", default=date.today().isoformat())
    parser.add_argument("--limit", type=int, default=None, help="Limita registros para teste manual.")
    parser.add_argument("--no-send", action="store_true", help="Nao envia Telegram.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    run_date = datetime.strptime(args.run_date, "%Y-%m-%d").date()
    result = run_weekly(args.output_dir, run_date, limit=args.limit, send=not args.no_send)
    print(json.dumps({k: v for k, v in result.items() if k.endswith("_path")}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
