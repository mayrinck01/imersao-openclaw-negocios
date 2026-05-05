#!/usr/bin/env python3
"""
BigDog — Mogo "Faturamento Detalhado" (Mensal)
Roda todo dia 1 e 2 de cada mês, qualquer horário.
Relatório sempre referente ao mês anterior.

Replica: Relatórios > Faturamento > Faturamento Detalhado
  - Período: primeiro ao último dia do mês anterior
  - Sem filtros adicionais (TipoPagamento, TipoPedido, Turno = vazios)
  - Tipo de Faturamento: Bruto (TipoFat=0)

Endpoint: GET /relatorios/BuscaDadosRelatorioDinamico
  idGeradorRelatorios=0, codRelatorio=58
"""

import sys, os, subprocess, json, calendar, time
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login
session = mogo_login()

# Calcular período: mês anterior
hoje = datetime.now()
mes_ant = hoje.month - 1 if hoje.month > 1 else 12
ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
ultimo_dia = calendar.monthrange(ano_ant, mes_ant)[1]
primeiro = datetime(ano_ant, mes_ant, 1)
ultimo = datetime(ano_ant, mes_ant, ultimo_dia)
data_ini = primeiro.strftime('%d/%m/%Y')
data_fim = ultimo.strftime('%d/%m/%Y')
mes_ref = primeiro.strftime('%m-%Y')

print(f"Faturamento Detalhado: {data_ini} a {data_fim}...")

# ── Buscar todos os registros ─────────────────────────────────────────────────
# Estratégia: semana a semana + paginação menor.
# O relatório mensal inteiro em lotes de 5000 pode estourar timeout do Mogo.
todos_registros = []
cursor = primeiro

while cursor <= ultimo:
    fim_semana = min(cursor + timedelta(days=6), ultimo)
    data_ini_semana = cursor.strftime('%d/%m/%Y')
    data_fim_semana = fim_semana.strftime('%d/%m/%Y')

    filtro = (
        f"DataDe{{{data_ini_semana}"
        f"|DataAte{{{data_fim_semana}"
        f"|TipoPagamento{{"
        f"|TipoFat{{0"
        f"|TipoPedido{{"
        f"|Turno{{"
    )

    pagina = 0
    lote = 2000

    while True:
        tentativa = 0
        while True:
            tentativa += 1
            try:
                r = session.get(f"{MOGO_URL}/relatorios/BuscaDadosRelatorioDinamico", params={
                    'idGeradorRelatorios': '0',
                    'codRelatorio': '58',
                    'filtro': filtro,
                    'gridparamns': json.dumps({
                        "Searching": True,
                        "RecordsCount": lote,
                        "PageIndex": pagina,
                        "SortingName": "",
                        "SortingOrder": "ASC"
                    }),
                    'colunas': '[]',
                    'dbNameFranquia': ''
                }, timeout=180)
                d = r.json()
                rows = d.get('rows') or []
                break
            except Exception as e:
                if tentativa >= 3:
                    print(f"  ERRO {data_ini_semana}–{data_fim_semana} lote {pagina+1}: {e}")
                    rows = []
                    d = {}
                    break
                print(f"  Retry {tentativa}/3 {data_ini_semana}–{data_fim_semana} lote {pagina+1}...")
                time.sleep(2 * tentativa)

        todos_registros.extend(rows)
        total_semana = int(d.get('records', 0) or 0)
        print(
            f"  {data_ini_semana}–{data_fim_semana} lote {pagina+1}: "
            f"{len(rows)} registros (acumulado: {len(todos_registros)} | total semana: {total_semana})"
        )

        if len(rows) < lote:
            break
        pagina += 1

    cursor = fim_semana + timedelta(days=1)

print(f"Total carregado: {len(todos_registros)}")

if not todos_registros:
    print("Nenhum registro encontrado.")
    sys.exit(0)

# ── Calcular totais ────────────────────────────────────────────────────────────
total_valor = 0.0
for reg in todos_registros:
    try:
        val = str(reg.get('val', '0')).replace('.', '').replace(',', '.').strip()
        total_valor += float(val)
    except:
        pass

total_fmt = f"R$ {total_valor:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

# ── Salvar Excel ───────────────────────────────────────────────────────────────
COLUNAS = [
    ('dt',      'Data'),
    ('hr',      'Hora'),
    ('sem',     'Dia da Semana'),
    ('val',     'Valor'),
    ('tp',      'Pagamento'),
    ('func',    'Funcionário'),
    ('cliente', 'Cliente'),
    ('origem',  'Origem'),
    ('turn',    'Turno'),
]


COLUNAS = order_columns_by_first_record(todos_registros[0] if todos_registros else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Faturamento Detalhado'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{mes_ref}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = pt_title("Faturamento Detalhado")

hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, reg in enumerate(todos_registros, 2):
    for c_idx, (col_key, _) in enumerate(COLUNAS, 1):
        ws.cell(row=r_idx, column=c_idx, value=reg.get(col_key, '') or '')

# Larguras
for i, w in enumerate([12, 8, 12, 14, 25, 20, 30, 15, 12], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Salvar JSON ────────────────────────────────────────────────────────────────
json_path = f"{pasta}/{mes_ref}.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({
        "periodo": {"de": data_ini, "ate": data_fim},
        "total_registros": len(todos_registros),
        "faturamento_bruto": total_fmt,
        "registros": todos_registros
    }, f, ensure_ascii=False, indent=2)
print(f"JSON salvo: {json_path}")

# ── Enviar email ───────────────────────────────────────────────────────────────
mes_nome = month_year_pt(datetime(ano_ant, mes_ant, 1))

corpo = f"📊 Faturamento Detalhado — {mes_nome}\n"
corpo += "=" * 50 + "\n\n"
corpo += f"📅 Período: {data_ini} a {data_fim}\n"
corpo += f"📦 Total de registros: {len(todos_registros)}\n"
corpo += f"💰 Faturamento bruto: {total_fmt}\n\n"
corpo += "(BigDog 🐕)"

cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "📊 Mogo Faturamento Detalhado — {mes_nome} — {total_fmt}" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}" '
    f'--attach "{json_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
