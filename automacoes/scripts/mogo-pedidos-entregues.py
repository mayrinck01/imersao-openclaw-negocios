#!/usr/bin/env python3
"""
BigDog — Mogo "Pedidos Entregues" Report
Roda todo dia às 09:00 BRT.

Replica exatamente: PDV > Entregas > Filtros (Data Agendada = ontem) > Pesquisar > Finalizados > Excel

Endpoint: POST /Pedido/ListPedidosParaEntrega
  Query:  cFiltroTipoEntrega=2, dtDe=DD/MM/YYYY, dtAte=DD/MM/YYYY, tipoFiltroData=Entrega
  Form:   _search=true, rows=1000, page=1, sidx=HoraInclusao, sord=desc
"""

import sys, os, subprocess, json, re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from mogo_excel import order_columns_by_first_record, format_currency_cells
from filter_pt import month_year_pt, pt_columns, pt_title
sys.path.insert(0, os.path.dirname(__file__))
from mogo_login import mogo_login, MOGO_URL

op_token = open('/root/.openclaw/credentials/1password-token.txt').read().strip()
env = os.environ.copy()
env['OP_SERVICE_ACCOUNT_TOKEN'] = op_token

# Login com novo fluxo de 3 etapas
session = mogo_login()

ontem = datetime.now() - timedelta(days=1)
ontem_fmt = ontem.strftime('%d-%m-%Y')   # para nome de arquivo
ontem_api = ontem.strftime('%d/%m/%Y')   # para chamada API

print(f"Buscando pedidos entregues em {ontem_api}...")

# ── Buscar pedidos finalizados ────────────────────────────────────────────────
r = session.post(
    f"{MOGO_URL}/Pedido/ListPedidosParaEntrega",
    params={
        'cFiltroTipoEntrega': '2',
        'dtDe': ontem_api,
        'dtAte': ontem_api,
        'tipoFiltroData': 'Entrega'
    },
    data={
        '_search': 'true',
        'nd': '1',
        'rows': '1000',
        'page': '1',
        'sidx': 'HoraInclusao',
        'sord': 'desc',
        'totalrows': ''
    },
    timeout=30
)

if r.status_code != 200:
    print(f"ERRO ao buscar pedidos ({r.status_code})")
    sys.exit(1)

try:
    data = r.json()
    pedidos = data.get('rows', [])
except Exception as e:
    print(f"ERRO ao parsear JSON: {e}")
    sys.exit(1)

total_pedidos = len(pedidos)
print(f"Pedidos entregues encontrados: {total_pedidos}")

if not pedidos:
    print(f"Nenhum pedido entregue em {ontem_api}.")
    sys.exit(0)

# ── Ordenar por HoraInclusao ──────────────────────────────────────────────────
pedidos.sort(key=lambda p: p.get('HoraInclusao', ''))

# ── Helpers para limpar HTML dos campos ──────────────────────────────────────
def extrair_origem(html):
    """Extrai texto de origem do HTML: <div title=IFood ...> → IFood"""
    if not html:
        return ''
    m = re.search(r"title='?\"?([^'\">\s]+)'?\"?", str(html))
    return m.group(1) if m else re.sub(r'<[^>]+>', '', str(html)).strip()

def extrair_foody(html):
    """Extrai status Foody: <div title='Pedido enviado ao Foody'>Enviado</div> → Enviado"""
    if not html:
        return ''
    m = re.search(r'>([^<]+)</div>', str(html))
    return m.group(1).strip() if m else ''

# ── Calcular totais ───────────────────────────────────────────────────────────
total_valor = 0.0
for p in pedidos:
    try:
        val = str(p.get('ValorFinal', '0')).replace('.', '').replace(',', '.').strip()
        total_valor += float(val)
    except:
        pass

# ── Salvar Excel ──────────────────────────────────────────────────────────────
COLUNAS = [
    ('NumeroPedido',         'Nº Pedido'),
    ('_Origem',              'Origem'),
    ('DataEntrega',          'Data Agendada'),
    ('HoraEntregaTxt',       'Hora Agendada'),
    ('NomeCliente',          'Nome'),
    ('_FoodyDelivery',       'Foody Delivery'),
    ('ObsEntrega_Descricao', 'Forma Entrega'),
    ('ObservacaoEntrega',    'Observação'),
    ('Logradouro',           'Logradouro'),
    ('Bairro',               'Bairro'),
    ('Numero',               'Nº'),
    ('StatusPago',           'Pago'),
    ('HoraInclusao',         'Inclusão'),
    ('DataPedido',           'Data Pedido'),
    ('HoraSaidaEntregadorTxt', 'Hora Saída Entrega'),
]


COLUNAS = order_columns_by_first_record(pedidos[0] if pedidos else {}, COLUNAS)

pasta = '/root/workspaces/cake-brain/relatorios/Mogo/Pedidos Entregues'
os.makedirs(pasta, exist_ok=True)
xlsx_path = f"{pasta}/{ontem_fmt}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Finalizados"

hfill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=9)
for c, (_, header) in enumerate(COLUNAS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal='center')

for r_idx, pedido in enumerate(pedidos, 2):
    for c_idx, (col_name, _) in enumerate(COLUNAS, 1):
        if col_name == '_Origem':
            val = extrair_origem(pedido.get('OrigemPedido', ''))
        elif col_name == '_FoodyDelivery':
            val = extrair_foody(pedido.get('SincronizadoFoodyDelivery', ''))
        else:
            val = pedido.get(col_name, '') or ''
        ws.cell(row=r_idx, column=c_idx, value=val)

# Ajustar largura das colunas
larguras = [12, 12, 14, 14, 30, 14, 22, 20, 35, 18, 6, 8, 10, 14, 18]
for i, larg in enumerate(larguras, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = larg

format_currency_cells(wb)
wb.save(xlsx_path)
print(f"Excel salvo: {xlsx_path}")

# ── Montar email ──────────────────────────────────────────────────────────────
total_fmt = f"R$ {total_valor:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

corpo = f"✅ Pedidos Entregues — {ontem_api}\n"
corpo += "=" * 50 + "\n\n"
corpo += f"📦 Total: {total_pedidos} pedido(s)\n"
corpo += f"💰 Valor total: {total_fmt}\n\n"

# Por origem
origens = {}
for p in pedidos:
    o = extrair_origem(p.get('OrigemPedido','')) or 'Próprio'
    origens[o] = origens.get(o, 0) + 1
if origens:
    corpo += "📊 Por origem:\n"
    for origem, qtde in sorted(origens.items(), key=lambda x: -x[1]):
        corpo += f"  • {origem}: {qtde}\n"
    corpo += "\n"

corpo += "─" * 50 + "\n"
for p in pedidos:
    origem = extrair_origem(p.get('OrigemPedido',''))
    saida = p.get('HoraSaidaEntregadorTxt','') or ''
    corpo += (
        f"#{p.get('NumeroPedido','')} | {p.get('HoraEntregaTxt','')} | "
        f"{p.get('NomeCliente','')[:22]} | {p.get('Bairro','')} | "
        f"{origem} | Saída: {saida}\n"
    )

corpo += "\n(BigDog 🐕)"

# ── Enviar email ──────────────────────────────────────────────────────────────
cmd = [
    'bash', '-c',
    f'GOG_KEYRING_PASSWORD="" gog gmail send '
    f'--account cakebigdog@gmail.com '
    f'--client cakebigdog '
    f'--to joao@cakeco.com.br '
    f'--subject "✅ Mogo Entregues — {total_pedidos} pedido(s) — {total_fmt} — {ontem_api}" '
    f'--body "{corpo}" '
    f'--attach "{xlsx_path}"'
]
res = subprocess.run(cmd, capture_output=True, text=True, env=env)
if 'message_id' in res.stdout:
    print(f"✅ Email enviado para joao@cakeco.com.br")
else:
    print(f"ERRO email: {res.stderr[:200]}")
