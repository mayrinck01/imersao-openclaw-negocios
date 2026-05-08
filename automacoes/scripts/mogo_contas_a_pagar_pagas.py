#!/usr/bin/env python3
"""
BigDog — Mogo "Contas a Pagar Pagas" (Mensal)

Roda mensalmente no dia 6 às 04:08 BRT.
Relatório referente ao mês anterior.

Replica: Financeiro > Contas a Pagar
  - Tipo de data: Pagamento
  - Período: primeiro ao último dia do mês anterior
  - Contas pagas no período

Endpoints:
  - GET /Financeiro/BillsToPayJqGrid
  - GET /Financeiro/GetTotaisBillsToPay
"""

from __future__ import annotations

import argparse
import calendar
import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from filter_pt import month_year_pt  # noqa: E402
from mogo_excel import format_currency_cells  # noqa: E402
from mogo_login import MOGO_URL, mogo_login  # noqa: E402


ACCOUNT = "cakebigdog@gmail.com"
CLIENT = "cakebigdog"
DEFAULT_TO = "joao@cakeco.com.br"
OUTPUT_DIR = Path("/root/workspaces/cake-brain/relatorios/Mogo/Contas a Pagar Pagas")
PAGE_SIZE = 1000


@dataclass(frozen=True)
class MonthPeriod:
    month: int
    year: int
    start_br: str
    end_br: str
    file_prefix: str
    label_pt: str


def month_period(month: int | None = None, year: int | None = None, *, previous_month_from: date | None = None) -> MonthPeriod:
    if month is None or year is None:
        ref = previous_month_from or date.today()
        month = ref.month - 1
        year = ref.year
        if month == 0:
            month = 12
            year -= 1
    last_day = calendar.monthrange(year, month)[1]
    start_br = f"01/{month:02d}/{year:04d}"
    end_br = f"{last_day:02d}/{month:02d}/{year:04d}"
    file_prefix = f"{year:04d}-{month:02d}"
    label_pt = month_year_pt(datetime(year, month, 1))
    return MonthPeriod(month, year, start_br, end_br, file_prefix, label_pt)


def build_bills_to_pay_params(period: MonthPeriod, *, page: int = 1, rows: int = PAGE_SIZE) -> dict[str, str]:
    return {
        "_search": "true",
        "nd": "1",
        "rows": str(rows),
        "page": str(page),
        "sidx": "DataPagamento",
        "sord": "asc",
        "totalrows": "",
        "chkVencidos": "on",
        "chkReceber": "on",
        "chkRecebido": "on",
        "chkIgnEstoque": "on",
        "chkCaixaVista": "on",
        "chkRatiada": "on",
        "chkNaoContasCaixa": "on",
        "chkEntradaEstoqueXML": "on",
        "inpListIdContas": "",
        "inpPesqCrPdp": "",
        "inpPesqCrC": "",
        "inpPesqCrD": "",
        "inpPesqCrV": "",
        "inpPesqCrN": "",
        "dataDe": period.start_br,
        "dataAte": period.end_br,
        "selDate": "pagamento",
        "chkTituloComp": "on",
        "validaConciliacao": "-1",
    }


def strip_html(value: Any) -> str:
    if value is None:
        return ""
    text = re.sub(r"<[^>]+>", "", str(value))
    return re.sub(r"\s+", " ", text).strip()


def parse_br_date(value: Any) -> date | None:
    text = strip_html(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def parse_money(value: Any) -> float:
    text = strip_html(value)
    if not text:
        return 0.0
    text = text.replace("R$", "").replace(" ", "").strip()
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def fmt_brl(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _is_paid(row: dict[str, Any]) -> bool:
    status = strip_html(row.get("StatusBillsToPay")).lower()
    if status == "pago":
        return True
    if row.get("Pago") is True:
        return True
    return False


def normalize_paid_rows(rows: list[dict[str, Any]], period: MonthPeriod) -> list[dict[str, Any]]:
    start = datetime.strptime(period.start_br, "%d/%m/%Y").date()
    end = datetime.strptime(period.end_br, "%d/%m/%Y").date()
    normalized: list[dict[str, Any]] = []

    for row in rows:
        payment_date = parse_br_date(row.get("DataPagamento"))
        if not payment_date or not (start <= payment_date <= end):
            continue
        if not _is_paid(row):
            continue

        normalized.append({
            "id": row.get("Id"),
            "descricao": strip_html(row.get("Descricao")),
            "fornecedor": strip_html(row.get("Fornecedor_RazaoSocial") or row.get("Fornecedor")),
            "plano_contas": strip_html(row.get("PlanoDeContas_Descricao") or row.get("PlanoDeContas")),
            "centro_custo": strip_html(row.get("CentroDeCusto")),
            "conta_bancaria": strip_html(row.get("ContaBancaria_Descricao") or row.get("ContaBancaria")),
            "data_emissao": strip_html(row.get("DataEmissao")),
            "vencimento": strip_html(row.get("Vencimento")),
            "data_pagamento": payment_date.strftime("%d/%m/%Y"),
            "valor": parse_money(row.get("Valor")),
            "valor_pago": parse_money(row.get("ValorPago")),
            "saldo": parse_money(row.get("Saldo")),
            "desconto": parse_money(row.get("Desconto")),
            "juro": parse_money(row.get("Juro")),
            "status": strip_html(row.get("StatusBillsToPay")),
            "numero_nf": strip_html(row.get("NumeroNF")),
            "historico": strip_html(row.get("Historico")),
            "observacao": strip_html(row.get("Observacao")),
            "funcionario": strip_html(row.get("Funcionario")),
            "conciliado": strip_html(row.get("Conciliado")),
        })

    normalized.sort(key=lambda r: (r["data_pagamento"], r["fornecedor"], r["descricao"]))
    return normalized


def _group(rows: list[dict[str, Any]], key: str) -> list[tuple[str, int, float]]:
    grouped: dict[str, dict[str, float | int]] = defaultdict(lambda: {"count": 0, "total": 0.0})
    for row in rows:
        label = row.get(key) or "Sem classificação"
        grouped[str(label)]["count"] += 1
        grouped[str(label)]["total"] += float(row.get("valor_pago") or 0.0)
    return sorted(((k, int(v["count"]), float(v["total"])) for k, v in grouped.items()), key=lambda x: (-x[2], -x[1], x[0].lower()))


def build_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "count": len(rows),
        "total_valor": sum(float(r.get("valor") or 0.0) for r in rows),
        "total_pago": sum(float(r.get("valor_pago") or 0.0) for r in rows),
        "total_saldo": sum(float(r.get("saldo") or 0.0) for r in rows),
        "by_fornecedor": _group(rows, "fornecedor"),
        "by_conta_bancaria": _group(rows, "conta_bancaria"),
        "by_centro_custo": _group(rows, "centro_custo"),
        "by_plano_contas": _group(rows, "plano_contas"),
    }


def fetch_rows(session, period: MonthPeriod) -> list[dict[str, Any]]:
    all_rows: list[dict[str, Any]] = []
    page = 1
    records = None
    while True:
        params = build_bills_to_pay_params(period, page=page, rows=PAGE_SIZE)
        response = session.get(f"{MOGO_URL}/Financeiro/BillsToPayJqGrid", params=params, timeout=(20, 180))
        if response.status_code != 200:
            raise RuntimeError(f"Mogo BillsToPayJqGrid retornou HTTP {response.status_code}: {response.text[:200]}")
        payload = response.json()
        rows = payload.get("rows") or []
        records = int(payload.get("records") or records or 0)
        all_rows.extend(rows)
        print(f"  Página {page}: {len(rows)} registros (acumulado {len(all_rows)} / records {records})")
        if not rows or len(rows) < PAGE_SIZE or (records and len(all_rows) >= records):
            break
        page += 1
    return all_rows


def fetch_totals(session, period: MonthPeriod) -> dict[str, Any]:
    params = build_bills_to_pay_params(period, page=1, rows=PAGE_SIZE)
    # O endpoint de totais não precisa de paginação, mas respeita o mesmo filtro da tela.
    for key in ("_search", "nd", "rows", "page", "sidx", "sord", "totalrows"):
        params.pop(key, None)
    response = session.get(f"{MOGO_URL}/Financeiro/GetTotaisBillsToPay", params=params, timeout=(20, 120))
    if response.status_code != 200:
        raise RuntimeError(f"Mogo GetTotaisBillsToPay retornou HTTP {response.status_code}: {response.text[:200]}")
    return response.json()


COLUMNS = [
    ("data_pagamento", "Data Pagamento"),
    ("fornecedor", "Fornecedor"),
    ("descricao", "Descrição"),
    ("plano_contas", "Plano de Contas"),
    ("centro_custo", "Centro de Custo"),
    ("conta_bancaria", "Conta Bancária"),
    ("data_emissao", "Data Emissão"),
    ("vencimento", "Vencimento"),
    ("valor", "Valor"),
    ("valor_pago", "Valor Pago"),
    ("saldo", "Saldo"),
    ("desconto", "Desconto"),
    ("juro", "Juro"),
    ("status", "Status"),
    ("numero_nf", "Nº NF"),
    ("historico", "Histórico"),
    ("observacao", "Observação"),
    ("funcionario", "Funcionário"),
    ("conciliado", "Conciliado"),
    ("id", "ID Mogo"),
]


def write_outputs(rows: list[dict[str, Any]], summary: dict[str, Any], totals: dict[str, Any], period: MonthPeriod, output_dir: Path = OUTPUT_DIR) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / f"{period.file_prefix}-contas-a-pagar-pagas.xlsx"
    json_path = output_dir / f"{period.file_prefix}-contas-a-pagar-pagas.json"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagas"

    hfill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=9)
    for c_idx, (_, header) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, 2):
        for c_idx, (field, _) in enumerate(COLUMNS, 1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(field, ""))

    for c_idx, (field, header) in enumerate(COLUMNS, 1):
        max_len = len(header)
        for item in rows[:500]:
            max_len = max(max_len, len(str(item.get(field, ""))))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len + 2, 12), 45)

    money_cols = {"valor", "valor_pago", "saldo", "desconto", "juro"}
    for c_idx, (field, _) in enumerate(COLUMNS, 1):
        if field in money_cols:
            for cell in ws.iter_cols(min_col=c_idx, max_col=c_idx, min_row=2, max_row=ws.max_row):
                for c in cell:
                    c.number_format = 'R$ #,##0.00'

    # Aba resumo
    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Relatório", "Contas a Pagar — Pagas"])
    ws2.append(["Período", f"{period.start_br} a {period.end_br}"])
    ws2.append(["Registros", summary["count"]])
    ws2.append(["Valor", summary["total_valor"]])
    ws2.append(["Valor Pago", summary["total_pago"]])
    ws2.append(["Saldo", summary["total_saldo"]])
    ws2.append([])
    for title, key in [
        ("Por fornecedor", "by_fornecedor"),
        ("Por conta bancária", "by_conta_bancaria"),
        ("Por centro de custo", "by_centro_custo"),
        ("Por plano de contas", "by_plano_contas"),
    ]:
        ws2.append([title])
        ws2.append(["Nome", "Qtd", "Valor Pago"])
        for label, count, total in summary[key][:30]:
            ws2.append([label, count, total])
        ws2.append([])
    for row in ws2.iter_rows():
        for cell in row:
            if isinstance(cell.value, (float, int)) and cell.column in (2, 3):
                cell.number_format = 'R$ #,##0.00' if cell.column == 3 else '0'
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 18

    format_currency_cells(wb)
    wb.save(xlsx_path)

    json_path.write_text(json.dumps({
        "periodo": {"de": period.start_br, "ate": period.end_br, "tipo_data": "pagamento"},
        "summary": summary,
        "totais_mogo": totals,
        "registros": rows,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    return xlsx_path, json_path


def _top_lines(items: list[tuple[str, int, float]], limit: int = 8) -> str:
    if not items:
        return "  —\n"
    return "".join(f"  • {label}: {count} registro(s) | {fmt_brl(total)}\n" for label, count, total in items[:limit])


def build_email_body(period: MonthPeriod, summary: dict[str, Any], totals: dict[str, Any]) -> str:
    body = f"""Relatório Mogo — Contas a Pagar Pagas — {period.label_pt}

📅 Período: {period.start_br} → {period.end_br}
🗂️ Tipo de data: Pagamento
📦 Registros pagos: {summary['count']}

💰 Valor total: {fmt_brl(summary['total_valor'])}
✅ Valor pago:  {fmt_brl(summary['total_pago'])}
⏳ Saldo:       {fmt_brl(summary['total_saldo'])}

Totais retornados pelo Mogo:
  • Pagas: {fmt_brl(float(totals.get('pagas') or 0.0))}
  • Total pago: {fmt_brl(float(totals.get('totalPago') or 0.0))}

Top fornecedores:
{_top_lines(summary['by_fornecedor'])}
Top contas bancárias:
{_top_lines(summary['by_conta_bancaria'])}
Top centros de custo:
{_top_lines(summary['by_centro_custo'])}
(BigDog 🐕)
"""
    return body


def send_email(to: str, subject: str, body: str, attachments: list[Path]) -> None:
    body_file = Path("/tmp/mogo-contas-a-pagar-pagas-email.txt")
    body_file.write_text(body, encoding="utf-8")
    try:
        cmd = [
            "gog", "gmail", "send",
            "--account", ACCOUNT,
            "--client", CLIENT,
            "--to", to,
            "--subject", subject,
            "--body-file", str(body_file),
        ]
        for attachment in attachments:
            cmd += ["--attach", str(attachment)]
        env = os.environ.copy()
        env["GOG_KEYRING_PASSWORD"] = ""
        result = subprocess.run(cmd, capture_output=True, text=True, env=env, timeout=180)
        if result.returncode != 0:
            raise RuntimeError(result.stderr[:500] or result.stdout[:500])
        print("✅ Email enviado")
    finally:
        try:
            body_file.unlink()
        except FileNotFoundError:
            pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Relatório mensal Mogo — Contas a Pagar Pagas")
    parser.add_argument("mes", nargs="?", type=int, help="Mês de referência (1-12). Default: mês anterior")
    parser.add_argument("ano", nargs="?", type=int, help="Ano de referência. Default: mês anterior")
    parser.add_argument("--no-email", action="store_true", help="Gera arquivos sem enviar email")
    parser.add_argument("--to", default=DEFAULT_TO, help="Destinatário do email")
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR), help="Diretório de saída")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    period = month_period(month=args.mes, year=args.ano)
    print(f"Contas a Pagar Pagas: {period.start_br} → {period.end_br}")

    session = mogo_login()
    raw_rows = fetch_rows(session, period)
    rows = normalize_paid_rows(raw_rows, period)
    totals = fetch_totals(session, period)
    summary = build_summary(rows)

    print(f"Registros brutos: {len(raw_rows)}")
    print(f"Registros pagos validados: {summary['count']}")
    print(f"Valor pago: {fmt_brl(summary['total_pago'])}")

    if not rows:
        print("Nenhuma conta paga encontrada no período.")
        return 0

    xlsx_path, json_path = write_outputs(rows, summary, totals, period, Path(args.output_dir))
    print(f"Excel salvo: {xlsx_path}")
    print(f"JSON salvo: {json_path}")

    if not args.no_email:
        subject = f"💸 Mogo Contas a Pagar Pagas — {period.label_pt} — {summary['count']} registros — {fmt_brl(summary['total_pago'])}"
        send_email(args.to, subject, build_email_body(period, summary, totals), [xlsx_path, json_path])

    print("✅ Relatório concluído")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
