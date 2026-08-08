#!/usr/bin/env python3
"""Gera o relatório mensal Mogo 71 — Pedidos X Cancelamentos por Usuário."""

import datetime as dt
import json
import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


sys.path.insert(0, os.path.dirname(__file__))

from mogo_excel import order_columns_by_records


REPORT_CODE = 71
LOCAL_FOLDER_NAME = "Pedidos X Cancelamentos por Usuario"
DEFAULT_OUTPUT_DIR = Path("/root/workspaces/cake-brain/relatorios/Mogo") / LOCAL_FOLDER_NAME
COLUMNS = [
    ("A0", "Funcionário"),
    ("A1", "Quantidade de Pedidos"),
    ("A2", "Quantidade de Cancelamentos"),
]


def previous_month_period(today=None):
    today = today or dt.date.today()
    first_this_month = today.replace(day=1)
    last_previous_month = first_this_month - dt.timedelta(days=1)
    first_previous_month = last_previous_month.replace(day=1)
    return first_previous_month, last_previous_month


def build_filter(start_date, end_date):
    return f"DataDe{{{start_date:%d/%m/%Y}|DataAte{{{end_date:%d/%m/%Y}"


def _quantity(value):
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    return int(float(text))


def fetch_rows(session, mogo_url, start_date, end_date):
    grid_params = json.dumps(
        {
            "Searching": True,
            "RecordsCount": 9999,
            "PageIndex": 0,
            "SortingName": "",
            "SortingOrder": "ASC",
        }
    )
    response = session.get(
        f"{mogo_url}/relatorios/BuscaDadosRelatorioDinamico",
        params={
            "idGeradorRelatorios": "0",
            "codRelatorio": str(REPORT_CODE),
            "filtro": build_filter(start_date, end_date),
            "gridparamns": grid_params,
            "colunas": "[]",
            "dbNameFranquia": "",
        },
        timeout=60,
    )
    response.raise_for_status()
    rows = response.json().get("rows") or []
    if not rows:
        raise RuntimeError("Mogo não retornou dados para Pedidos X Cancelamentos por Usuário")
    return rows


def export_report(rows, start_date, end_date, output_dir=DEFAULT_OUTPUT_DIR):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    month_ref = start_date.strftime("%m-%Y")
    xlsx_path = output_dir / f"{month_ref}.xlsx"
    json_path = output_dir / f"{month_ref}.json"

    columns = order_columns_by_records(rows, COLUMNS)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Pedidos X Cancelamentos"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for column_index, (_key, header) in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, 2):
        for column_index, (key, _header) in enumerate(columns, 1):
            value = row.get(key, "")
            if key in {"A1", "A2"}:
                value = _quantity(value)
            sheet.cell(row=row_index, column=column_index, value=value)

    widths = {"A0": 28, "A1": 24, "A2": 30}
    for column_index, (key, _header) in enumerate(columns, 1):
        letter = openpyxl.utils.get_column_letter(column_index)
        sheet.column_dimensions[letter].width = widths.get(key, 18)

    workbook.save(xlsx_path)

    named_rows = [
        {
            "funcionario": row.get("A0", ""),
            "quantidade_pedidos": _quantity(row.get("A1")),
            "quantidade_cancelamentos": _quantity(row.get("A2")),
        }
        for row in rows
    ]
    payload = {
        "periodo": {"de": start_date.strftime("%d/%m/%Y"), "ate": end_date.strftime("%d/%m/%Y")},
        "total_funcionarios": len(named_rows),
        "total_pedidos": sum(row["quantidade_pedidos"] for row in named_rows),
        "total_cancelamentos": sum(row["quantidade_cancelamentos"] for row in named_rows),
        "registros": named_rows,
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {"xlsx": xlsx_path, "json": json_path, **{key: payload[key] for key in payload if key != "registros"}}


def main():
    from mogo_login import MOGO_URL, mogo_login

    start_date, end_date = previous_month_period()
    print(f"Pedidos X Cancelamentos por Usuário: {start_date:%d/%m/%Y} a {end_date:%d/%m/%Y}...")
    rows = fetch_rows(mogo_login(), MOGO_URL, start_date, end_date)
    result = export_report(rows, start_date, end_date)
    print(f"Registros: {result['total_funcionarios']}")
    print(f"Pedidos: {result['total_pedidos']}")
    print(f"Cancelamentos: {result['total_cancelamentos']}")
    print(f"Excel salvo: {result['xlsx']}")
    print(f"JSON salvo: {result['json']}")


if __name__ == "__main__":
    main()
